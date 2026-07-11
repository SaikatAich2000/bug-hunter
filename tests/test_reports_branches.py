"""Edge-case and branch tests for the Reports subsystem (engine, routes, xlsx, catalog).

app.* imports live inside test bodies because the client fixture re-imports app modules per test.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

import pytest

from openpyxl import load_workbook

from tests.conftest import BOOTSTRAP_EMAIL, BOOTSTRAP_PASSWORD


def _make_project(client, name="CovProj"):
    r = client.post("/api/projects", json={"name": name, "color": "#123456"})
    assert r.status_code == 201, r.text
    return r.json()


def _make_item(client, project_id, **extra):
    body = {
        "title": "cov item",
        "project_id": project_id,
        "priority": "High",
        "environment": "DEV",
        "item_type": "Bug",
    }
    body.update(extra)
    r = client.post("/api/bugs", json=body)
    assert r.status_code == 201, r.text
    return r.json()


def _change_status(client, bug_id, new_status):
    r = client.put(f"/api/bugs/{bug_id}", json={"status": new_status})
    assert r.status_code == 200, r.text
    return r.json()


def _run(client, report_key, filters=None):
    r = client.post("/api/reports/run", json={
        "report_key": report_key,
        "filters": filters or {},
    })
    assert r.status_code == 200, r.text
    return r.json()


def test_cov_parse_resolution_status_edge_cases():
    from app.reports.engine import _parse_resolution_status
    assert _parse_resolution_status("") is None
    assert _parse_resolution_status(None) is None
    assert _parse_resolution_status("created the bug") is None
    assert _parse_resolution_status("assigned to Bob") is None
    assert _parse_resolution_status(
        "#5 'X' — status: 'New' → 'Resolved'") == "Resolved"


def test_cov_run_report_unknown_key_raises():
    from app.reports.engine import Filters, UnknownReportError, run_report
    with pytest.raises(UnknownReportError):
        run_report("definitely_not_a_report", Filters(), db=None)


def test_cov_run_route_maps_unknown_report_error_to_400(admin_client, monkeypatch):
    """UnknownReportError maps to 400; unreachable via HTTP, so run_report is patched to force it."""
    import app.routes.reports as reports_route
    from app.reports.engine import UnknownReportError

    def _boom(_key, _filters, _db):
        raise UnknownReportError("forced unknown")

    monkeypatch.setattr(reports_route, "run_report", _boom)
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",   # valid, so it passes the catalog check
        "filters": {},
    })
    assert r.status_code == 400
    assert "forced unknown" in r.json()["detail"]


def test_cov_catalog_get_report_meta():
    from app.reports.catalog import get_report_meta
    meta = get_report_meta("item_detail")            # hit
    assert meta is not None and meta["key"] == "item_detail"
    assert get_report_meta("nope_not_real") is None  # miss


def test_cov_xlsx_coerce_branches():
    from app.reports.xlsx import _coerce, _defang_formula_text
    assert _coerce(None) == ""
    assert _coerce(True) is True                      # bool kept as-is
    assert _coerce(7) == 7
    assert _coerce(3.5) == 3.5
    dt = datetime(2026, 1, 2, 3, 4, 5, tzinfo=timezone.utc)
    assert _coerce(dt) == dt.isoformat(timespec="seconds")
    assert _coerce(["a", "b"]) == "['a', 'b']"
    # Formula-leading characters are prefixed with a quote to prevent injection.
    assert _defang_formula_text("=cmd|calc") == "'=cmd|calc"
    assert _coerce("=danger") == "'=danger"


def test_cov_xlsx_build_workbook_with_empty_summary():
    """An empty summary dict (never produced by real reports) hits the early return in _write_summary_block."""
    from app.reports.engine import ReportColumn, ReportResult
    from app.reports.xlsx import build_workbook_bytes
    result = ReportResult(
        report_key="custom",
        report_label="Custom Empty Report",   # avoid the word "Summary" here
        columns=[ReportColumn("id", "ID", 8, kind="number"),
                 ReportColumn("title", "Title", 30)],
        rows=[{"id": 1, "title": "one"}, {"id": 2, "title": "two"}],
        summary={},                       # empty -> early return
        filters={},
    )
    data = build_workbook_bytes(result)
    wb = load_workbook(io.BytesIO(data), read_only=True)
    assert "Filters Applied" in wb.sheetnames
    # The summary block returned early, so no "Summary" header row should exist.
    fws = wb["Filters Applied"]
    dump = " ".join(
        str(v) for row in fws.iter_rows(values_only=True)
        for v in row if v is not None
    )
    assert "Summary" not in dump


def test_cov_run_statuses_filter_validator(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="new-one")  # status defaults to New
    # The validator strips unknown statuses; "New" survives, "TotallyBogusStatus" is dropped.
    body = _run(admin_client, "item_detail",
                {"statuses": ["New", "TotallyBogusStatus"]})
    titles = {row["title"] for row in body["rows"]}
    assert titles == {"new-one"}
    assert body["rows"][0]["status"] == "New"


def test_cov_item_detail_one_sided_date_windows(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="dated")
    today = datetime.now(timezone.utc).date().isoformat()
    b_from = _run(admin_client, "item_detail", {"date_from": today})
    assert any(r["title"] == "dated" for r in b_from["rows"])
    b_to = _run(admin_client, "item_detail", {"date_to": today})
    assert any(r["title"] == "dated" for r in b_to["rows"])
    # A date_to in the far past should exclude today's row.
    b_old = _run(admin_client, "item_detail", {"date_to": "2000-01-01"})
    assert b_old["total"] == 0


def test_cov_twice_resolved_bug_resolution_paths(admin_client):
    p = _make_project(admin_client)
    b = _make_item(admin_client, p["id"], title="flap")
    # New -> Resolved -> Reopened -> Resolved (two resolution events).
    _change_status(admin_client, b["id"], "Resolved")
    _change_status(admin_client, b["id"], "Reopened")
    _change_status(admin_client, b["id"], "Resolved")
    # Second bug never resolves — exercises the row-is-None path in _ttr_row.
    wip = _make_item(admin_client, p["id"], title="wip-only")
    _change_status(admin_client, wip["id"], "In Progress")

    # Newest resolution wins; earlier rows hit the already-seen continue.
    detail = _run(admin_client, "item_detail", {})
    row = next(r for r in detail["rows"] if r["title"] == "flap")
    assert row["resolved_by"] == "Test Admin"
    assert row["resolved_at"]  # non-empty

    # TTR dedupes by bug, so the twice-resolved bug yields exactly one row.
    ttr = _run(admin_client, "time_to_resolution", {})
    flap_rows = [r for r in ttr["rows"] if r["title"] == "flap"]
    assert len(flap_rows) == 1
    assert ttr["summary"]["count"] == 1

    # Timeline does NOT dedupe, so both Resolved transitions are tallied.
    tl = _run(admin_client, "timeline", {})
    assert tl["summary"]["total_resolved"] == 2  # both Resolved transitions
    assert tl["summary"]["total_created"] == 2   # flap + wip-only filed today


def test_cov_pending_snapshot_status_filter_and_assignee(admin_client):
    me = admin_client.get("/api/auth/me").json()
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="open-assigned",
               assignee_ids=[me["id"]])

    # "New" intersects the open set, so the item appears.
    body = _run(admin_client, "pending_snapshot", {"statuses": ["New"]})
    titles = {r["title"] for r in body["rows"]}
    assert "open-assigned" in titles
    assert body["summary"]["by_assignee"].get("Test Admin", 0) >= 1

    # "Closed" never intersects the open set → early-return empty ReportResult.
    empty = _run(admin_client, "pending_snapshot", {"statuses": ["Closed"]})
    assert empty["total"] == 0
    assert empty["summary"]["total_items"] == 0


def test_cov_throughput_all_query_filters(admin_client):
    me = admin_client.get("/api/auth/me").json()
    p = _make_project(admin_client, name="ThruProj")
    ev = admin_client.post("/api/events", json={"name": "ThruEvent"}).json()
    target = _make_item(
        admin_client, p["id"], title="thru-target",
        priority="Critical", environment="PROD",
        assignee_ids=[me["id"]], event_id=ev["id"],
    )
    _change_status(admin_client, target["id"], "Resolved")
    body = _run(admin_client, "throughput", {
        "project_ids": [p["id"]],
        "event_id": ev["id"],
        "assignee_ids": [me["id"]],
        "reporter_ids": [me["id"]],
        "priorities": ["Critical"],
        "environments": ["PROD"],
    })
    assert body["summary"]["total_resolved"] == 1
    admin_row = next(r for r in body["rows"] if r["user_name"] == "Test Admin")
    assert admin_row["resolved"] == 1
    assert admin_row["bugs"] == 1


def test_cov_project_breakdown_resolved_and_final(admin_client):
    p = _make_project(admin_client, name="PBProj")
    _make_item(admin_client, p["id"], title="pb-open")       # stays New
    done = _make_item(admin_client, p["id"], title="pb-done")
    _change_status(admin_client, done["id"], "Resolved")     # hits resolved + final counters
    body = _run(admin_client, "project_breakdown", {})
    row = next(r for r in body["rows"] if r["project"] == "PBProj")
    assert row["created"] == 2
    assert row["open"] == 1        # New item
    assert row["resolved"] == 1    # Resolved item
    assert row["final"] == 1       # Resolved is a final status


def test_cov_aging_status_filter_intersection_and_empty(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="aging-open")     # New (open)
    closed = _make_item(admin_client, p["id"], title="aging-closed")
    _change_status(admin_client, closed["id"], "Closed")

    # "New" intersects the open set; only the New item qualifies.
    inter = _run(admin_client, "aging", {"statuses": ["New"]})
    titles = {r["title"] for r in inter["rows"]}
    assert titles == {"aging-open"}

    # No overlap with the open set → empty, never closed items with fake ages.
    empty_set = _run(admin_client, "aging", {"statuses": ["Closed"]})
    assert empty_set["rows"] == []


def test_cov_export_xlsx_blank_label_and_zebra(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="zebra-one")
    _make_item(admin_client, p["id"], title="zebra-two")  # second row triggers zebra fill
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail",
        # Label sanitizes to empty, so _safe_filename appends no suffix.
        "filters": {"label": "@@@!!!"},
    })
    assert r.status_code == 200
    cd = r.headers.get("content-disposition", "")
    assert "report-item_detail-" in cd
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    main_ws = wb[wb.sheetnames[0]]
    rows = list(main_ws.iter_rows(values_only=True))
    assert len(rows) >= 4  # banner + header + 2 data rows


def test_cov_export_xlsx_over_row_limit_returns_413(admin_client, monkeypatch):
    import app.config as config
    p = _make_project(admin_client)
    for i in range(3):
        _make_item(admin_client, p["id"], title=f"cap-{i}")
    # Patch the class attribute: module-level refs go stale under per-test reimport.
    monkeypatch.setattr(config.Settings, "MAX_REPORT_ROWS", 1)
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail",
        "filters": {},
    })
    assert r.status_code == 413, r.text
    assert "row" in r.json()["detail"].lower()


# Bulk-inserts via the app's DB session — 1001 HTTP POSTs would be too slow.
def test_cov_run_inline_render_cap_truncates(admin_client):
    p = _make_project(admin_client, name="BulkProj")
    from app.database import SessionLocal
    from app.models import Bug
    # 1001 items: engine returns all, route caps the rendered payload at 1000.
    db = SessionLocal()
    try:
        db.add_all([
            Bug(project_id=p["id"], title=f"bulk-{i}", description="",
                item_type="Bug", status="New", priority="Low",
                environment="DEV")
            for i in range(1001)
        ])
        db.commit()
    finally:
        db.close()
    body = _run(admin_client, "item_detail", {})
    assert body["total"] == 1001               # full count from to_api()
    assert len(body["rows"]) == 1000           # rows capped at the render cap
    assert body["truncated"] is True
    assert body["truncated_cap"] == 1000


def test_cov_export_xlsx_build_error_returns_500(admin_client, monkeypatch):
    import app.reports.xlsx as xlsx
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="boom")
    # OPENPYXL_AVAILABLE is read at call time, so patching it forces XlsxBuildError -> 500.
    monkeypatch.setattr(xlsx, "OPENPYXL_AVAILABLE", False)
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail",
        "filters": {},
    })
    assert r.status_code == 500
    # The error is logged server-side; the response body is intentionally generic.
    detail = r.json()["detail"].lower()
    assert "openpyxl" not in detail
    assert "workbook" in detail
