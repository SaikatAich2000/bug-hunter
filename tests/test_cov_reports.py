"""Coverage-focused tests for the Reports subsystem.

This file targets the specific uncovered branches/lines in:
  * app/reports/engine.py
  * app/routes/reports.py
  * app/reports/xlsx.py
  * app/reports/catalog.py

that the broader behaviour suite (tests/test_reports.py,
tests/test_reports_engine_helpers.py) leaves untouched. Every test is
hermetic — it talks to the temp SQLite DB the `client`/`admin_client`
fixtures spin up, and the few cases that need a non-DB unit are called
directly against the engine/xlsx helpers (exactly as
test_reports_engine_helpers.py already does).

Why some app.* modules are imported INSIDE the test bodies: the `client`
fixture deletes & re-imports every ``app.*`` module per test, so a module
captured at this file's top level goes stale. Importing inside the test —
and, for config, patching the ``config.Settings`` CLASS attribute — pins
the exact generation the running app reads (see tests/test_password_policy.py
for the same pattern).
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

import pytest

from openpyxl import load_workbook

from tests.conftest import BOOTSTRAP_EMAIL, BOOTSTRAP_PASSWORD


# ---------------------------------------------------------------------------
# Small API helpers (mirror tests/test_reports.py so behaviour stays aligned)
# ---------------------------------------------------------------------------
def _make_project(client, name="CovProj"):
    r = client.post("/api/projects", json={"name": name, "color": "#123456"})
    assert r.status_code == 201, r.text
    return r.json()


def _make_item(client, project_id, **extra):
    body = {
        "title": "cov item",
        "project_id": project_id,
        "priority": "High",
        "environment": "DEV",
        "item_type": "Bug",
    }
    body.update(extra)
    r = client.post("/api/bugs", json=body)
    assert r.status_code == 201, r.text
    return r.json()


def _change_status(client, bug_id, new_status):
    r = client.put(f"/api/bugs/{bug_id}", json={"status": new_status})
    assert r.status_code == 200, r.text
    return r.json()


def _run(client, report_key, filters=None):
    r = client.post("/api/reports/run", json={
        "report_key": report_key,
        "filters": filters or {},
    })
    assert r.status_code == 200, r.text
    return r.json()


# ---------------------------------------------------------------------------
# engine: _parse_resolution_status empty + non-matching detail (lines 458, 461)
# run_report unknown-key raise (line 1163)  — direct unit calls
# ---------------------------------------------------------------------------
def test_cov_parse_resolution_status_edge_cases():
    from app.reports.engine import _parse_resolution_status
    # line 458: falsy detail -> None
    assert _parse_resolution_status("") is None
    assert _parse_resolution_status(None) is None
    # line 461: present but non-matching detail -> None
    assert _parse_resolution_status("created the bug") is None
    assert _parse_resolution_status("assigned to Bob") is None
    # sanity: a real status line still parses
    assert _parse_resolution_status(
        "#5 'X' — status: 'New' → 'Resolved'") == "Resolved"


def test_cov_run_report_unknown_key_raises():
    from app.reports.engine import Filters, UnknownReportError, run_report
    with pytest.raises(UnknownReportError):
        run_report("definitely_not_a_report", Filters(), db=None)  # line 1163


def test_cov_run_route_maps_unknown_report_error_to_400(admin_client, monkeypatch):
    """routes._run_or_400 lines 139-140: the catalog gate passes for a valid
    key, but if run_report itself raises UnknownReportError the route turns it
    into a 400. Unreachable through pure HTTP (catalog and _DISPATCH are kept
    in sync), so we force run_report — the very name the route calls — to raise.
    """
    import app.routes.reports as reports_route
    from app.reports.engine import UnknownReportError

    def _boom(_key, _filters, _db):
        raise UnknownReportError("forced unknown")

    monkeypatch.setattr(reports_route, "run_report", _boom)
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",   # valid -> passes the catalog check
        "filters": {},
    })
    assert r.status_code == 400
    assert "forced unknown" in r.json()["detail"]


# ---------------------------------------------------------------------------
# catalog: get_report_meta hit + miss (line 124)
# ---------------------------------------------------------------------------
def test_cov_catalog_get_report_meta():
    from app.reports.catalog import get_report_meta
    meta = get_report_meta("item_detail")          # line 124 (hit)
    assert meta is not None and meta["key"] == "item_detail"
    assert get_report_meta("nope_not_real") is None  # line 124 (miss path)


# ---------------------------------------------------------------------------
# xlsx: _coerce branches (lines 80, 83, 88-90) + defang positive (62)
# _write_summary_block empty-summary early return (line 197) via full pipeline
# ---------------------------------------------------------------------------
def test_cov_xlsx_coerce_branches():
    from app.reports.xlsx import _coerce, _defang_formula_text
    assert _coerce(None) == ""                       # line 80
    assert _coerce(True) is True                      # line 83 (bool kept as-is)
    assert _coerce(7) == 7
    assert _coerce(3.5) == 3.5
    dt = datetime(2026, 1, 2, 3, 4, 5, tzinfo=timezone.utc)
    assert _coerce(dt) == dt.isoformat(timespec="seconds")  # lines 88-89
    # line 90: non-str / non-(int,float) / non-datetime -> str()+defang
    assert _coerce(["a", "b"]) == "['a', 'b']"
    # defang positive branch (line 62) — formula-leading text gets quoted
    assert _defang_formula_text("=cmd|calc") == "'=cmd|calc"
    assert _coerce("=danger") == "'=danger"


def test_cov_xlsx_build_workbook_with_empty_summary():
    """A ReportResult whose summary is empty exercises the early-return in
    _write_summary_block (line 197). No report produces an empty summary in
    practice, so build the result by hand and run the real writer."""
    from app.reports.engine import ReportColumn, ReportResult
    from app.reports.xlsx import build_workbook_bytes
    result = ReportResult(
        report_key="custom",
        report_label="Custom Empty Report",   # avoid the word "Summary" here
        columns=[ReportColumn("id", "ID", 8, kind="number"),
                 ReportColumn("title", "Title", 30)],
        rows=[{"id": 1, "title": "one"}, {"id": 2, "title": "two"}],
        summary={},                       # <- triggers line 197
        filters={},
    )
    data = build_workbook_bytes(result)
    wb = load_workbook(io.BytesIO(data), read_only=True)
    assert "Filters Applied" in wb.sheetnames
    # No "Summary" header row should appear because the block returned early.
    fws = wb["Filters Applied"]
    dump = " ".join(
        str(v) for row in fws.iter_rows(values_only=True)
        for v in row if v is not None
    )
    assert "Summary" not in dump


# ---------------------------------------------------------------------------
# routes: statuses validator (line 84) via /run
# ---------------------------------------------------------------------------
def test_cov_run_statuses_filter_validator(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="new-one")  # status defaults to New
    # Mix a valid status with a bogus one — the validator (line 84) keeps
    # only the allowed one; apply_status branch (engine line 326) then fires.
    body = _run(admin_client, "item_detail",
                {"statuses": ["New", "TotallyBogusStatus"]})
    titles = {row["title"] for row in body["rows"]}
    assert titles == {"new-one"}
    assert body["rows"][0]["status"] == "New"


# ---------------------------------------------------------------------------
# engine: _apply_date_range one-sided windows (339->341 and 341->343)
# item_detail routes through _apply_date_range (throughput uses its own).
# ---------------------------------------------------------------------------
def test_cov_item_detail_one_sided_date_windows(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="dated")
    today = datetime.now(timezone.utc).date().isoformat()
    # date_from only -> 341->343 (date_to branch skipped).
    b_from = _run(admin_client, "item_detail", {"date_from": today})
    assert any(r["title"] == "dated" for r in b_from["rows"])
    # date_to only -> 339->341 (date_from branch skipped).
    b_to = _run(admin_client, "item_detail", {"date_to": today})
    assert any(r["title"] == "dated" for r in b_to["rows"])
    # A date_to in the far past excludes today's row (proves the bound binds).
    b_old = _run(admin_client, "item_detail", {"date_to": "2000-01-01"})
    assert b_old["total"] == 0


# ---------------------------------------------------------------------------
# engine: resolution-info dedup (lines 502, 505-506) via item_detail, and
# time_to_resolution seen-bug skip + row-None skip (lines 1096, 1099), and
# timeline non-resolution branch (997->994) — all from one twice-resolved bug.
# ---------------------------------------------------------------------------
def test_cov_twice_resolved_bug_resolution_paths(admin_client):
    p = _make_project(admin_client)
    b = _make_item(admin_client, p["id"], title="flap")
    # New -> Resolved (resolution #1) -> Reopened (non-res) -> Resolved (#2).
    _change_status(admin_client, b["id"], "Resolved")
    _change_status(admin_client, b["id"], "Reopened")
    _change_status(admin_client, b["id"], "Resolved")
    # A second bug that only ever goes to a NON-resolved state: its single
    # status_changed row reaches _ttr_row for an as-yet-unseen bug, which
    # returns None -> the `row is None` skip fires (engine line 1099).
    wip = _make_item(admin_client, p["id"], title="wip-only")
    _change_status(admin_client, wip["id"], "In Progress")

    # item_detail -> _fetch_resolution_info sees 3 status_changed rows for the
    # same bug: newest resolution captured (lines 505-506), the rest hit the
    # already-seen continue (line 502). Row must still surface with a resolver.
    detail = _run(admin_client, "item_detail", {})
    row = next(r for r in detail["rows"] if r["title"] == "flap")
    assert row["resolved_by"] == "Test Admin"
    assert row["resolved_at"]  # non-empty

    # time_to_resolution -> for that bug the query yields 3 rows ordered by
    # created_at asc: first Resolved builds a row & marks seen; Reopened ->
    # _ttr_row None (line 1099); second Resolved -> bug_id already seen
    # (line 1096). Exactly one TTR row for the bug.
    ttr = _run(admin_client, "time_to_resolution", {})
    flap_rows = [r for r in ttr["rows"] if r["title"] == "flap"]
    assert len(flap_rows) == 1
    assert ttr["summary"]["count"] == 1

    # timeline -> the Reopened status change is a non-resolution row, so the
    # resolved-bucketing condition is False for it (branch 997->994). The two
    # Resolved transitions DO count (timeline tallies every resolution event,
    # it does not dedup by bug), so total_resolved == 2.
    tl = _run(admin_client, "timeline", {})
    assert tl["summary"]["total_resolved"] == 2  # both Resolved transitions
    assert tl["summary"]["total_created"] == 2   # flap + wip-only filed today


# ---------------------------------------------------------------------------
# engine: pending_snapshot statuses-intersection + empty early-return
# (lines 595, 598) and by_assignee accumulation (line 624).
# ---------------------------------------------------------------------------
def test_cov_pending_snapshot_status_filter_and_assignee(admin_client):
    me = admin_client.get("/api/auth/me").json()
    p = _make_project(admin_client)
    # An open item assigned to admin so by_assignee (line 624) accumulates.
    _make_item(admin_client, p["id"], title="open-assigned",
               assignee_ids=[me["id"]])

    # statuses filter that intersects the open set (line 595) -> "New" kept.
    body = _run(admin_client, "pending_snapshot", {"statuses": ["New"]})
    titles = {r["title"] for r in body["rows"]}
    assert "open-assigned" in titles
    assert body["summary"]["by_assignee"].get("Test Admin", 0) >= 1

    # statuses filter that does NOT intersect the open set -> empty open_set,
    # early-return ReportResult (line 598). "Closed" is never an open status.
    empty = _run(admin_client, "pending_snapshot", {"statuses": ["Closed"]})
    assert empty["total"] == 0
    assert empty["summary"]["total_items"] == 0


# ---------------------------------------------------------------------------
# engine: throughput query filter branches
# (lines 670, 672, 674, 678, 680, 682) — every bug-side filter at once.
# ---------------------------------------------------------------------------
def test_cov_throughput_all_query_filters(admin_client):
    me = admin_client.get("/api/auth/me").json()
    p = _make_project(admin_client, name="ThruProj")
    ev = admin_client.post("/api/events", json={"name": "ThruEvent"}).json()
    target = _make_item(
        admin_client, p["id"], title="thru-target",
        priority="Critical", environment="PROD",
        assignee_ids=[me["id"]], event_id=ev["id"],
    )
    _change_status(admin_client, target["id"], "Resolved")
    body = _run(admin_client, "throughput", {
        "project_ids": [p["id"]],     # line 670
        "event_id": ev["id"],         # line 672
        "assignee_ids": [me["id"]],   # line 674
        "reporter_ids": [me["id"]],   # line 678
        "priorities": ["Critical"],   # line 680
        "environments": ["PROD"],     # line 682
    })
    assert body["summary"]["total_resolved"] == 1
    admin_row = next(r for r in body["rows"] if r["user_name"] == "Test Admin")
    assert admin_row["resolved"] == 1
    assert admin_row["bugs"] == 1


# ---------------------------------------------------------------------------
# engine: project_breakdown open=False branch + resolved + final increments
# (branch 866->868, lines 869, 871) — needs a resolved Bug in the set.
# ---------------------------------------------------------------------------
def test_cov_project_breakdown_resolved_and_final(admin_client):
    p = _make_project(admin_client, name="PBProj")
    _make_item(admin_client, p["id"], title="pb-open")       # stays New (open)
    done = _make_item(admin_client, p["id"], title="pb-done")
    _change_status(admin_client, done["id"], "Resolved")     # resolved + final
    body = _run(admin_client, "project_breakdown", {})
    row = next(r for r in body["rows"] if r["project"] == "PBProj")
    assert row["created"] == 2
    assert row["open"] == 1        # the New item
    assert row["resolved"] == 1    # the Resolved item (line 869)
    assert row["final"] == 1       # Resolved is a final status (line 871)


# ---------------------------------------------------------------------------
# engine: aging statuses-intersection (line 918) and empty open_set ->
# skip the status WHERE (branch 920->922).
# ---------------------------------------------------------------------------
def test_cov_aging_status_filter_intersection_and_empty(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="aging-open")     # New (open)
    closed = _make_item(admin_client, p["id"], title="aging-closed")
    _change_status(admin_client, closed["id"], "Closed")

    # statuses intersects the open set (line 918): only the New item qualifies.
    inter = _run(admin_client, "aging", {"statuses": ["New"]})
    titles = {r["title"] for r in inter["rows"]}
    assert titles == {"aging-open"}

    # statuses with NO overlap with the open set -> an "open items" report has
    # nothing to show (mirrors pending_snapshot), rather than silently leaking
    # closed items stamped with a bogus age.
    empty_set = _run(admin_client, "aging", {"statuses": ["Closed"]})
    assert empty_set["rows"] == []


# ---------------------------------------------------------------------------
# routes: export.xlsx label that sanitises to empty (branch 173->175),
# and the zebra-fill path with >=2 main-sheet rows (xlsx line 139).
# ---------------------------------------------------------------------------
def test_cov_export_xlsx_blank_label_and_zebra(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="zebra-one")
    _make_item(admin_client, p["id"], title="zebra-two")  # 2nd row -> zebra (139)
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail",
        # label is all formula/non-safe chars -> sanitises to "" -> the
        # `if suffix:` guard in _safe_filename is False (branch 173->175).
        "filters": {"label": "@@@!!!"},
    })
    assert r.status_code == 200
    cd = r.headers.get("content-disposition", "")
    # base stays just the report key (no label suffix appended).
    assert "report-item_detail-" in cd
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    main_ws = wb[wb.sheetnames[0]]
    rows = list(main_ws.iter_rows(values_only=True))
    # banner + header + 2 data rows.
    assert len(rows) >= 4


# ---------------------------------------------------------------------------
# routes: 413 when the row count exceeds MAX_REPORT_ROWS (line 199).
# Seed a handful of rows, then patch the CLASS attribute down to 1.
# ---------------------------------------------------------------------------
def test_cov_export_xlsx_over_row_limit_returns_413(admin_client, monkeypatch):
    import app.config as config
    p = _make_project(admin_client)
    for i in range(3):
        _make_item(admin_client, p["id"], title=f"cap-{i}")
    # Patch the Settings CLASS attribute so the freshly-built get_settings()
    # instance the route reads sees the tiny cap (module-level refs are stale
    # under the per-test reimport; see test_password_policy.py).
    monkeypatch.setattr(config.Settings, "MAX_REPORT_ROWS", 1)
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail",
        "filters": {},
    })
    assert r.status_code == 413, r.text
    assert "row" in r.json()["detail"].lower()


# ---------------------------------------------------------------------------
# routes: inline /run truncation when rows exceed the 1000 render cap
# (lines 158-160). Bulk-insert via the app's own session so the test stays
# fast (1001 HTTP POSTs would be needlessly slow) and hermetic. MAX_REPORT_ROWS
# keeps its default (50000) so the engine LIMIT doesn't trim below 1000.
# ---------------------------------------------------------------------------
def test_cov_run_inline_render_cap_truncates(admin_client):
    p = _make_project(admin_client, name="BulkProj")
    from app.database import SessionLocal
    from app.models import Bug
    # 1001 items -> engine returns them all (limit = MAX_REPORT_ROWS+1), the
    # route caps the rendered payload at 1000 and flags it truncated.
    db = SessionLocal()
    try:
        db.add_all([
            Bug(project_id=p["id"], title=f"bulk-{i}", description="",
                item_type="Bug", status="New", priority="Low",
                environment="DEV")
            for i in range(1001)
        ])
        db.commit()
    finally:
        db.close()
    body = _run(admin_client, "item_detail", {})
    # `total` reflects the full result (computed before the route trims), the
    # rendered `rows` list is capped at 1000 and flagged truncated.
    assert body["total"] == 1001               # full count from to_api()
    assert len(body["rows"]) == 1000           # line 158: rows[:rendered_cap]
    assert body["truncated"] is True           # line 159
    assert body["truncated_cap"] == 1000       # line 160


# ---------------------------------------------------------------------------
# routes: XlsxBuildError -> 500 (lines 209-211) + xlsx _ensure_openpyxl raise
# (line 68). Simulate a server with openpyxl unavailable by flipping the
# module flag the live build path reads.
# ---------------------------------------------------------------------------
def test_cov_export_xlsx_build_error_returns_500(admin_client, monkeypatch):
    import app.reports.xlsx as xlsx
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="boom")
    # build_workbook_bytes -> _ensure_openpyxl() reads OPENPYXL_AVAILABLE live;
    # forcing it False raises XlsxBuildError (line 68), which the route maps to
    # a 500 (lines 209-211).
    monkeypatch.setattr(xlsx, "OPENPYXL_AVAILABLE", False)
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail",
        "filters": {},
    })
    assert r.status_code == 500
    # The 500 body is GENERIC — server config/dependency state (the "openpyxl
    # not installed" cause) is logged server-side, not echoed to the client.
    detail = r.json()["detail"].lower()
    assert "openpyxl" not in detail
    assert "workbook" in detail
