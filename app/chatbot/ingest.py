"""Sleuth document ingest — the admin "upload a doc full of bugs" feature.

Admin-only (enforced in router.py). Two extraction paths tried in order: the
cloud AI when configured, else a deterministic parser (xlsx/JSON/CSV/free text)
that needs no network. Both yield the same cleaned `spec` shape, so validation,
creation, and audit are identical.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from itertools import islice
from typing import Any, Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Activity, Bug, Project, User
from app.schemas import (
    ALLOWED_ENVIRONMENTS, ALLOWED_ITEM_TYPES, ALLOWED_PRIORITIES,
    MIN_TITLE_LENGTH, sanitize_html,
)

logger = logging.getLogger("bug_hunter.sleuth.ingest")

# Hard cap on items per upload — bounds the DB against a runaway spreadsheet.
MAX_SPECS = 500
# Soft cap on uploaded document size; the route also enforces the body limit.
MAX_DOC_BYTES = 5 * 1024 * 1024  # 5 MB of text/spreadsheet is plenty


# Synonym maps so the parser accepts various spellings of priority/type/env.
_PRIORITY_SYNONYMS = {
    "critical": "Critical", "blocker": "Critical", "p0": "Critical", "urgent": "Critical",
    "high": "High", "p1": "High", "major": "High",
    "medium": "Medium", "p2": "Medium", "normal": "Medium", "med": "Medium",
    "low": "Low", "p3": "Low", "minor": "Low", "trivial": "Low",
}
_TYPE_SYNONYMS = {
    "bug": "Bug", "defect": "Bug", "issue": "Bug",
    "requirement": "Requirement", "story": "Requirement", "feature": "Requirement",
    "task": "Task", "todo": "Task", "chore": "Task",
}
_ENV_SYNONYMS = {
    "prod": "PROD", "production": "PROD", "live": "PROD",
    "uat": "UAT", "staging": "UAT", "stage": "UAT", "qa": "UAT", "test": "UAT",
    "dev": "DEV", "development": "DEV", "local": "DEV",
}

# Field-name aliases for structured rows (CSV header / JSON keys / xlsx header).
_TITLE_KEYS = ("title", "summary", "name", "bug", "issue", "task", "item")
_DESC_KEYS = ("description", "desc", "details", "detail", "notes", "note", "body")
_PRIORITY_KEYS = ("priority", "prio", "severity", "sev")
_TYPE_KEYS = ("type", "item_type", "itemtype", "category", "kind")
_ENV_KEYS = ("environment", "env")


def _norm_from(value: Any, synonyms: dict[str, str], allowed: list[str], default: str) -> str:
    """Resolve a free-text value to a canonical enum: exact match, then synonym
    table, then default."""
    if value is None:
        return default
    s = str(value).strip().lower()
    if not s:
        return default
    for canonical in allowed:
        if canonical.lower() == s:
            return canonical
    return synonyms.get(s, default)


def _first_key(row: dict, keys: tuple[str, ...]) -> Any:
    """Return the first non-empty value among `keys`, case-insensitively. Tries
    exact matches first, then a word-level contains check so headers like
    "Bug Summary" or "Sev." map to the right field."""
    low = {str(k).strip().lower(): v for k, v in row.items()}
    for k in keys:
        if k in low and low[k] not in (None, ""):
            return low[k]
    # Fuzzy fallback: a column whose name contains the key as a word.
    for k in keys:
        for col, v in low.items():
            if v in (None, ""):
                continue
            if k in re.split(r"[^a-z0-9]+", col):
                return v
    return None


def _clean_spec(raw: Any) -> Optional[dict]:
    """Normalise one raw entry into a validated spec, or None if it has no title."""
    if isinstance(raw, str):
        raw = {"title": raw}
    if not isinstance(raw, dict):
        return None
    title = _first_key(raw, _TITLE_KEYS)
    title = str(title).strip() if title is not None else ""
    if len(title) < MIN_TITLE_LENGTH:
        return None
    desc = _first_key(raw, _DESC_KEYS)
    return {
        "title": title[:200],
        "description": str(desc).strip() if desc is not None else "",
        "priority": _norm_from(_first_key(raw, _PRIORITY_KEYS), _PRIORITY_SYNONYMS, ALLOWED_PRIORITIES, "Medium"),
        "item_type": _norm_from(_first_key(raw, _TYPE_KEYS), _TYPE_SYNONYMS, ALLOWED_ITEM_TYPES, "Bug"),
        "environment": _norm_from(_first_key(raw, _ENV_KEYS), _ENV_SYNONYMS, ALLOWED_ENVIRONMENTS, "DEV"),
    }


def _clean_all(rows: list[Any]) -> list[dict]:
    out: list[dict] = []
    for raw in rows:
        spec = _clean_spec(raw)
        if spec is not None:
            out.append(spec)
        if len(out) >= MAX_SPECS:
            break
    return out


# --- Format detection + per-format parsers (deterministic path) ---
def _specs_from_json(text: str) -> Optional[list[dict]]:
    try:
        data = json.loads(text)
    except ValueError:
        return None
    if isinstance(data, dict):
        for key in ("bugs", "items", "issues", "tasks", "rows", "data"):
            if isinstance(data.get(key), list):
                return _clean_all(data[key])
        # A single object describing one bug.
        return _clean_all([data])
    if isinstance(data, list):
        return _clean_all(data)
    return None


def _specs_from_csv(text: str) -> list[dict]:
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    return _rows_to_specs(rows)


# Leading list markers (bullets, numbers, checkboxes, blockquotes, headings).
_LINE_MARKER = re.compile(r"^\s*(?:[-*+>]\s+|\d+[.)]\s+|\[[ xX]\]\s+|#{1,6}\s+)")
# A line that's just a separator / rule.
_SEPARATOR = re.compile(r"^[\s\-=*_#.]+$")
# A trailing/leading priority tag like [High] or (Critical).
_PRIORITY_TAG = re.compile(r"[\[(]\s*(critical|blocker|p0|urgent|high|p1|major|"
                           r"medium|p2|normal|low|p3|minor|trivial)\s*[\])]", re.IGNORECASE)


def _split_line_meta(line: str) -> dict:
    """Turn one free-text line into a raw spec dict, pulling out a `[Priority]`
    tag and a `Title | Description` split."""
    raw: dict[str, Any] = {}
    tag = _PRIORITY_TAG.search(line)
    if tag:
        raw["priority"] = tag.group(1)
        line = _PRIORITY_TAG.sub("", line).strip()
    if "|" in line:
        title, _, desc = line.partition("|")
        raw["title"] = title.strip()
        raw["description"] = desc.strip()
    else:
        raw["title"] = line.strip()
    return raw


def _specs_from_lines(text: str) -> list[dict]:
    raws: list[dict] = []
    for line in text.splitlines():
        stripped = _LINE_MARKER.sub("", line).strip()
        if not stripped or _SEPARATOR.match(stripped):
            continue
        raws.append(_split_line_meta(stripped))
    return _clean_all(raws)


# Words that, when seen in a first row, mark it as a header rather than data.
_HEADER_WORDS = frozenset(
    _TITLE_KEYS + _DESC_KEYS + _PRIORITY_KEYS + _TYPE_KEYS + _ENV_KEYS
    + ("status", "state", "assignee", "owner", "id", "#", "no", "no.")
)


def _looks_like_header(row: list[str]) -> bool:
    """True if the row names a recognized column (title, priority, status and
    so on), the reliable signal that it's a header, not data."""
    return any(c.strip().lower() in _HEADER_WORDS for c in row if c.strip())


# Bound the workbook scan so a decompression-bomb xlsx (billions of declared
# cells) can't blow up memory before MAX_SPECS applies. Far above any real sheet.
_MAX_XLSX_ROWS = 5000
_MAX_XLSX_COLS = 100


def _xlsx_rows(raw: bytes) -> Optional[list[list[str]]]:
    """Read the first sheet's non-empty rows as lists of strings, or None if the
    bytes aren't a readable workbook."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — not an xlsx / corrupt → fall through
        return None
    try:
        ws = wb.active
        # read_only iter_rows ignores max_row/col, so bound explicitly.
        rows = [
            [("" if c is None else str(c)) for c in row[:_MAX_XLSX_COLS]]
            for row in islice(ws.iter_rows(values_only=True), _MAX_XLSX_ROWS)
        ]
    finally:
        wb.close()
    return [r for r in rows if any(c.strip() for c in r)]


def _rows_to_specs(rows: list[list[str]]) -> list[dict]:
    """Convert table rows to specs. A header row maps columns by name; otherwise
    the first cell is the title. Multi-column sheets treat row 0 as a header so
    it doesn't become a spurious work item."""
    if not rows:
        return []
    multi_col = max((len(r) for r in rows), default=1) >= 2
    if _looks_like_header(rows[0]) or multi_col:
        header = [c.strip().lower() for c in rows[0]]
        return _clean_all([dict(zip(header, r)) for r in rows[1:]])
    return _clean_all([r[0] for r in rows])


def _specs_from_xlsx(raw: bytes) -> Optional[list[dict]]:
    """Parse the first sheet of an xlsx into specs, or None if the bytes aren't
    a readable workbook."""
    rows = _xlsx_rows(raw)
    if rows is None:
        return None
    return _rows_to_specs(rows)


def _looks_like_xlsx(filename: str, raw: bytes) -> bool:
    # xlsx is a zip (magic "PK\x03\x04"); check the extension too.
    return filename.lower().endswith(".xlsx") or raw[:4] == b"PK\x03\x04"


def parse_document(filename: str, raw: bytes) -> list[dict]:
    """Deterministic parse of an upload into cleaned specs, trying the suggested
    format then falling back to line parsing for misnamed text files."""
    filename = filename or ""
    if _looks_like_xlsx(filename, raw):
        specs = _specs_from_xlsx(raw)
        if specs is not None:
            return specs
    # Decode text once for the remaining formats.
    text = raw.decode("utf-8", errors="replace")
    stripped = text.lstrip()
    if filename.lower().endswith(".json") or stripped[:1] in ("[", "{"):
        specs = _specs_from_json(text)
        if specs is not None:
            return specs
    if filename.lower().endswith(".csv") or ("," in (text.splitlines() or [""])[0]):
        csv_specs = _specs_from_csv(text)
        if csv_specs:
            return csv_specs
    return _specs_from_lines(text)


# --- AI extraction (optional, isolated, mockable) ---
_AI_SYSTEM = (
    "You read a document an admin uploaded (it may be a spreadsheet rendered as "
    "a pipe-delimited table, CSV, JSON, or free prose) and extract the software "
    "work items (bugs / requirements / tasks) it describes. Return ONE JSON "
    "object and nothing else:\n"
    '  {"bugs": [{"title": str, "description": str, "priority": '
    '"Low|Medium|High|Critical", "item_type": "Bug|Requirement|Task", '
    '"environment": "DEV|UAT|PROD"}]}\n'
    "Rules: the FIRST row of a table is usually a header naming the columns — "
    "use it to map fields, never emit the header itself as an item. Use the "
    "document's own wording for each title. Read every data row. Infer "
    "priority/type/environment when the text implies them, else use Medium / "
    "Bug / DEV. Skip headings, preamble, blank rows and anything that isn't an "
    "actual work item. Never invent items that aren't in the document."
)


def ai_extract_specs(text: str) -> Optional[list[dict]]:
    """Read the document with the cloud assistant when configured. Returns cleaned
    specs, or None when the layer is off/unavailable/unhelpful (caller falls back)."""
    try:
        from app.chatbot import cloud_llm
        if not cloud_llm.is_available():
            return None
        parsed = cloud_llm.complete_json(_AI_SYSTEM, text[:20000])
    except Exception:  # noqa: BLE001 — an AI fault must never break ingest
        logger.exception("Sleuth ingest AI extraction failed; using parser")
        return None
    if not parsed or not isinstance(parsed.get("bugs"), list):
        return None
    specs = _clean_all(parsed["bugs"])
    return specs or None


def _document_text(filename: str, raw: bytes) -> str:
    """Plain-text view of the upload for the AI reader; spreadsheets become
    pipe-delimited tables, everything else is decoded as UTF-8."""
    if _looks_like_xlsx(filename or "", raw):
        rows = _xlsx_rows(raw)
        if not rows:
            return ""
        return "\n".join(" | ".join(c.strip() for c in r) for r in rows)
    return raw.decode("utf-8", errors="replace")


def extract_specs(filename: str, raw: bytes) -> tuple[list[dict], str]:
    """Top-level extraction: AI first (better at headers/messy layouts/prose),
    falling back to the deterministic parser. Returns (specs, 'ai'|'parser')."""
    text = _document_text(filename, raw)
    if text.strip():
        ai = ai_extract_specs(text)
        if ai:
            return ai, "ai"
    return parse_document(filename, raw), "parser"


def resolve_project_for_preview(db: Session, project_id: Optional[int]) -> Optional[Project]:
    """Resolve the project ingested items would land in (for the preview),
    without creating anything. None when there's no project."""
    return _resolve_project(db, project_id)


def _resolve_project(db: Session, project_id: Optional[int]) -> Optional[Project]:
    if project_id is not None:
        return db.get(Project, project_id)
    # Default to the lowest-id (bootstrap "General") project so ingest needs no pick.
    return db.scalar(select(Project).order_by(Project.id))


def create_bugs_from_specs(
    db: Session, specs: list[dict], actor: User, project_id: Optional[int] = None,
) -> dict:
    """Create one audited Bug per spec under a project, reported by the admin.
    Returns a summary dict; raises ValueError if the project can't be resolved."""
    project = _resolve_project(db, project_id)
    if project is None:
        raise ValueError("No project to file these items under — create a project first")

    created: list[dict] = []
    for spec in specs[:MAX_SPECS]:
        bug = Bug(
            project_id=project.id,
            reporter_id=actor.id,
            title=spec["title"],
            description=sanitize_html(spec.get("description", "")),
            item_type=spec.get("item_type", "Bug"),
            status="New",
            priority=spec.get("priority", "Medium"),
            environment=spec.get("environment", "DEV"),
        )
        db.add(bug)
        db.flush()
        db.add(Activity(
            bug_id=bug.id, entity_type="bug", entity_id=bug.id,
            actor_user_id=actor.id, actor_name=actor.name, action="bug_created",
            detail=f"{bug.item_type} #{bug.id} '{bug.title}' created via Sleuth document ingest",
        ))
        created.append({"id": bug.id, "title": bug.title})
    db.commit()
    return {
        "created": len(created),
        "project_id": project.id,
        "project_name": project.name,
        "items": created,
    }


__all__ = [
    "parse_document", "ai_extract_specs", "extract_specs",
    "create_bugs_from_specs", "resolve_project_for_preview",
    "MAX_SPECS", "MAX_DOC_BYTES",
]
