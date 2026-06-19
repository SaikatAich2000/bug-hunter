"""
Spins up a fresh SQLite DB, seeds users / projects / bugs, exercises the
parser, executor, Excel generator, classifier, and HTTP route. Read-only
by design at this stage — write-action tests live in _action_test.py.
Run from the bug-hunter root: python3 _chatbot_test.py
"""
from __future__ import annotations

import os as _os, sys as _sys
# Make the bug-hunter root importable when run directly.
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))

import os
import sys
import tempfile
import traceback
from datetime import datetime, timedelta, timezone

_tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_tmp.close()
os.environ["DATABASE_URL"] = f"sqlite:///{_tmp.name}"
os.environ["SESSION_SECRET"] = "test-secret-key-please-ignore-this"
os.environ["BOOTSTRAP_ADMIN_EMAIL"] = "admin@example.com"
os.environ["BOOTSTRAP_ADMIN_PASSWORD"] = "AdminPass123!"
os.environ["BOOTSTRAP_ADMIN_NAME"] = "Admin Person"
# Cloud assistant stays off in tests (avoid a real network call from a
# developer's local .env that enables it).
os.environ["SLEUTH_CLOUD_ENABLED"] = "0"

# Force a fresh app import bound to THIS file's dedicated DB (see the long note
# in test_sleuth_actions.py) — avoids a full-suite "no such table" from a shared
# engine bound to an earlier-collected module's torn-down DB.
import sys as _sys_purge
for _m in list(_sys_purge.modules):
    if _m == "app" or _m.startswith("app."):
        del _sys_purge.modules[_m]

from app.database import Base, engine, SessionLocal
from app import models
import pytest

from app.auth import hash_password
from app.chatbot import nlu, executor, excel
from app.chatbot.executor import build_context

PASSED: list[str] = []
FAILED: list[tuple[str, str]] = []


@pytest.fixture(autouse=True)
def _rebind_and_seed():
    """Re-bind this file's module-level app.* references to the current import
    generation each test, then seed its dedicated DB. conftest's `client`
    fixture purges `app.*` to rebind the engine per test; without re-binding and
    re-seeding here, parser's reads would hit a stale or empty engine. (Under
    pytest the ``__main__`` block that seeds in standalone mode never runs.)"""
    import importlib
    g = globals()
    db_mod = importlib.import_module("app.database")
    g["Base"], g["engine"], g["SessionLocal"] = (
        db_mod.Base, db_mod.engine, db_mod.SessionLocal,
    )
    g["models"] = importlib.import_module("app.models")
    g["hash_password"] = importlib.import_module("app.auth").hash_password
    g["nlu"] = importlib.import_module("app.chatbot.nlu")
    g["executor"] = importlib.import_module("app.chatbot.executor")
    g["excel"] = importlib.import_module("app.chatbot.excel")
    g["build_context"] = g["executor"].build_context
    seed()
    yield


def check(name: str, cond: bool, detail: str = "") -> None:
    if cond:
        PASSED.append(name)
        print(f"  PASS  {name}")
    else:
        FAILED.append((name, detail))
        raise AssertionError(f"{name}: {detail}")


def section(title: str) -> None:
    print(f"\n=== {title} ===")


def seed() -> None:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        admin = models.User(name="Admin Person", email="admin@example.com",
                            role="admin",
                            password_hash=hash_password("AdminPass123!"),
                            is_active=True)
        alice = models.User(name="Alice Wonderland", email="alice@example.com",
                            role="manager",
                            password_hash=hash_password("Alice12345!"),
                            is_active=True)
        bob = models.User(name="Bob Builder", email="bob@example.com",
                          role="user",
                          password_hash=hash_password("Bob12345!"),
                          is_active=True)
        carol = models.User(name="Carol Singer", email="carol@example.com",
                            role="user",
                            password_hash=hash_password("Carol12345!"),
                            is_active=True)
        db.add_all([admin, alice, bob, carol])
        db.commit()

        proj_a = models.Project(name="Apollo", description="Mission control")
        proj_b = models.Project(name="Beacon", description="Auth service")
        db.add_all([proj_a, proj_b])
        db.commit()

        now = datetime.now(timezone.utc)
        bugs = [
            models.Bug(title="Login button does nothing on Safari",
                       description="Safari 17 does not fire click",
                       status="New", priority="High", environment="PROD",
                       project_id=proj_b.id, reporter_id=alice.id,
                       created_at=now - timedelta(days=2)),
            models.Bug(title="Crash when uploading 100MB file",
                       description="OOM on attachment upload",
                       status="In Progress", priority="Critical",
                       environment="PROD",
                       project_id=proj_a.id, reporter_id=admin.id,
                       created_at=now - timedelta(days=5)),
            models.Bug(title="Typo on landing page",
                       description="welocme -> welcome",
                       status="Resolved", priority="Low", environment="DEV",
                       project_id=proj_b.id, reporter_id=carol.id,
                       created_at=now - timedelta(days=10)),
            models.Bug(title="Date filter off by one timezone",
                       description="UTC vs local",
                       status="Reopened", priority="Medium", environment="UAT",
                       project_id=proj_a.id, reporter_id=alice.id,
                       created_at=now - timedelta(hours=18)),
            models.Bug(title="Spam in audit log",
                       description="Phantom logout events",
                       status="Closed", priority="Low", environment="PROD",
                       project_id=proj_b.id, reporter_id=bob.id,
                       created_at=now - timedelta(days=30)),
            models.Bug(title="Critical security gap in auth",
                       description="Token reuse possible",
                       status="New", priority="Critical", environment="PROD",
                       project_id=proj_b.id, reporter_id=admin.id,
                       created_at=now - timedelta(hours=4)),
            models.Bug(title="Marked not a real issue",
                       description="user error",
                       status="Not a Bug", priority="Low", environment="DEV",
                       project_id=proj_a.id, reporter_id=bob.id,
                       created_at=now - timedelta(days=1)),
        ]
        db.add_all(bugs)
        db.commit()

        bugs[0].assignees = [bob]
        bugs[1].assignees = [bob, carol]
        bugs[3].assignees = [alice]
        bugs[5].assignees = [alice, bob]
        db.commit()

        from app.models import Activity
        db.add_all([
            Activity(actor_user_id=admin.id, actor_name="Admin Person",
                     action="bug.create",
                     entity_type="bug", entity_id=bugs[0].id,
                     detail="Created bug #1"),
            Activity(actor_user_id=alice.id, actor_name="Alice Wonderland",
                     action="bug.update",
                     entity_type="bug", entity_id=bugs[1].id,
                     detail="Changed status to In Progress"),
        ])
        db.commit()
        print(f"Seeded: 4 users, 2 projects, {len(bugs)} bugs")
    finally:
        db.close()


def test_nlu() -> None:
    section("NLU parser")
    db = SessionLocal()
    try:
        ctx = build_context(db)
        cases = [
            ("show open bugs assigned to bob",
             lambda pq: pq.intent == "list_bugs"
                        and any("bob" in n.lower() for n in pq.assignee_names)
                        and "New" in pq.statuses),
            ("how many critical bugs in prod?",
             lambda pq: pq.intent == "list_bugs" and pq.wants_count
                        and "Critical" in pq.priorities
                        and "PROD" in pq.environments),
            ("export all bugs in apollo to excel",
             lambda pq: pq.intent == "list_bugs" and pq.wants_export
                        and any("apollo" in n.lower() for n in pq.project_names)),
            ("bug #4",
             lambda pq: pq.intent == "bug_detail" and pq.bug_id == 4),
            ("hi",
             lambda pq: pq.intent == "greeting"),
            ("thanks!",
             lambda pq: pq.intent == "thanks"),
            ("help",
             lambda pq: pq.intent == "help"),
            ("list all managers",
             lambda pq: pq.intent == "list_users" and pq.role_filter == "manager"),
            ("summary",
             lambda pq: pq.intent == "stats"),
            ("recent activity",
             lambda pq: pq.intent == "recent_activity"),
            ("bugs reported by alice in last 7 days",
             lambda pq: pq.intent == "list_bugs"
                        and pq.time_window is not None
                        and any("alice" in n.lower() for n in pq.reporter_names)),
            ("show all p0 bugs",
             lambda pq: pq.intent == "list_bugs" and "Critical" in pq.priorities),
            ("blockers in production",
             lambda pq: pq.intent == "list_bugs"
                        and "Critical" in pq.priorities
                        and "PROD" in pq.environments),
            ("list projects",
             lambda pq: pq.intent == "list_projects"),
            ("show users",
             lambda pq: pq.intent == "list_users"),
        ]
        for msg, pred in cases:
            try:
                pq = nlu.parse(msg, ctx)
                ok = pred(pq)
                d = "" if ok else (
                    f"intent={pq.intent} st={pq.statuses} pr={pq.priorities} "
                    f"env={pq.environments} a={pq.assignee_names} "
                    f"r={pq.reporter_names} p={pq.project_names} "
                    f"cnt={pq.wants_count} exp={pq.wants_export} "
                    f"bid={pq.bug_id} role={pq.role_filter}")
                check(f"NLU: {msg!r}", ok, d)
            except Exception as e:
                check(f"NLU: {msg!r}", False, f"EXC: {e!r}")
    finally:
        db.close()


def test_executor() -> None:
    section("Executor end-to-end")
    db = SessionLocal()
    try:
        admin = db.query(models.User).filter_by(email="admin@example.com").one()
        bob = db.query(models.User).filter_by(email="bob@example.com").one()

        # 1. open bugs vs bob → IDs 1, 2, 6
        resp = executor.execute("show open bugs assigned to bob", db, admin)
        tbl = next((b for b in resp.blocks if b.kind == "table"), None)
        check("open bugs vs bob — table present", tbl is not None)
        if tbl:
            ids = sorted(tbl.payload.get("row_bug_ids", []))
            check("open bugs vs bob — IDs == [1,2,6]",
                  ids == [1, 2, 6], f"got {ids}")

        # 2. critical PROD count = 2
        resp = executor.execute("how many critical bugs in prod?", db, admin)
        text = " ".join(b.payload.get("text", "")
                        for b in resp.blocks if b.kind == "text")
        check("critical-PROD count text contains '2'", "2" in text, text)

        # 3. Excel export
        resp = executor.execute("export all bugs in apollo to excel", db, admin)
        fb = next((b for b in resp.blocks if b.kind == "file"), None)
        check("export apollo — file block present", fb is not None)
        if fb:
            data = excel.fetch_staged(fb.payload["download_token"])
            check("export apollo — fetchable", data is not None)
            if data:
                check("export apollo — non-empty", len(data[0]) > 500)
                check("export apollo — .xlsx", data[1].endswith(".xlsx"))

        # 4. Bug detail
        resp = executor.execute("bug #4", db, admin)
        text_blocks = [b for b in resp.blocks if b.kind == "text"]
        text = " ".join(b.payload.get("text", "") for b in text_blocks)
        check("bug #4 detail mentions Reopened or Date filter",
              "Reopened" in text or "Date filter" in text, text[:200])
        opens = [b.payload.get("open_bug_id") for b in text_blocks]
        check("bug #4 — open_bug_id == 4", 4 in opens, f"opens={opens}")

        # 5. Greeting
        resp = executor.execute("hi", db, admin)
        check("greeting returns text",
              any(b.kind == "text" for b in resp.blocks))

        # 6. List users
        resp = executor.execute("list all users", db, admin)
        tbl = next((b for b in resp.blocks if b.kind == "table"), None)
        check("list users — table present", tbl is not None)
        if tbl:
            names = [r[0] for r in tbl.payload["rows"]]
            check("list users — all 4 names",
                  set(names) == {"Admin Person", "Alice Wonderland",
                                 "Bob Builder", "Carol Singer"},
                  f"got {names}")

        # 7. Managers only
        resp = executor.execute("list managers", db, admin)
        tbl = next((b for b in resp.blocks if b.kind == "table"), None)
        if tbl:
            names = [r[0] for r in tbl.payload["rows"]]
            check("list managers — only Alice",
                  names == ["Alice Wonderland"], f"got {names}")

        # 8. Stats
        resp = executor.execute("summary", db, admin)
        text = " ".join(b.payload.get("text", "")
                        for b in resp.blocks if b.kind == "text")
        check("stats — has open/Open",
              "Open" in text or "open" in text, text[:200])

        # 9-10. Recent activity
        resp = executor.execute("recent activity", db, admin)
        check("recent activity (admin) — has blocks", len(resp.blocks) > 0)
        resp = executor.execute("recent activity", db, bob)
        check("recent activity (non-admin) — has blocks", len(resp.blocks) > 0)

        # 11. Unknown
        resp = executor.execute("xyzzy frobnicate qux", db, admin)
        check("unknown — has fallback text",
              any(b.kind == "text" for b in resp.blocks))

        # 12. Closed bugs only
        resp = executor.execute("show closed bugs", db, admin)
        tbl = next((b for b in resp.blocks if b.kind == "table"), None)
        if tbl:
            ids = sorted(tbl.payload.get("row_bug_ids", []))
            check("closed bugs — only Bug #5", ids == [5], f"got {ids}")

        # 13. About / statuses
        resp = executor.execute("what statuses are there", db, admin)
        text = " ".join(b.payload.get("text", "")
                        for b in resp.blocks if b.kind == "text")
        check("about-statuses — mentions New",
              "New" in text or "new" in text, text[:200])

        # 14. Priority synonym 'blockers'
        resp = executor.execute("show me the blockers", db, admin)
        tbl = next((b for b in resp.blocks if b.kind == "table"), None)
        if tbl:
            ids = sorted(tbl.payload.get("row_bug_ids", []))
            check("blockers — IDs == [2,6]", ids == [2, 6], f"got {ids}")

        # 15. Env synonym 'production'
        resp = executor.execute("bugs in production", db, admin)
        tbl = next((b for b in resp.blocks if b.kind == "table"), None)
        if tbl:
            ids = sorted(tbl.payload.get("row_bug_ids", []))
            check("PROD bugs — IDs == [1,2,5,6]",
                  ids == [1, 2, 5, 6], f"got {ids}")

        # 16. Text search
        resp = executor.execute("find bugs about login", db, admin)
        tbl = next((b for b in resp.blocks if b.kind == "table"), None)
        if tbl:
            ids = sorted(tbl.payload.get("row_bug_ids", []))
            check("text search 'login' — finds Bug #1", 1 in ids, f"got {ids}")

        # 17. SQL injection attempt
        before = db.query(models.Bug).count()
        resp = executor.execute("'; DROP TABLE bugs; --", db, admin)
        check("injection — no crash", len(resp.blocks) > 0)
        after = db.query(models.Bug).count()
        check("bugs table unchanged", before == after,
              f"before={before} after={after}")

        # 18. Long input
        long_msg = "show me " + ("really " * 200) + "long bugs"
        resp = executor.execute(long_msg, db, admin)
        check("long input — no crash", len(resp.blocks) > 0)

        # 19. Unicode
        resp = executor.execute("показать все баги", db, admin)
        check("unicode — no crash", len(resp.blocks) > 0)

        # 20. All bugs
        resp = executor.execute("show me all bugs", db, admin)
        tbl = next((b for b in resp.blocks if b.kind == "table"), None)
        check("all bugs — table present", tbl is not None)
        if tbl:
            ids = sorted(tbl.payload.get("row_bug_ids", []))
            check("all bugs — at least 6", len(ids) >= 6, f"got {ids}")
    finally:
        db.close()


def test_excel() -> None:
    section("Excel cache + generation")
    excel.clear_all_for_test()
    rows = [
        {"id": 1, "title": "Foo", "project": "P", "status": "New",
         "priority": "High", "environment": "PROD", "reporter": "alice",
         "assignees": "bob", "due_date": "", "created_at": "2025-01-01",
         "updated_at": "2025-01-02"},
        {"id": 2, "title": "Bar", "project": "P", "status": "Closed",
         "priority": "Low", "environment": "DEV", "reporter": "bob",
         "assignees": "", "due_date": "", "created_at": "2025-01-01",
         "updated_at": "2025-01-02"},
    ]
    token, size = excel.stage_workbook(rows, "test.xlsx", "Filter: any")
    check("excel — token returned", isinstance(token, str) and len(token) > 8)
    check("excel — size > 500", size > 500)
    fetched = excel.fetch_staged(token)
    check("excel — fetched OK", fetched is not None)
    if fetched:
        b, fn = fetched
        check("excel — filename", fn == "test.xlsx")
        check("excel — len matches size", len(b) == size)
    bad = excel.fetch_staged("nope-xyz")
    check("excel — unknown token None", bad is None)
    if fetched:
        import io
        import openpyxl as _opx
        wb = _opx.load_workbook(io.BytesIO(fetched[0]))
        ws = wb.active
        check("excel — workbook has rows", ws.max_row >= 4)
        headers = [ws.cell(row=2, column=c).value for c in range(1, 12)]
        check("excel — header has 'ID'", "ID" in (headers or []))


def test_router() -> None:
    section("HTTP route /api/chat")
    from fastapi.testclient import TestClient
    from app.main import app
    client = TestClient(app)
    r = client.post("/api/auth/login",
                    json={"email": "admin@example.com",
                          "password": "AdminPass123!"})
    check("login as admin — 200", r.status_code == 200,
          f"st={r.status_code} body={r.text[:200]}")

    bare = TestClient(app)
    r = bare.post("/api/chat/ask", json={"message": "hi"})
    check("/api/chat/ask without auth — 401/403",
          r.status_code in (401, 403), f"got {r.status_code}")

    r = client.post("/api/chat/ask", json={"message": "hi"})
    check("ask hi — 200", r.status_code == 200,
          f"st={r.status_code} body={r.text[:200]}")
    if r.status_code == 200:
        body = r.json()
        check("ask hi — has blocks",
              isinstance(body.get("blocks"), list) and len(body["blocks"]) > 0)
        check("ask hi — intent=greeting",
              body.get("intent") == "greeting", str(body.get("intent")))

    r = client.post("/api/chat/ask", json={"message": "help"})
    check("ask help — 200", r.status_code == 200)

    r = client.post("/api/chat/ask",
                    json={"message": "show open bugs assigned to bob"})
    check("ask: open bugs vs bob — 200", r.status_code == 200)
    if r.status_code == 200:
        body = r.json()
        tbl = next((b for b in body["blocks"] if b["kind"] == "table"), None)
        check("ask: open bugs vs bob — has table", tbl is not None)

    r = client.post("/api/chat/ask",
                    json={"message": "export all bugs in apollo to excel"})
    check("ask: export apollo — 200", r.status_code == 200)
    if r.status_code == 200:
        body = r.json()
        fb = next((b for b in body["blocks"] if b["kind"] == "file"), None)
        check("ask: export apollo — has file block", fb is not None)
        if fb:
            tok = fb["payload"]["download_token"]
            r2 = client.get(f"/api/chat/download/{tok}")
            check("download — 200", r2.status_code == 200)
            check("download — xlsx ct",
                  "spreadsheetml" in r2.headers.get("content-type", ""))
            check("download — attachment disposition",
                  "attachment" in r2.headers.get("content-disposition", ""))
            check("download — non-empty body", len(r2.content) > 500)
            r3 = bare.get(f"/api/chat/download/{tok}")
            check("download — unauth blocked", r3.status_code in (401, 403))

    r = client.post("/api/chat/ask", json={"message": ""})
    check("ask empty — 200/422", r.status_code in (200, 422))

    r = client.post("/api/chat/ask", json={"message": "x" * 5000})
    check("ask oversize — 422", r.status_code == 422)

    statuses = []
    for _ in range(35):
        rr = client.post("/api/chat/ask", json={"message": "hi"})
        statuses.append(rr.status_code)
    check("rate limit — 429 fires", 429 in statuses,
          f"sample={set(statuses)}")


if __name__ == "__main__":
    try:
        seed()
        test_nlu()
        test_executor()
        test_excel()
        test_router()
    except Exception:
        traceback.print_exc()
        FAILED.append(("HARNESS", "uncaught crash"))

    print(f"\n=== RESULTS ===")
    print(f"Passed: {len(PASSED)}")
    print(f"Failed: {len(FAILED)}")
    if FAILED:
        for n, d in FAILED:
            print(f"  FAIL  {n}  {d}")
        sys.exit(1)
    print("All checks passed")
    sys.exit(0)
