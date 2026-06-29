"""Security hardening tests.

Covers login timing (no user-enumeration via latency), CSV/XLSX formula
injection, body-size middleware, X-Forwarded-For trust gating, email
masking in logs, per-account lockout, HaveIBeenPwned breach checks, and
EXIF/metadata stripping on image uploads.

Each section pairs unit tests with route-level integration tests where
practical. In-memory state (lockout buckets, HIBP backend) is reset
between cases via ``reset_security_state`` so the suite is
order-independent.
"""
from __future__ import annotations

import io
import logging
from unittest import mock

import pytest


@pytest.fixture(autouse=True)
def reset_security_state():
    """Reset in-memory lockout state after each test.

    The lockout buckets live in process memory, so a burst of failures in
    one test would carry over and lock the same email out in a later test.
    Cleanup runs after yield so a failing test doesn't skip it.
    """
    yield
    try:
        from app import account_lockout
        account_lockout._reset_for_tests()
    except ImportError:
        pass


def _make_bug(client, title: str, project_id: int = 1) -> int:
    res = client.post("/api/bugs", json={
        "project_id": project_id, "title": title,
        "description": "desc", "item_type": "Bug",
        "status": "New", "priority": "Medium", "environment": "DEV",
    })
    assert res.status_code == 201, res.text
    return res.json()["id"]


# ---------------------------------------------------------------------------
# Login timing equality
# ---------------------------------------------------------------------------
class TestLoginTimingEquality:
    """Unknown-email and wrong-password branches must both run one bcrypt
    verification so response latency can't be used to enumerate accounts."""

    def test_unknown_email_still_runs_bcrypt(self, client):
        # Spy on verify_password rather than paying the real bcrypt cost.
        from app.routes import auth as auth_routes
        with mock.patch.object(auth_routes, "verify_password",
                               wraps=auth_routes.verify_password) as spy:
            res = client.post("/api/auth/login", json={
                "email": "no-such-user@example.com",
                "password": "anything-at-all-9",
            })
        assert res.status_code == 401
        assert spy.call_count == 1, (
            "Unknown-email branch did not run bcrypt — timing oracle still open"
        )
        # Must use the dummy hash, not None or an empty string.
        args, _ = spy.call_args
        assert args[1] == auth_routes._DUMMY_PASSWORD_HASH

    def test_wrong_password_branch_unchanged(self, client):
        # Confirm the real-user/wrong-password path still hits verify_password.
        from tests.conftest import BOOTSTRAP_EMAIL
        from app.routes import auth as auth_routes
        with mock.patch.object(auth_routes, "verify_password",
                               wraps=auth_routes.verify_password) as spy:
            res = client.post("/api/auth/login", json={
                "email": BOOTSTRAP_EMAIL,
                "password": "definitely-not-the-right-pwd-9",
            })
        assert res.status_code == 401
        assert spy.call_count == 1

    def test_unified_401_message_for_both_branches(self, client):
        from tests.conftest import BOOTSTRAP_EMAIL
        a = client.post("/api/auth/login", json={
            "email": "no-such-user@example.com", "password": "abc12345",
        })
        b = client.post("/api/auth/login", json={
            "email": BOOTSTRAP_EMAIL, "password": "wrong-password-9",
        })
        assert a.status_code == b.status_code == 401
        assert a.json()["detail"] == b.json()["detail"]


# ---------------------------------------------------------------------------
# Spreadsheet formula injection (XLSX)
#
# openpyxl treats cell values starting with =, +, -, @, \t, or \r as
# formulas when opened in Excel. Titles are defanged in
# app/reports/xlsx.py::_defang_formula_text before being written.
# ---------------------------------------------------------------------------
class TestXlsxFormulaInjectionGuard:

    @pytest.mark.parametrize("trigger", ["=", "+", "-", "@", "\t", "\r"])
    def test_defang_helper_neutralises_formula_triggers(self, trigger):
        from app.reports.xlsx import _defang_formula_text
        assert _defang_formula_text(trigger + "cmd|calc!A1").startswith("'" + trigger)

    def test_defang_helper_passes_through_normal_text(self):
        from app.reports.xlsx import _defang_formula_text
        assert _defang_formula_text("Login button broken") == "Login button broken"

    def test_defang_helper_handles_empty_string(self):
        from app.reports.xlsx import _defang_formula_text
        assert _defang_formula_text("") == ""

    def test_export_xlsx_prefixes_malicious_title(self, admin_client):
        """A formula-shaped bug title must appear in the XLSX with a leading
        single-quote so spreadsheet apps treat it as text, not a formula."""
        import io
        from openpyxl import load_workbook
        bug_id = _make_bug(admin_client, "=cmd|'calc.exe'!A1")
        assert bug_id  # sanity
        res = admin_client.post("/api/reports/export.xlsx", json={
            "report_key": "item_detail", "filters": {},
        })
        assert res.status_code == 200
        wb = load_workbook(io.BytesIO(res.content), read_only=True)
        found_defanged = False
        for row in wb[wb.sheetnames[0]].iter_rows(values_only=True):
            for cell in row:
                if not isinstance(cell, str):
                    continue
                if "cmd|" in cell:
                    assert cell.startswith("'="), (
                        f"Un-neutralised formula in XLSX cell: {cell!r}"
                    )
                    found_defanged = True
        assert found_defanged, "expected the malicious title to appear (defanged) somewhere"


# ---------------------------------------------------------------------------
# Body size middleware
# ---------------------------------------------------------------------------
class TestBodySizeMiddleware:

    def test_normal_request_under_limit_succeeds(self, admin_client):
        res = admin_client.get("/api/auth/me")
        assert res.status_code == 200

    def test_oversize_content_length_rejected_with_413(self, client, monkeypatch):
        # Middleware checks Content-Length before reading the body, so we
        # don't need to actually send 70 MB. Lowering the limit and sending a
        # small body over it is done in the next test; this one just confirms
        # the default limit is at least 50 MB (as documented).
        from app.main import settings
        assert settings.MAX_REQUEST_BODY_BYTES >= 50 * 1024 * 1024

    def test_oversize_body_rejected(self, tmp_path, monkeypatch):
        """Use a tiny limit (1 KB) so the test doesn't need to allocate
        hundreds of MB to exercise the 413 path."""
        import sys
        db_file = tmp_path / "bodysize.db"
        monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_file}")
        monkeypatch.setenv("EMAIL_BACKEND", "disabled")
        monkeypatch.setenv("SESSION_SECRET", "test_secret")
        monkeypatch.setenv("BOOTSTRAP_ADMIN_EMAIL", "admin@test.local")
        monkeypatch.setenv("BOOTSTRAP_ADMIN_PASSWORD", "Admin1234")
        monkeypatch.setenv("MAX_REQUEST_BODY_BYTES", "1024")
        for mod in list(sys.modules):
            if mod == "app" or mod.startswith("app."):
                del sys.modules[mod]
        from app.config import get_settings
        get_settings.cache_clear()  # type: ignore[attr-defined]
        from fastapi.testclient import TestClient
        from app.main import app
        with TestClient(app) as c:
            c.post("/api/auth/login", json={
                "email": "admin@test.local", "password": "Admin1234",
            })
            payload = {"current_password": "Admin1234", "new_password": "x" * 2048}
            res = c.post("/api/auth/change-password", json=payload)
            # Middleware fires before the route handler, so the status is 413
            # regardless of what the route's own validation would say.
            assert res.status_code == 413, (
                f"Expected 413, got {res.status_code}: {res.text}"
            )
            assert "too large" in res.json()["detail"].lower()

    def test_malformed_content_length_returns_400(self):
        """A non-integer Content-Length must yield 400, not an uncaught
        ValueError from int(). Dispatched directly because httpx/uvicorn
        enforce numeric Content-Length on the wire."""
        import asyncio
        from starlette.requests import Request
        from app.main import BodySizeLimitMiddleware

        middleware = BodySizeLimitMiddleware(None)
        scope = {
            "type": "http", "method": "POST", "path": "/api/auth/login",
            "headers": [(b"content-length", b"not-a-number")],
            "query_string": b"", "scheme": "http",
            "server": ("testserver", 80), "client": ("test", 0),
            "raw_path": b"/api/auth/login",
        }
        request = Request(scope)

        async def call_next(_req):     # pragma: no cover — should not run
            raise AssertionError("middleware must short-circuit before call_next")

        response = asyncio.run(middleware.dispatch(request, call_next))
        assert response.status_code == 400
        body = response.body.decode("utf-8")
        assert "try again" in body.lower()


class TestAccountLockoutEdgeCases:

    def test_env_int_falls_back_on_garbage(self, monkeypatch):
        from app.account_lockout import _env_int
        monkeypatch.setenv("BH_BOGUS_INT_VAR", "not-a-number")
        assert _env_int("BH_BOGUS_INT_VAR", 42) == 42

    def test_env_int_uses_value(self, monkeypatch):
        from app.account_lockout import _env_int
        monkeypatch.setenv("BH_GOOD_INT_VAR", "99")
        assert _env_int("BH_GOOD_INT_VAR", 0) == 99

    def test_old_failures_get_evicted(self, monkeypatch):
        """Failures outside the rolling window must be evicted so an
        infrequent typo never accumulates into a lockout."""
        import time
        from app import account_lockout
        monkeypatch.setattr(account_lockout, "_LOGIN_FAIL_WINDOW_SECONDS", 0.05)
        monkeypatch.setattr(account_lockout, "_LOGIN_FAIL_LIMIT", 3)
        account_lockout._reset_for_tests()
        for _ in range(2):
            account_lockout.record_failure("evictee@x.com")
        time.sleep(0.07)
        account_lockout.record_failure("evictee@x.com")
        bucket = account_lockout._buckets.get("evictee@x.com")
        assert bucket is not None
        assert len(bucket.fails) == 1
        account_lockout.check_locked("evictee@x.com")  # must not raise

    def test_bucket_dict_cap_drops_oldest_entry(self, monkeypatch):
        from app import account_lockout
        monkeypatch.setattr(account_lockout, "_LOCKOUT_BUCKETS_MAX", 5)
        monkeypatch.setattr(account_lockout, "_LOGIN_FAIL_LIMIT", 10)
        account_lockout._reset_for_tests()
        # Fill past the cap → eviction kicks in on the 6th entry.
        for i in range(7):
            account_lockout.record_failure(f"user{i}@x.com")
        # 6th and 7th entries must have evicted the oldest to stay at cap.
        assert len(account_lockout._buckets) <= 5

    def test_clear_unknown_email_is_noop(self):
        """clear() on an email with no bucket must not blow up."""
        from app import account_lockout
        account_lockout._reset_for_tests()
        account_lockout.clear("never-recorded@x.com")  # must not raise
        assert "never-recorded@x.com" not in account_lockout._buckets


class TestPasswordBreachFetchRange:
    """Exercise the real ``_fetch_range`` implementation. All other breach
    tests patch this seam for hermeticity; these stub the httpx layer so
    the actual branching inside ``_fetch_range`` gets covered."""

    @pytest.fixture(autouse=True)
    def _enable(self, monkeypatch):
        monkeypatch.setenv("PASSWORD_BREACH_CHECK_ENABLED", "true")

    @staticmethod
    def _patch_httpx_client(monkeypatch, response_status=200, response_text="",
                            raise_on_get=None):
        from app import password_breach

        class _FakeResponse:
            status_code = response_status
            text = response_text

        class _FakeClient:
            def __init__(self, **_kw):
                pass
            def __enter__(self): return self
            def __exit__(self, *_a): return False
            def get(self, _url, **_kw):
                if raise_on_get is not None:
                    raise raise_on_get
                return _FakeResponse()

        monkeypatch.setattr(password_breach.httpx, "Client", _FakeClient)

    def test_fetch_range_returns_text_on_200(self, monkeypatch):
        from app import password_breach
        self._patch_httpx_client(monkeypatch, 200, "ABCD:1\n")
        assert password_breach._fetch_range("5BAA6") == "ABCD:1\n"

    def test_fetch_range_returns_none_on_non_200(self, monkeypatch):
        from app import password_breach
        self._patch_httpx_client(monkeypatch, 503, "")
        assert password_breach._fetch_range("5BAA6") is None

    def test_fetch_range_returns_none_on_httperror(self, monkeypatch):
        import httpx
        from app import password_breach
        self._patch_httpx_client(
            monkeypatch, raise_on_get=httpx.HTTPError("simulated")
        )
        assert password_breach._fetch_range("5BAA6") is None

    def test_fetch_range_returns_none_on_oserror(self, monkeypatch):
        from app import password_breach
        self._patch_httpx_client(
            monkeypatch, raise_on_get=OSError("connection refused")
        )
        assert password_breach._fetch_range("5BAA6") is None


class TestImageStripEdgeCases:

    def test_pillow_missing_returns_original(self, monkeypatch):
        """With Pillow absent the helper must fail open and return the
        original bytes unchanged."""
        import sys
        from app.image_strip import strip_image_metadata
        monkeypatch.setitem(sys.modules, "PIL", None)
        raw = b"fake-jpeg-bytes"
        assert strip_image_metadata(raw, "image/jpeg") == raw

    def test_format_none_returns_original(self, monkeypatch):
        """An image Pillow opens but whose .format is None (e.g. a raw
        in-memory buffer) must be returned unchanged."""
        from PIL import Image as PILImage
        from app.image_strip import strip_image_metadata

        class _FakeImg:
            format = None
            info: dict = {}
            size = (8, 8)
            def load(self):
                pass

        monkeypatch.setattr(PILImage, "open", lambda _src: _FakeImg())
        raw = b"any-bytes"
        assert strip_image_metadata(raw, "image/jpeg") == raw

    def test_save_oserror_returns_original(self, monkeypatch):
        """An OSError from Pillow's save() must cause the helper to log
        and return the original bytes rather than propagating the exception."""
        from PIL import Image as PILImage
        from app.image_strip import strip_image_metadata

        class _BoomImg:
            format = "JPEG"
            info: dict = {}
            def load(self):
                pass
            def save(self, _out, **_kw): raise OSError("disk full mid-encode")

        monkeypatch.setattr(PILImage, "open", lambda _src: _BoomImg())
        raw = b"any-bytes"
        assert strip_image_metadata(raw, "image/jpeg") == raw

    def test_content_type_with_charset_param_still_handled(self):
        """Some browsers send ``image/jpeg; charset=binary``. The helper
        must strip the parameter before matching the MIME prefix.

        No real JPEG here — just confirms the helper gets past the
        content-type guard. Pillow can't decode the bytes, so it fails
        open and returns the original. The full round-trip is covered by
        TestExifStrip.
        """
        from app.image_strip import strip_image_metadata
        raw = b"not-a-jpeg"
        assert strip_image_metadata(raw, "image/jpeg; charset=binary") == raw


# ---------------------------------------------------------------------------
# X-Forwarded-For trust gate
# ---------------------------------------------------------------------------
class TestXffTrustGate:

    def test_xff_ignored_when_trust_disabled(self, admin_client):
        # TRUST_PROXY_FORWARDED_FOR defaults to False.
        # An XFF header on login must not affect the recorded session IP.
        admin_client.post("/api/auth/logout")
        res = admin_client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin1234"},
            headers={"X-Forwarded-For": "203.0.113.99"},
        )
        assert res.status_code == 200
        sessions = admin_client.get("/api/sessions").json()
        assert sessions, "expected at least one session row"
        ips = [s["ip_address"] for s in sessions]
        assert "203.0.113.99" not in ips, (
            f"XFF was honoured despite TRUST_PROXY_FORWARDED_FOR=False: {ips}"
        )

    def test_xff_honoured_when_trust_enabled(self, tmp_path, monkeypatch):
        """With TRUST_PROXY_FORWARDED_FOR=true the right-most XFF entry (the
        one the trusted proxy appended) must be recorded. The left-most entry
        is client-controlled and must not be stored."""
        import sys
        db_file = tmp_path / "xff.db"
        monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_file}")
        monkeypatch.setenv("EMAIL_BACKEND", "disabled")
        monkeypatch.setenv("SESSION_SECRET", "test_secret")
        monkeypatch.setenv("BOOTSTRAP_ADMIN_EMAIL", "admin@test.local")
        monkeypatch.setenv("BOOTSTRAP_ADMIN_PASSWORD", "Admin1234")
        monkeypatch.setenv("TRUST_PROXY_FORWARDED_FOR", "true")
        monkeypatch.setenv("TRUST_PROXY_HOP_COUNT", "1")
        for mod in list(sys.modules):
            if mod == "app" or mod.startswith("app."):
                del sys.modules[mod]
        from app.config import get_settings
        get_settings.cache_clear()  # type: ignore[attr-defined]
        from fastapi.testclient import TestClient
        from app.main import app
        with TestClient(app) as c:
            res = c.post(
                "/api/auth/login",
                json={"email": "admin@test.local", "password": "Admin1234"},
                # "<attacker-forged>, <real proxy-appended peer>"
                headers={"X-Forwarded-For": "198.51.100.7, 10.0.0.1"},
            )
            assert res.status_code == 200, res.text
            sessions = c.get("/api/sessions").json()
            assert any(s["ip_address"] == "10.0.0.1" for s in sessions), (
                f"right-most XFF was NOT honoured despite TRUST=true: {sessions}"
            )
            assert not any(s["ip_address"] == "198.51.100.7" for s in sessions), (
                f"client-controlled left-most XFF must NOT be trusted: {sessions}"
            )


# ---------------------------------------------------------------------------
# Email masking in logs
# ---------------------------------------------------------------------------
class TestEmailMasking:

    @pytest.mark.parametrize("raw,expected", [
        ("alice@example.com",  "a***@example.com"),
        ("a@example.com",      "a***@example.com"),
        ("@example.com",       "@example.com"),
        ("",                   "***"),
        ("plain-no-at-sign",   "***"),
        ("BOB@example.com",    "B***@example.com"),
    ])
    def test_mask_helper(self, raw, expected):
        from app.routes.auth import _mask_email
        assert _mask_email(raw) == expected

    def test_inactive_login_logs_masked_email(self, admin_client, caplog):
        admin_client.post("/api/users", json={
            "name": "Disabled User", "email": "disabled@test.local",
            "role": "user", "password": "Disabled12",
        })
        users = admin_client.get("/api/users").json()
        target = next(u for u in users if u["email"] == "disabled@test.local")
        r = admin_client.put(f"/api/users/{target['id']}", json={"is_active": False})
        assert r.status_code == 200
        admin_client.post("/api/auth/logout")

        with caplog.at_level(logging.INFO, logger="bug_hunter.auth"):
            res = admin_client.post("/api/auth/login", json={
                "email": "disabled@test.local", "password": "Disabled12",
            })
        assert res.status_code == 401
        relevant = [r.message for r in caplog.records if "Login refused" in r.message]
        assert relevant, "expected the masked-email log line"
        for line in relevant:
            assert "disabled@test.local" not in line, (
                f"Raw email leaked into log: {line!r}"
            )
            assert "d***@test.local" in line


# ---------------------------------------------------------------------------
# Account lockout
# ---------------------------------------------------------------------------
class TestAccountLockout:

    def test_unit_check_locked_passes_when_no_state(self):
        from app import account_lockout
        account_lockout.check_locked("never-seen@example.com")

    def test_unit_threshold_triggers_lockout(self):
        from app import account_lockout
        email = "victim@example.com"
        for _ in range(account_lockout._LOGIN_FAIL_LIMIT):
            account_lockout.record_failure(email)
        with pytest.raises(Exception) as excinfo:
            account_lockout.check_locked(email)
        assert getattr(excinfo.value, "status_code", None) == 429
        assert "Retry-After" in (excinfo.value.headers or {})

    def test_unit_clear_resets_bucket(self):
        from app import account_lockout
        email = "transient@example.com"
        for _ in range(account_lockout._LOGIN_FAIL_LIMIT):
            account_lockout.record_failure(email)
        account_lockout.clear(email)
        account_lockout.check_locked(email)

    def test_unit_unknown_email_also_counts(self):
        """Ticking only known emails would leak account existence."""
        from app import account_lockout
        ghost = "definitely-not-a-real-user@example.com"
        for _ in range(account_lockout._LOGIN_FAIL_LIMIT):
            account_lockout.record_failure(ghost)
        with pytest.raises(Exception) as excinfo:
            account_lockout.check_locked(ghost)
        assert getattr(excinfo.value, "status_code", None) == 429

    def test_unit_disabled_when_limit_is_zero(self, monkeypatch):
        from app import account_lockout
        monkeypatch.setattr(account_lockout, "_LOGIN_FAIL_LIMIT", 0)
        for _ in range(50):
            account_lockout.record_failure("anyone@example.com")
        # limit == 0 is the operator opt-out; must never raise.
        account_lockout.check_locked("anyone@example.com")

    def test_http_login_429_after_threshold(self, client):
        """Drive the lockout via the route. The IP rate limit (8/60s)
        may fire first, but either way at least one response must be 429."""
        from tests.conftest import BOOTSTRAP_EMAIL
        codes = []
        for _ in range(15):
            r = client.post("/api/auth/login", json={
                "email": BOOTSTRAP_EMAIL, "password": "Wrong-pwd-9",
            })
            codes.append(r.status_code)
        assert 429 in codes, f"Expected 429 somewhere in {codes}"

    def test_successful_login_clears_lockout(self, client):
        """A successful login after some (sub-threshold) failures must
        clear the bucket so those failures don't carry over."""
        from tests.conftest import BOOTSTRAP_EMAIL, BOOTSTRAP_PASSWORD
        from app import account_lockout
        for _ in range(3):
            account_lockout.record_failure(BOOTSTRAP_EMAIL)
        res = client.post("/api/auth/login", json={
            "email": BOOTSTRAP_EMAIL, "password": BOOTSTRAP_PASSWORD,
        })
        assert res.status_code == 200
        account_lockout.check_locked(BOOTSTRAP_EMAIL)


# ---------------------------------------------------------------------------
# HIBP breach check
# ---------------------------------------------------------------------------
class TestPasswordBreachCheck:
    """conftest disables PASSWORD_BREACH_CHECK_ENABLED globally for hermeticity.
    Integration tests re-enable it per test body (the autouse approach is
    overridden by the client fixture's later setenv). Unit tests bypass the
    client fixture, so the module default (enabled=True) applies."""

    def test_unit_known_breached_hash_matches(self):
        # SHA-1("password") = 5BAA61E4C9B93F3F0682250B6CF8331B7EE68FD8
        # prefix 5BAA6, suffix 1E4C9B93F3F0682250B6CF8331B7EE68FD8
        from app import password_breach
        body = "1E4C9B93F3F0682250B6CF8331B7EE68FD8:3861493\n"
        with mock.patch.object(password_breach, "_fetch_range", return_value=body):
            assert password_breach.is_password_breached("password") is True

    def test_unit_unknown_password_passes(self):
        from app import password_breach
        with mock.patch.object(password_breach, "_fetch_range", return_value=""):
            assert password_breach.is_password_breached("rare-uniq-pwd-9") is False

    def test_unit_padding_count_zero_treated_as_safe(self):
        from app import password_breach
        # COUNT=0 entries are anti-correlation padding from the HIBP API.
        sha1_prefix_match = "1E4C9B93F3F0682250B6CF8331B7EE68FD8"
        body = f"{sha1_prefix_match}:0\n"
        with mock.patch.object(password_breach, "_fetch_range", return_value=body):
            assert password_breach.is_password_breached("password") is False

    def test_unit_fail_open_on_network_error(self):
        from app import password_breach
        with mock.patch.object(password_breach, "_fetch_range", return_value=None):
            assert password_breach.is_password_breached("anything-9") is False

    def test_unit_disabled_short_circuits(self, monkeypatch):
        from app import password_breach
        monkeypatch.setenv("PASSWORD_BREACH_CHECK_ENABLED", "false")
        # Returns False before reaching _fetch_range; no mock needed.
        assert password_breach.is_password_breached("password") is False

    @staticmethod
    def _force_match(pw: str) -> str:
        """Return a HIBP /range/ response body whose suffix matches the
        SHA-1 of ``pw``. The mock returns this body unconditionally."""
        import hashlib
        digest = hashlib.sha1(pw.encode("utf-8")).hexdigest().upper()  # NOSONAR
        return f"{digest[5:]}:9999\n"

    def test_unit_changeme_allowlisted_despite_breach_match(self):
        """'changeme' is whitelisted case-insensitively (mirrors
        schemas._check_password_strength) and must short-circuit before any
        network call, even though it appears in the breach corpus."""
        from app import password_breach
        body = self._force_match("changeme")
        with mock.patch.object(password_breach, "_fetch_range", return_value=body) as m:
            assert password_breach.is_password_breached("changeme") is False
            assert password_breach.is_password_breached("CHANGEME") is False
        m.assert_not_called()

    def test_change_password_rejects_breached(self, admin_client, monkeypatch):
        monkeypatch.setenv("PASSWORD_BREACH_CHECK_ENABLED", "true")
        from app import password_breach
        # GoodFresh123 passes local strength rules, but the mocked HIBP
        # endpoint says it's breached.
        new_pw = "GoodFresh123"
        with mock.patch.object(
            password_breach, "_fetch_range",
            return_value=self._force_match(new_pw),
        ):
            res = admin_client.post("/api/auth/change-password", json={
                "current_password": "Admin1234",
                "new_password": new_pw,
            })
        assert res.status_code == 400, res.text
        assert "breach" in res.json()["detail"].lower()

    def test_create_user_rejects_breached(self, admin_client, monkeypatch):
        monkeypatch.setenv("PASSWORD_BREACH_CHECK_ENABLED", "true")
        from app import password_breach
        new_pw = "GoodFresh123"
        with mock.patch.object(
            password_breach, "_fetch_range",
            return_value=self._force_match(new_pw),
        ):
            res = admin_client.post("/api/users", json={
                "name": "Tester", "email": "tester@test.local",
                "role": "user", "password": new_pw,
            })
        assert res.status_code == 400, res.text
        assert "breach" in res.json()["detail"].lower()

    def test_change_password_accepts_safe_new_password(self, admin_client, monkeypatch):
        monkeypatch.setenv("PASSWORD_BREACH_CHECK_ENABLED", "true")
        from app import password_breach
        with mock.patch.object(password_breach, "_fetch_range", return_value=""):
            res = admin_client.post("/api/auth/change-password", json={
                "current_password": "Admin1234",
                "new_password": "FreshSafe123",
            })
        assert res.status_code == 204


# ---------------------------------------------------------------------------
# EXIF strip
# ---------------------------------------------------------------------------
def _jpeg_with_exif(gps_value: str = "secret-gps-tag") -> bytes:
    """Return an 8x8 JPEG with ``gps_value`` stored in EXIF tag 270
    (ImageDescription). Uses Pillow's Image.Exif directly (no extras)."""
    from PIL import Image

    img = Image.new("RGB", (8, 8), (200, 100, 50))
    exif = img.getexif()
    exif[270] = gps_value
    out = io.BytesIO()
    img.save(out, format="JPEG", exif=exif.tobytes())
    return out.getvalue()


def _png_with_text(text: str = "stash-this") -> bytes:
    """Build a PNG carrying a tEXt chunk to grep for afterwards."""
    from PIL import Image
    from PIL.PngImagePlugin import PngInfo

    img = Image.new("RGB", (8, 8), (10, 20, 30))
    meta = PngInfo()
    meta.add_text("Source", text)
    out = io.BytesIO()
    img.save(out, format="PNG", pnginfo=meta)
    return out.getvalue()


class TestExifStrip:

    def test_unit_jpeg_metadata_removed(self):
        from app.image_strip import strip_image_metadata
        marker = "GPS-LEAK-XY-ZZ"
        raw = _jpeg_with_exif(marker)
        assert marker.encode() in raw, "test fixture should embed the marker"
        clean = strip_image_metadata(raw, "image/jpeg")
        assert marker.encode() not in clean, (
            "EXIF strip failed: marker still present in output bytes"
        )

    def test_unit_png_metadata_removed(self):
        from app.image_strip import strip_image_metadata
        marker = "PNG-TEXT-LEAK-XY"
        raw = _png_with_text(marker)
        assert marker.encode() in raw
        clean = strip_image_metadata(raw, "image/png")
        assert marker.encode() not in clean

    def test_unit_non_image_passes_through(self):
        from app.image_strip import strip_image_metadata
        raw = b"%PDF-1.7\n%fake pdf bytes"
        assert strip_image_metadata(raw, "application/pdf") == raw

    def test_unit_unknown_content_type_passes_through(self):
        from app.image_strip import strip_image_metadata
        raw = b"<svg></svg>"
        assert strip_image_metadata(raw, "image/svg+xml") == raw

    def test_unit_garbage_image_passes_through(self):
        from app.image_strip import strip_image_metadata
        raw = b"this is definitely not a jpeg"
        assert strip_image_metadata(raw, "image/jpeg") == raw

    def test_unit_empty_bytes_returns_empty(self):
        from app.image_strip import strip_image_metadata
        assert strip_image_metadata(b"", "image/jpeg") == b""

    def test_http_upload_strips_jpeg_exif(self, admin_client):
        bug_id = _make_bug(admin_client, "EXIF carrier upload")
        marker = "FIELD-MARKER-9X"
        raw = _jpeg_with_exif(marker)
        assert marker.encode() in raw
        res = admin_client.post(
            f"/api/bugs/{bug_id}/attachments",
            files={"file": ("evidence.jpg", raw, "image/jpeg")},
        )
        assert res.status_code == 201, res.text
        att_id = res.json()["id"]
        dl = admin_client.get(f"/api/bugs/{bug_id}/attachments/{att_id}/download")
        assert dl.status_code == 200
        assert marker.encode() not in dl.content, (
            "EXIF marker survived the upload roundtrip"
        )
