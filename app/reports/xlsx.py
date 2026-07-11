"""XLSX writer for reports: Summary sheet, optional Items drill-down, Filters
Applied sheet. Used by routes/reports.py and chatbot/excel.py."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:   # pragma: no cover — broken installs only
    OPENPYXL_AVAILABLE = False

from app.reports.engine import ReportResult, ReportColumn


class XlsxBuildError(Exception):
    """Raised when the workbook can't be assembled (openpyxl missing)."""


_HEADER_FILL = "1F2A44"   # Bug Hunter dark accent
_HEADER_FG = "FFFFFF"
_BANNER_FILL = "C9764F"   # Bug Hunter orange
_BANNER_FG = "FFFFFF"
_ZEBRA_FILL = "F2F4F8"


# Formula-injection guard: a leading char here makes Excel treat the cell as
# a formula (e.g. `=cmd|'/c calc'!A1`). OWASP fix: prefix with a quote.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r", "\n")


def _defang_formula_text(s: str) -> str:
    """Prefix with a quote if s starts with a formula trigger (checks both
    the raw and whitespace-stripped first char — Excel trims leading spaces)."""
    if not s:
        return s
    stripped = s.lstrip()
    if s[0] in _FORMULA_TRIGGERS or (stripped and stripped[0] in _FORMULA_TRIGGERS):
        return "'" + s
    return s


def _ensure_openpyxl() -> None:
    if not OPENPYXL_AVAILABLE:
        raise XlsxBuildError(
            "openpyxl is not installed on this server. "
            "Add it to requirements.txt and redeploy."
        )


def _coerce(value: Any) -> Any:
    """Convert a row value to something openpyxl accepts. None becomes '',
    datetimes become ISO strings, and all strings are passed through
    _defang_formula_text to block formula injection."""
    if value is None:
        return ""
    if isinstance(value, bool):
        # bool is a subclass of int — keep this branch above the int one.
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        return _defang_formula_text(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    return _defang_formula_text(str(value))


def _write_table(
    ws,
    columns: list[ReportColumn],
    rows: list[dict[str, Any]],
    *,
    banner: str,
    start_row: int = 1,
) -> int:
    """Write a banner + header + rows. Returns the next free row index."""
    ncols = max(1, len(columns))

    # Defang banner/headers too, even though they're server-controlled today.
    banner_cell = ws.cell(row=start_row, column=1, value=_defang_formula_text(banner))
    banner_cell.font = Font(bold=True, color=_BANNER_FG, size=12)
    banner_cell.fill = PatternFill("solid", fgColor=_BANNER_FILL)
    banner_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=ncols,
    )
    ws.row_dimensions[start_row].height = 22

    # Header row.
    header_row = start_row + 1
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(bold=True, color=_HEADER_FG)
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=idx, value=_defang_formula_text(col.label))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = max(8, col.width)

    # Data rows.
    align_by_col = [
        Alignment(horizontal=col.align or "left", vertical="top", wrap_text=False)
        for col in columns
    ]
    for r_idx, row in enumerate(rows):
        sheet_row = header_row + 1 + r_idx
        zebra = (r_idx % 2 == 1)
        for c_idx, col in enumerate(columns, start=1):
            v = _coerce(row.get(col.key, ""))
            cell = ws.cell(row=sheet_row, column=c_idx, value=v)
            cell.alignment = align_by_col[c_idx - 1]
            if zebra:
                cell.fill = PatternFill("solid", fgColor=_ZEBRA_FILL)

    # Freeze banner + header so they don't scroll.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return header_row + 1 + len(rows)


_FILTER_LABELS = [
    ("label", "Run Label"),
    ("date_from", "Date From"),
    ("date_to", "Date To"),
    ("item_types", "Item Types"),
    ("statuses", "Statuses"),
    ("priorities", "Priorities"),
    ("environments", "Environments"),
    ("project_ids", "Project IDs"),
    ("assignee_ids", "Assignee IDs"),
    ("reporter_ids", "Reporter IDs"),
    ("event_id", "Event ID"),
    ("include_not_a_bug", "Include 'Not a Bug'"),
    ("text_search", "Text Search"),
]


def _format_filter_display(v: Any) -> str:
    """Render a Filters.to_meta() value for the Filters Applied sheet."""
    if v is None or v == "" or v == [] or v == {}:
        return "(not set)"
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return str(v)


def _format_summary_display(v: Any) -> Any:
    """Render a summary entry for the Filters Applied sheet."""
    if isinstance(v, dict):
        return ", ".join(f"{k}: {vv}" for k, vv in v.items())
    return v


def _write_filters_block(ws, result: ReportResult, start_row: int) -> int:
    """Write the Filter/Value rows. Returns the row index after the block."""
    row = start_row
    ws.cell(row=row, column=1, value="Filter").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
    row += 1
    f = result.filters or {}
    for key, label in _FILTER_LABELS:
        ws.cell(row=row, column=1, value=label)
        # text_search/label are user free-text — defang like data cells.
        ws.cell(row=row, column=2, value=_coerce(_format_filter_display(f.get(key))))
        row += 1
    return row


def _write_summary_block(ws, result: ReportResult, start_row: int) -> int:
    if not result.summary:
        return start_row
    row = start_row + 1   # blank separator row
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
    row += 1
    for k, v in result.summary.items():
        ws.cell(row=row, column=1, value=str(k).replace("_", " ").title())
        ws.cell(row=row, column=2, value=_coerce(_format_summary_display(v)))
        row += 1
    return row


def _write_filters_sheet(ws, result: ReportResult) -> None:
    ws.title = "Filters Applied"
    ws.cell(row=1, column=1, value=f"Bug Hunter — {result.report_label}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).isoformat(timespec='seconds')}")
    ws.cell(row=3, column=1, value=f"Total rows: {result.total}")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50

    row = _write_filters_block(ws, result, start_row=5)
    _write_summary_block(ws, result, start_row=row)


def build_workbook_bytes(result: ReportResult) -> bytes:
    """Serialise a ReportResult to xlsx bytes. Three sheets at most."""
    _ensure_openpyxl()
    wb = Workbook()
    main_ws = wb.active
    main_ws.title = (result.report_label or "Report")[:31] or "Report"
    banner = f"{result.report_label} — {result.total} row{'' if result.total == 1 else 's'}"
    if result.filters and result.filters.get("label"):
        banner = f"{banner} · {result.filters['label']}"
    _write_table(main_ws, result.columns, result.rows, banner=banner)

    if result.detail_rows:
        detail_ws = wb.create_sheet("Items")
        det_banner = f"{result.report_label} — Items ({len(result.detail_rows)} row{'' if len(result.detail_rows) == 1 else 's'})"
        _write_table(detail_ws, result.detail_columns, result.detail_rows, banner=det_banner)

    filters_ws = wb.create_sheet("Filters Applied")
    _write_filters_sheet(filters_ws, result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = [
    "XlsxBuildError",
    "build_workbook_bytes",
    "_defang_formula_text",
    "OPENPYXL_AVAILABLE",
]
