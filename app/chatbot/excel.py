"""Sleuth Excel writer: builds xlsx in memory, staged under a short-lived
token for /api/chat/download/{token}. Cache is TTL- and size-bounded,
guarded by a threading.Lock.
"""
from __future__ import annotations

import io
import secrets
import threading
import time
from typing import Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - exercised only on broken installs
    OPENPYXL_AVAILABLE = False


class ExcelGenerationError(Exception):
    """Raised when the workbook can't be built (e.g. openpyxl missing)."""


_TTL_SECONDS = 30 * 60
_MAX_ENTRIES = 50  # hard cap; oldest evicted when exceeded


_cache_lock = threading.Lock()
_cache: dict[str, tuple[bytes, str, float, int]] = {}
# token -> (xlsx_bytes, filename, expires_at_epoch, owner_user_id)


def _evict_expired_locked(now: float) -> None:
    """Drop expired entries. Caller must hold _cache_lock."""
    dead = [tok for tok, (_, _, exp, _owner) in _cache.items() if exp <= now]
    for tok in dead:
        _cache.pop(tok, None)


def _evict_oldest_locked() -> None:
    """Drop the soonest-expiring entry when over the cap. Caller holds the lock."""
    if len(_cache) < _MAX_ENTRIES:
        return
    oldest = min(_cache.items(), key=lambda kv: kv[1][2])
    _cache.pop(oldest[0], None)


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
_HEADER_STYLE_FILL = "1F2A44"   # Bug Hunter dark accent
_HEADER_STYLE_FG = "FFFFFF"


# Formula-injection defense: leading trigger chars get a quote prefix,
# same as routes/bugs.py CSV export and reports/xlsx.py.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r", "\n")


def _defang_formula_text(s: str) -> str:
    # check raw AND lstripped first char — apps trim leading spaces, " =cmd" still fires
    if not s:
        return s
    stripped = s.lstrip()
    if s[0] in _FORMULA_TRIGGERS or (stripped and stripped[0] in _FORMULA_TRIGGERS):
        return "'" + s
    return s


# Column order matches executor._bug_row.
_COLUMNS: list[tuple[str, str, int]] = [
    ("id",          "ID",          8),
    ("title",       "Title",       50),
    ("project",     "Project",     20),
    ("status",      "Status",      14),
    ("priority",    "Priority",    12),
    ("environment", "Env",         8),
    ("reporter",    "Reporter",    24),
    ("assignees",   "Assignees",   40),
    ("due_date",    "Due Date",    12),
    ("created_at",  "Created",     22),
    ("updated_at",  "Updated",     22),
]


def _build_workbook(rows: list[dict[str, Any]], description: str) -> bytes:
    if not OPENPYXL_AVAILABLE:
        # user-facing message — don't leak dependency name / server internals
        raise ExcelGenerationError(
            "The spreadsheet exporter is unavailable on this server right now."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Bugs"

    # banner row with the filter description
    banner = f"Bug Hunter export — {description}" if description else "Bug Hunter export"
    ws.cell(row=1, column=1, value=banner).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))

    # Header row.
    header_fill = PatternFill("solid", fgColor=_HEADER_STYLE_FILL)
    header_font = Font(bold=True, color=_HEADER_STYLE_FG)
    for idx, (_key, header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Data rows.
    for r, row in enumerate(rows, start=3):
        for c, (key, _h, _w) in enumerate(_COLUMNS, start=1):
            val = row.get(key, "")
            if isinstance(val, str):
                val = _defang_formula_text(val)
            # openpyxl chokes on None for some types; coerce to ""
            ws.cell(row=r, column=c, value="" if val is None else val)

    # freeze banner + header rows
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def stage_workbook(rows: list[dict[str, Any]], filename: str,
                   owner_id: int, description: str = "") -> tuple[str, int]:
    """Build the workbook and stage it under a fresh token bound to ``owner_id``.
    Returns (token, size_bytes); raises ExcelGenerationError on failure."""
    payload = _build_workbook(rows, description)
    return stage_bytes(payload, filename, owner_id)


def stage_bytes(payload: bytes, filename: str, owner_id: int) -> tuple[str, int]:
    """Stage pre-built xlsx bytes under a fresh owner-bound token. Returns (token, size)."""
    token = secrets.token_urlsafe(20)
    expires = time.time() + _TTL_SECONDS
    with _cache_lock:
        _evict_expired_locked(time.time())
        _evict_oldest_locked()
        _cache[token] = (payload, filename, expires, owner_id)
    return token, len(payload)


def fetch_staged(token: str, owner_id: int) -> Optional[tuple[bytes, str]]:
    """Return (bytes, filename) if token is valid AND owned by ``owner_id``, else None.

    Owner check makes the token a per-user capability; mismatch returns None so
    the router 404s and token existence can't be probed.
    """
    if not token:
        return None
    now = time.time()
    with _cache_lock:
        _evict_expired_locked(now)
        entry = _cache.get(token)
        if entry is None:
            return None
        payload, filename, expires, entry_owner = entry
        if expires <= now:
            _cache.pop(token, None)
            return None
        if entry_owner != owner_id:
            return None
        return payload, filename


def clear_all_for_test() -> None:
    """Hook for tests; never called from production code paths."""
    with _cache_lock:
        _cache.clear()


__all__ = [
    "ExcelGenerationError",
    "stage_workbook",
    "stage_bytes",
    "fetch_staged",
    "OPENPYXL_AVAILABLE",
]
