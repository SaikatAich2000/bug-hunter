"""Security and correctness tests for known behavioural gaps.

Covers: session lifecycle on password change/reset, stale reset tokens,
update_bug partial-write risk, activity ordering consistency, frontend
API contract, input validation, CSRF posture, and transactional semantics.
"""
from __future__ import annotations

import io


def _admin_login(c):
    r = c.post("/api/auth/login", json={
        "email": "admin@test.local", "password": "Admin1234",
    })
    assert r.status_code == 200, r.text


def _make_bug(c, project_id, title="A bug for tests"):
    r = c.post("/api/bugs", json={
        "project_id": project_id, "title": title,
        "priority": "Medium", "environment": "DEV",
    })
    assert r.status_code == 201, r.text
    return r.json()


def _make_project(c, name="P"):
    r = c.post("/api/projects", json={"name": name})
    assert r.status_code == 201, r.text
    return r.json()


def _make_user(c, name="Someone", email="some@x.com", role="user", password="TestUserPwd9X"):
    body = {"name": name, "email": email, "role": role, "password": password}
    # Tag the new user to every existing project so these tests keep working
    # under project-scoped access without needing per-test wiring.
    pids = [p["id"] for p in c.get("/api/projects").json()]
    if pids:
        body["project_ids"] = pids
    r = c.post("/api/users", json=body)
    assert r.status_code == 201, r.text
    return r.json()


# Sessions survive a password change/reset. If an account is compromised,
# the owner changing the password won't evict an attacker's active session.
class TestSessionLifecycle:
    def test_existing_session_still_works_after_password_change(self, admin_client):
        """Changing the password does not invalidate existing sessions, so a
        compromised account stays compromised even after the owner resets."""
        # Confirm session works before the change.
        r = admin_client.get("/api/auth/me")
        assert r.status_code == 200

        r = admin_client.post("/api/auth/change-password", json={
            "current_password": "Admin1234",
            "new_password": "NewBetter999",
        })
        assert r.status_code == 204

        # Old session cookie still works — this is the current (gap) behaviour.
        r = admin_client.get("/api/auth/me")
        assert r.status_code == 200, \
            "BUG: old session is invalidated after password change (good!) — update test"

    def test_outstanding_reset_tokens_remain_after_password_change(self, admin_client):
        """A reset link issued before a regular password change stays valid
        and can override the new password. Anyone who intercepts the link
        can take over the account.

        We can't retrieve the token (it's email-only), so this just confirms
        the system allows multiple concurrent outstanding tokens to exist."""
        # Three consecutive reset requests; all should succeed and each
        # should produce an audit row.
        for _ in range(3):
            r = admin_client.post("/api/auth/forgot-password", json={
                "email": "admin@test.local",
            })
            assert r.status_code == 204
        r = admin_client.get("/api/audit?q=password_reset_requested")
        rows = r.json()
        n = sum(1 for r in rows if r["action"] == "password_reset_requested")
        assert n >= 3, \
            f"BUG: forgot-password doesn't dedupe — {n} concurrent tokens issued"


# update_bug mutates the ORM object before the role check runs, risking a
# partial write if the session is committed in the wrong order.
class TestUpdateBugOrdering:
    def test_failed_reporter_change_does_not_persist_other_changes(self, admin_client):
        """A non-admin PUT that mixes a legitimate field change with an
        unauthorized reporter_id change must be rejected atomically. The
        title change must not persist."""
        p = _make_project(admin_client, "U1")
        owner = _make_user(admin_client, name="Owner", email="owner@x.com",
                           password="TestUserPwd9X")
        other = _make_user(admin_client, name="Other", email="other@x.com",
                           password="TestUserPwd9X")
        # Admin creates the bug and assigns owner as reporter so owner can edit.
        bug = _make_bug(admin_client, p["id"], title="Original title here")
        admin_client.put(f"/api/bugs/{bug['id']}", json={"reporter_id": owner["id"]})

        # Switch to the owner session.
        admin_client.post("/api/auth/logout")
        admin_client.post("/api/auth/login", json={
            "email": "owner@x.com", "password": "TestUserPwd9X",
        })

        # Combine a valid title change with an unauthorized reporter change.
        # The whole request must be rejected (403) with no partial write.
        r = admin_client.put(f"/api/bugs/{bug['id']}", json={
            "title": "Hacked title here",
            "reporter_id": other["id"],
        })
        assert r.status_code == 403

        r = admin_client.get(f"/api/bugs/{bug['id']}")
        assert r.json()["title"] == "Original title here", \
            "BUG: title was persisted despite the 403 from the role check"


# Bug.activities has no secondary sort key, so same-second events can appear
# in different orders depending on which endpoint fetches them.
class TestActivityOrdering:
    def test_activity_order_consistent_between_get_bug_and_list_activity(self, admin_client):
        """Activities with the same timestamp must appear in the same order
        from both the bug detail and the activity list endpoints."""
        p = _make_project(admin_client, "AO1")
        bug = _make_bug(admin_client, p["id"], title="Activity ordering test")
        # Rapid updates to increase the chance of timestamp collisions.
        for status in ("In Progress", "Resolved", "Reopened", "Closed"):
            admin_client.put(f"/api/bugs/{bug['id']}", json={"status": status})

        detail = admin_client.get(f"/api/bugs/{bug['id']}").json()
        activity = admin_client.get(f"/api/bugs/{bug['id']}/activity").json()

        detail_ids = [a["id"] for a in detail["activities"]]
        activity_ids = [a["id"] for a in activity]
        assert detail_ids == activity_ids, \
            f"BUG: activity order differs between endpoints.\n" \
            f"  detail:   {detail_ids}\n  activity: {activity_ids}"


# Frontend contract tests — exercised through the API responses the SPA consumes.
class TestFrontendContract:
    def test_bug_detail_attachments_only_includes_bug_level(self, admin_client):
        """Top-level 'attachments' in a BugDetail must only contain
        bug-level files (comment_id NULL). Comment attachments live under
        their comment, and attachment_count reflects both."""
        p = _make_project(admin_client, "FE1")
        bug = _make_bug(admin_client, p["id"])
        admin_client.post(f"/api/bugs/{bug['id']}/attachments",
                          files={"file": ("bug.txt", io.BytesIO(b"a"), "text/plain")})
        cr = admin_client.post(f"/api/bugs/{bug['id']}/comments", json={"body": "x"})
        cid = cr.json()["id"]
        admin_client.post(f"/api/bugs/{bug['id']}/attachments",
                          files={"file": ("c.txt", io.BytesIO(b"b"), "text/plain")},
                          data={"comment_id": str(cid)})

        d = admin_client.get(f"/api/bugs/{bug['id']}").json()
        bug_level_filenames = {a["filename"] for a in d["attachments"]}
        assert bug_level_filenames == {"bug.txt"}, bug_level_filenames
        comment_atts = d["comments"][0]["attachments"]
        assert {a["filename"] for a in comment_atts} == {"c.txt"}
        assert d["attachment_count"] == 2

    def test_xlsx_export_preserves_commas_and_newlines_in_description(self, admin_client):
        """Commas and newlines in a description must survive the openpyxl
        round-trip as a single cell. This is the XLSX equivalent of what
        the old CSV export test checked."""
        import io
        from openpyxl import load_workbook
        p = _make_project(admin_client, "FE2")
        admin_client.post("/api/bugs", json={
            "project_id": p["id"],
            "title": "row, has, commas",
            "description": "line1\nline2,still line2",
            "priority": "Low", "environment": "DEV",
        })
        r = admin_client.post("/api/reports/export.xlsx", json={
            "report_key": "item_detail", "filters": {},
        })
        assert r.status_code == 200, r.text
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        text = " ".join(
            str(v) for row in wb[wb.sheetnames[0]].iter_rows(values_only=True)
            for v in row if v is not None
        )
        assert "line1" in text and "still line2" in text
        assert "row, has, commas" in text


# Input validation — boundary checks on exposed fields.
class TestInputValidation:
    def test_user_create_with_extreme_long_email(self, admin_client):
        """email max=254. Pass 300 chars — must reject."""
        r = admin_client.post("/api/users", json={
            "name": "longmail", "email": "a" * 250 + "@x.com",
            "role": "user", "password": "TestUserPwd9X",
        })
        assert r.status_code == 422

    def test_bug_create_with_oversized_description(self, admin_client):
        # Cap is 1 MB to allow rich HTML with multiple base64 screenshots.
        # Anything over the cap should be 422.
        p = _make_project(admin_client, "IV1")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "x" * 10,
            "description": "a" * 1_000_001,
            "priority": "Low", "environment": "DEV",
        })
        assert r.status_code == 422

    def test_comment_with_oversized_body(self, admin_client):
        # 200 KB cap (room for pasted screenshots). Anything over is 422.
        p = _make_project(admin_client, "IV2")
        bug = _make_bug(admin_client, p["id"])
        r = admin_client.post(f"/api/bugs/{bug['id']}/comments",
                              json={"body": "x" * 200_001})
        assert r.status_code == 422

    def test_audit_limit_clamped_at_10000(self, admin_client):
        # Ceiling of 10 000 lets operators pull the full trail; above that is 422.
        r = admin_client.get("/api/audit?limit=10001")
        assert r.status_code == 422
        r = admin_client.get("/api/audit?limit=10000")
        assert r.status_code == 200

    def test_user_role_normalization_upper_to_lower(self, admin_client):
        """Schema lowercases role: 'ADMIN' should become 'admin'."""
        r = admin_client.post("/api/users", json={
            "name": "rolecase", "email": "rc@x.com", "role": "ADMIN",
            "password": "TestUserPwd9X",
        })
        assert r.status_code == 201, r.text
        assert r.json()["role"] == "admin"


# CSRF posture. True cross-site simulation isn't possible in a unit test,
# so these verify cookie attributes and that no token mechanism is expected.
class TestCSRFPosture:
    def test_no_csrf_token_on_state_changing_routes(self, admin_client):
        """Bug Hunter relies on SameSite=Lax rather than CSRF tokens.
        A normal authenticated POST with no extra header must succeed."""
        p = _make_project(admin_client, "CSRF1")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "no csrf token here",
            "priority": "Low", "environment": "DEV",
        })
        assert r.status_code == 201

    def test_options_preflight_handled(self, admin_client):
        """Preflight OPTIONS must return a deterministic response.

        CORS middleware is only registered when CORS_ORIGINS is explicitly
        configured. With no configured origins (the default), preflight hits
        the route directly — any non-5xx response is accepted, just no
        surprises like 500s.
        """
        r = admin_client.options("/api/bugs", headers={
            "Origin": "https://example.com",
            "Access-Control-Request-Method": "POST",
            "Access-Control-Request-Headers": "Content-Type",
        })
        # 200/204: CORS middleware replied. 400: origin rejected. 401/405: no CORS middleware.
        assert r.status_code in (200, 204, 400, 401, 405), (
            f"unexpected: {r.status_code}"
        )


# Transactional / idempotency behaviour.
class TestTransactional:
    def test_bug_update_no_change_does_not_log_activity(self, admin_client):
        """A PUT that changes nothing must not create an activity row."""
        p = _make_project(admin_client, "TX1")
        bug = _make_bug(admin_client, p["id"], title="No-change update test")
        n0 = len(admin_client.get(f"/api/bugs/{bug['id']}/activity").json())
        # Send back the same status the bug already has.
        r = admin_client.put(f"/api/bugs/{bug['id']}", json={"status": "New"})
        assert r.status_code == 200
        n1 = len(admin_client.get(f"/api/bugs/{bug['id']}/activity").json())
        assert n0 == n1, f"BUG: no-change update added an activity row ({n0} → {n1})"

    def test_bug_assignee_set_to_same_value_does_not_log_change(self, admin_client):
        """Re-submitting the same assignee list shouldn't log a fake 'assignees changed' event."""
        p = _make_project(admin_client, "TX2")
        u = _make_user(admin_client, name="Assigned1", email="ass1@x.com")
        bug = _make_bug(admin_client, p["id"], title="Assignee no-change test")
        admin_client.put(f"/api/bugs/{bug['id']}", json={"assignee_ids": [u["id"]]})
        n0 = len(admin_client.get(f"/api/bugs/{bug['id']}/activity").json())
        admin_client.put(f"/api/bugs/{bug['id']}", json={"assignee_ids": [u["id"]]})
        n1 = len(admin_client.get(f"/api/bugs/{bug['id']}/activity").json())
        assert n0 == n1, f"BUG: same-assignee update created a phantom activity row"
