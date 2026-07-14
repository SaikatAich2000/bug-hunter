"""Regression tests: auth, CRUD, comments, attachments, audit, stats, security, edge cases."""
from __future__ import annotations

import io
import time


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _login_admin(c):
    r = c.post("/api/auth/login", json={
        "email": "admin@test.local", "password": "Admin1234",
    })
    assert r.status_code == 200, r.text


def _login_as(c, email, password):
    r = c.post("/api/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, r.text


def _logout(c):
    c.post("/api/auth/logout")


def _create_user(c, name, email, role="user", password="TestUserPwd9X", is_active=True):
    # Default password satisfies server-side strength requirements.
    if len(name) < 2:                # server requires name>=2
        name = name + "_user"
    body = {"name": name, "email": email, "role": role,
            "password": password, "is_active": is_active}
    # Assign to all projects so non-admins can see every bug in these tests.
    pids = [p["id"] for p in c.get("/api/projects").json()]
    if pids:
        body["project_ids"] = pids
    r = c.post("/api/users", json=body)
    assert r.status_code == 201, r.text
    return r.json()


def _create_project(c, name="P1", color="#c9764f"):
    r = c.post("/api/projects", json={"name": name, "color": color})
    assert r.status_code == 201, r.text
    return r.json()


def _create_bug(c, project_id, title="A Bug Title", **extra):
    if len(title) < 3:               # server requires title>=3
        title = title + "_xx"
    body = {"project_id": project_id, "title": title, "priority": "Medium",
            "environment": "DEV"}
    body.update(extra)
    r = c.post("/api/bugs", json=body)
    assert r.status_code == 201, r.text
    return r.json()


class TestAuth:
    def test_logout_when_already_logged_out_is_204(self, client):
        """Logout from a fresh client (no cookie) must still return 204."""
        r = client.post("/api/auth/logout")
        assert r.status_code == 204

    def test_login_with_uppercase_email_succeeds(self, client):
        """Email case-insensitivity: login with mixed case email."""
        r = client.post("/api/auth/login", json={
            "email": "Admin@Test.LOCAL", "password": "Admin1234",
        })
        assert r.status_code == 200, r.text

    def test_login_with_inactive_user_is_unified_401(self, admin_client):
        """Deactivated user gets the same 401 as a wrong password (no account-state leak)."""
        u = _create_user(admin_client, "Deact", "deact@x.com",
                         password="DeactivatedZx9Q", is_active=False)
        _logout(admin_client)
        r = admin_client.post("/api/auth/login", json={
            "email": "deact@x.com", "password": "DeactivatedZx9Q",
        })
        assert r.status_code == 401, f"got {r.status_code}: {r.text}"
        # Same message as wrong-password so disabled accounts aren't enumerable.
        assert r.json()["detail"] == "Invalid email or password"

    def test_session_cookie_is_httponly(self, client):
        """Auth cookie must be HttpOnly to prevent XSS theft."""
        r = client.post("/api/auth/login", json={
            "email": "admin@test.local", "password": "Admin1234",
        })
        assert r.status_code == 200
        set_cookies = r.headers.get_list("set-cookie")
        bh_cookie = next((c for c in set_cookies if c.startswith("bh_session=")), None)
        assert bh_cookie is not None
        assert "HttpOnly" in bh_cookie, f"cookie missing HttpOnly: {bh_cookie}"

    def test_session_cookie_has_samesite_lax(self, client):
        """Auth cookie must be SameSite=Lax for CSRF defence."""
        r = client.post("/api/auth/login", json={
            "email": "admin@test.local", "password": "Admin1234",
        })
        set_cookies = r.headers.get_list("set-cookie")
        bh_cookie = next((c for c in set_cookies if c.startswith("bh_session=")), None)
        assert bh_cookie and "samesite=lax" in bh_cookie.lower()

    def test_change_password_with_short_new_password_fails(self, admin_client):
        """New password must be ≥ 8 chars."""
        r = admin_client.post("/api/auth/change-password", json={
            "current_password": "Admin1234", "new_password": "Short1",
        })
        assert r.status_code == 422, r.text

    def test_change_password_then_old_session_still_valid(self, admin_client):
        """After a password change, the current session must stay valid (no forced logout)."""
        r = admin_client.post("/api/auth/change-password", json={
            "current_password": "Admin1234", "new_password": "Newpass789",
        })
        assert r.status_code == 204
        r = admin_client.get("/api/auth/me")
        assert r.status_code == 200

    def test_login_with_extremely_long_password(self, client):
        """A 1000-char password must not crash bcrypt; the sha256 prehash prevents that."""
        r = client.post("/api/auth/login", json={
            "email": "admin@test.local", "password": "x" * 1000,
        })
        assert r.status_code in (401, 422)  # rejected, but not 500

    def test_invalid_email_format_login(self, client):
        r = client.post("/api/auth/login", json={
            "email": "notanemail", "password": "Whatever1",
        })
        assert r.status_code == 422

    def test_unauthed_change_password_is_401(self, client):
        r = client.post("/api/auth/change-password", json={
            "current_password": "Admin1234", "new_password": "Newpass1234",
        })
        assert r.status_code == 401

    def test_reset_password_with_invalid_token(self, client):
        r = client.post("/api/auth/reset-password", json={
            "token": "definitely-not-a-real-token", "new_password": "Newpass789",
        })
        assert r.status_code == 400


class TestUsers:
    def test_create_user_with_duplicate_email_is_409(self, admin_client):
        _create_user(admin_client, "U1", "dup@x.com")
        r = admin_client.post("/api/users", json={
            "name": "U2", "email": "dup@x.com", "role": "user",
            "password": "TestUserPwd9X",
        })
        assert r.status_code == 409, r.text

    def test_create_user_with_invalid_role(self, admin_client):
        r = admin_client.post("/api/users", json={
            "name": "Bad", "email": "bad@x.com", "role": "superadmin",
            "password": "TestUserPwd9X",
        })
        assert r.status_code == 422

    def test_create_user_short_password(self, admin_client):
        r = admin_client.post("/api/users", json={
            "name": "U", "email": "u@x.com", "role": "user", "password": "abc",
        })
        assert r.status_code == 422

    def test_create_user_empty_password(self, admin_client):
        """Empty password must not silently bypass min-length."""
        r = admin_client.post("/api/users", json={
            "name": "U", "email": "u@x.com", "role": "user", "password": "",
        })
        assert r.status_code == 422

    def test_create_user_invalid_email(self, admin_client):
        r = admin_client.post("/api/users", json={
            "name": "U", "email": "no-at-sign", "role": "user",
            "password": "TestUserPwd9X",
        })
        assert r.status_code == 422

    def test_create_user_with_whitespace_only_name(self, admin_client):
        """Whitespace-only name must be rejected."""
        r = admin_client.post("/api/users", json={
            "name": "    ", "email": "ws@x.com", "role": "user",
            "password": "TestUserPwd9X",
        })
        assert r.status_code == 422

    def test_admin_password_reset_via_admin(self, admin_client):
        """Admin updating a user's password should let the user log in with new pw."""
        u = _create_user(admin_client, "Bob", "bob@x.com", password="OldPass123")
        r = admin_client.put(f"/api/users/{u['id']}", json={"password": "BrandNew99"})
        assert r.status_code == 200, r.text
        _logout(admin_client)
        r = admin_client.post("/api/auth/login", json={
            "email": "bob@x.com", "password": "BrandNew99",
        })
        assert r.status_code == 200

    def test_user_emails_normalized_to_lowercase_on_create(self, admin_client):
        """Email is stored lowercased: must allow login with original case."""
        u = _create_user(admin_client, "Mix", "MixCase@X.COM", password="TestUserPwd9X")
        assert u["email"] == "mixcase@x.com", \
            f"expected lowercased email; got {u['email']!r}"

    def test_user_list_search_is_case_insensitive(self, admin_client):
        _create_user(admin_client, "Charlie", "charlie@x.com")
        r = admin_client.get("/api/users?q=CHARLIE")
        assert r.status_code == 200
        names = [u["name"] for u in r.json()]
        assert "Charlie" in names

    def test_regular_user_cannot_update_users(self, user_client):
        """A regular user must not be able to PUT another user."""
        # id=1 is the bootstrap admin
        r = user_client.put("/api/users/1", json={"role": "admin"})
        assert r.status_code == 403

    def test_regular_user_cannot_delete_users(self, user_client):
        r = user_client.delete("/api/users/1")
        assert r.status_code == 403

    def test_get_nonexistent_user_404(self, admin_client):
        r = admin_client.get("/api/users/999999")
        assert r.status_code == 404

    def test_email_is_unique_on_update(self, admin_client):
        """Updating user2's email to user1's email must 409."""
        u1 = _create_user(admin_client, "U1", "one@x.com")
        u2 = _create_user(admin_client, "U2", "two@x.com")
        r = admin_client.put(f"/api/users/{u2['id']}", json={"email": "one@x.com"})
        assert r.status_code == 409, r.text


class TestProjects:
    def test_create_project_with_invalid_color(self, admin_client):
        """Color must match #RRGGBB."""
        r = admin_client.post("/api/projects", json={
            "name": "BadColor", "color": "red",
        })
        assert r.status_code == 422

    def test_create_project_with_3char_hex_color_rejected(self, admin_client):
        """Schema requires #RRGGBB — short form #fff must be rejected."""
        r = admin_client.post("/api/projects", json={
            "name": "ShortHex", "color": "#fff",
        })
        assert r.status_code == 422

    def test_create_project_with_duplicate_name_is_409(self, admin_client):
        _create_project(admin_client, name="Unique")
        r = admin_client.post("/api/projects", json={"name": "Unique"})
        assert r.status_code == 409

    def test_delete_project_with_bugs_is_409(self, admin_client):
        p = _create_project(admin_client, name="HasBugs")
        _create_bug(admin_client, p["id"])
        r = admin_client.delete(f"/api/projects/{p['id']}")
        assert r.status_code == 409
        assert "bug" in r.json()["detail"].lower()

    def test_update_nonexistent_project_404(self, admin_client):
        r = admin_client.put("/api/projects/999999", json={"name": "Whatever"})
        assert r.status_code == 404

    def test_manager_can_create_project(self, admin_client):
        _create_user(admin_client, "Mgr", "mgr@x.com", role="manager",
                     password="TestUserPwd9X")
        _logout(admin_client)
        _login_as(admin_client, "mgr@x.com", "TestUserPwd9X")
        r = admin_client.post("/api/projects", json={"name": "MgrProject"})
        assert r.status_code == 201

    def test_regular_user_cannot_update_project(self, user_client):
        r = user_client.put("/api/projects/1", json={"name": "Hacked"})
        assert r.status_code == 403

    def test_project_name_too_short(self, admin_client):
        """Project name has min_length=2."""
        r = admin_client.post("/api/projects", json={"name": "A"})
        assert r.status_code == 422


class TestBugs:
    def test_create_bug_with_nonexistent_project(self, admin_client):
        r = admin_client.post("/api/bugs", json={
            "project_id": 999999, "title": "Doomed",
            "priority": "Low", "environment": "DEV",
        })
        assert r.status_code == 400

    def test_create_bug_with_invalid_status_normalizes(self, admin_client):
        """Status normalization: 'new' should become 'New'."""
        p = _create_project(admin_client, name="N1")
        bug = _create_bug(admin_client, p["id"], status="new")
        assert bug["status"] == "New", \
            f"expected status normalized to 'New', got {bug['status']!r}"

    def test_create_bug_with_invalid_status(self, admin_client):
        p = _create_project(admin_client, name="N2")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "X",
            "priority": "Medium", "environment": "DEV",
            "status": "Bogus",
        })
        assert r.status_code == 422

    def test_create_bug_with_invalid_environment(self, admin_client):
        p = _create_project(admin_client, name="N3")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "X",
            "priority": "Medium", "environment": "QA",
        })
        assert r.status_code == 422

    def test_create_bug_with_invalid_due_date_format(self, admin_client):
        p = _create_project(admin_client, name="N4")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "X",
            "priority": "Medium", "environment": "DEV",
            "due_date": "31/12/2025",
        })
        assert r.status_code == 422

    def test_create_bug_with_unknown_assignee_id(self, admin_client):
        p = _create_project(admin_client, name="N5")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "Long enough title",
            "priority": "Medium", "environment": "DEV",
            "assignee_ids": [999999],
        })
        assert r.status_code == 400

    def test_assignee_ids_dedup(self, admin_client):
        p = _create_project(admin_client, name="N6")
        u = _create_user(admin_client, "A", "a@x.com")
        bug = _create_bug(admin_client, p["id"], assignee_ids=[u["id"], u["id"], u["id"]])
        assert len(bug["assignees"]) == 1

    def test_regular_user_cant_set_other_user_as_reporter(self, admin_client):
        """Regular users can only file bugs as themselves."""
        p = _create_project(admin_client, name="N7")
        _create_user(admin_client, "Other", "other@x.com")
        _create_user(admin_client, "Reg", "reg@x.com", password="TestUserPwd9X")
        _logout(admin_client)
        _login_as(admin_client, "reg@x.com", "TestUserPwd9X")
        users_resp = admin_client.get("/api/users").json()
        other_id = next(u["id"] for u in users_resp if u["email"] == "other@x.com")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "Sneaky",
            "priority": "Low", "environment": "DEV",
            "reporter_id": other_id,
        })
        assert r.status_code == 403

    def test_regular_user_can_self_report(self, admin_client):
        """Regular user passing their own id as reporter is fine."""
        p = _create_project(admin_client, name="N8")
        u = _create_user(admin_client, "Self", "self@x.com", password="TestUserPwd9X")
        _logout(admin_client)
        _login_as(admin_client, "self@x.com", "TestUserPwd9X")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "Mine",
            "priority": "Low", "environment": "DEV",
            "reporter_id": u["id"],
        })
        assert r.status_code == 201

    def test_get_nonexistent_bug_404(self, admin_client):
        r = admin_client.get("/api/bugs/999999")
        assert r.status_code == 404

    def test_delete_bug_returns_404_after(self, admin_client):
        p = _create_project(admin_client, name="N9")
        bug = _create_bug(admin_client, p["id"])
        admin_client.delete(f"/api/bugs/{bug['id']}")
        r = admin_client.get(f"/api/bugs/{bug['id']}")
        assert r.status_code == 404

    def test_bug_list_search_by_id_with_hash(self, admin_client):
        p = _create_project(admin_client, name="N10")
        bug = _create_bug(admin_client, p["id"], title="Findme")
        r = admin_client.get(f"/api/bugs?q=%23{bug['id']}")  # %23 is '#'
        assert r.status_code == 200
        items = r.json()["items"]
        assert len(items) == 1
        assert items[0]["id"] == bug["id"]

    def test_bug_list_search_by_id_without_hash(self, admin_client):
        p = _create_project(admin_client, name="N11")
        bug = _create_bug(admin_client, p["id"], title="Findme2")
        r = admin_client.get(f"/api/bugs?q={bug['id']}")
        items = r.json()["items"]
        assert any(b["id"] == bug["id"] for b in items)

    def test_bug_list_search_text(self, admin_client):
        p = _create_project(admin_client, name="N12")
        _create_bug(admin_client, p["id"], title="UNIQUE_NEEDLE_string")
        r = admin_client.get("/api/bugs?q=needle")
        items = r.json()["items"]
        assert any("UNIQUE_NEEDLE_string" in b["title"] for b in items)

    def test_bug_list_invalid_pagination(self, admin_client):
        r = admin_client.get("/api/bugs?page=0")
        assert r.status_code == 400
        r = admin_client.get("/api/bugs?page_size=999")
        assert r.status_code == 400

    def test_bug_list_pagination_total_pages(self, admin_client):
        p = _create_project(admin_client, name="N13")
        for i in range(5):
            _create_bug(admin_client, p["id"], title=f"bug{i}")
        r = admin_client.get("/api/bugs?page=1&page_size=2")
        body = r.json()
        assert body["total"] >= 5
        assert body["pages"] == (body["total"] + 1) // 2

    def test_bug_filter_combinations(self, admin_client):
        p1 = _create_project(admin_client, name="P-A")
        p2 = _create_project(admin_client, name="P-B")
        _create_bug(admin_client, p1["id"], title="bug-one", priority="High", environment="DEV")
        _create_bug(admin_client, p1["id"], title="bug-two", priority="Low", environment="UAT")
        _create_bug(admin_client, p2["id"], title="bug-three", priority="High", environment="PROD")
        r = admin_client.get(f"/api/bugs?project_id={p1['id']}&priority=High")
        items = r.json()["items"]
        assert len(items) == 1 and items[0]["title"] == "bug-one"

    def test_bug_update_invalid_field_value(self, admin_client):
        p = _create_project(admin_client, name="N14")
        bug = _create_bug(admin_client, p["id"])
        r = admin_client.put(f"/api/bugs/{bug['id']}", json={"status": "Garbage"})
        assert r.status_code == 422

    def test_bug_update_to_same_reporter_should_not_fail_for_user(self, admin_client):
        """PUTting a bug with the user's own reporter_id must succeed, not 403."""
        p = _create_project(admin_client, name="N15")
        u = _create_user(admin_client, "Owner", "owner@x.com", password="TestUserPwd9X")
        _logout(admin_client)
        _login_as(admin_client, "owner@x.com", "TestUserPwd9X")
        bug = _create_bug(admin_client, p["id"], title="mine")
        r = admin_client.put(f"/api/bugs/{bug['id']}", json={
            "title": "renamed",
            "reporter_id": u["id"],
        })
        assert r.status_code == 200, r.text

    def test_bug_clear_due_date_with_empty_string(self, admin_client):
        """Frontend sends '' when due date is cleared. Should accept and store None."""
        p = _create_project(admin_client, name="N16")
        bug = _create_bug(admin_client, p["id"], due_date="2025-01-01")
        r = admin_client.put(f"/api/bugs/{bug['id']}", json={"due_date": ""})
        assert r.status_code == 200, r.text
        assert r.json()["due_date"] is None

    def test_user_can_edit_bug_they_are_assignee_of(self, admin_client):
        """Regression: assignees must be able to edit; previously can_edit only checked reporter."""
        p = _create_project(admin_client, name="N17")
        u = _create_user(admin_client, "Helper", "helper@x.com", password="TestUserPwd9X")
        bug = _create_bug(admin_client, p["id"], assignee_ids=[u["id"]])
        _logout(admin_client)
        _login_as(admin_client, "helper@x.com", "TestUserPwd9X")
        r = admin_client.put(f"/api/bugs/{bug['id']}", json={"status": "In Progress"})
        assert r.status_code == 200

    def test_xlsx_export_works(self, admin_client):
        """A freshly-created bug must appear in the Reports XLSX export."""
        import io
        from openpyxl import load_workbook
        p = _create_project(admin_client, name="N18")
        _create_bug(admin_client, p["id"], title="csv-test")
        r = admin_client.post("/api/reports/export.xlsx", json={
            "report_key": "item_detail", "filters": {},
        })
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        text = " ".join(
            str(v) for row in wb[wb.sheetnames[0]].iter_rows(values_only=True)
            for v in row if v is not None
        )
        assert "csv-test" in text


class TestCommentsAttachments:
    def test_comment_on_nonexistent_bug(self, admin_client):
        r = admin_client.post("/api/bugs/999999/comments", json={"body": "hi"})
        assert r.status_code == 404

    def test_comment_with_empty_body_rejected(self, admin_client):
        p = _create_project(admin_client, name="C1")
        bug = _create_bug(admin_client, p["id"])
        r = admin_client.post(f"/api/bugs/{bug['id']}/comments", json={"body": ""})
        assert r.status_code == 422

    def test_comment_with_only_whitespace_rejected(self, admin_client):
        p = _create_project(admin_client, name="C2")
        bug = _create_bug(admin_client, p["id"])
        r = admin_client.post(f"/api/bugs/{bug['id']}/comments", json={"body": "   "})
        assert r.status_code == 422

    def test_attachment_empty_file_rejected(self, admin_client):
        p = _create_project(admin_client, name="C3")
        bug = _create_bug(admin_client, p["id"])
        files = {"file": ("empty.txt", io.BytesIO(b""), "text/plain")}
        r = admin_client.post(f"/api/bugs/{bug['id']}/attachments", files=files)
        assert r.status_code == 400

    def test_attachment_for_wrong_comment_id(self, admin_client):
        p = _create_project(admin_client, name="C4")
        bug1 = _create_bug(admin_client, p["id"])
        bug2 = _create_bug(admin_client, p["id"], title="bug2")
        cr = admin_client.post(f"/api/bugs/{bug2['id']}/comments", json={"body": "x"})
        cid = cr.json()["id"]
        # Attaching to bug1 with bug2's comment_id must 400.
        files = {"file": ("a.txt", io.BytesIO(b"x"), "text/plain")}
        r = admin_client.post(
            f"/api/bugs/{bug1['id']}/attachments",
            files=files, data={"comment_id": str(cid)},
        )
        assert r.status_code == 400

    def test_download_attachment_works(self, admin_client):
        p = _create_project(admin_client, name="C5")
        bug = _create_bug(admin_client, p["id"])
        files = {"file": ("data.bin", io.BytesIO(b"Hello World"), "application/octet-stream")}
        r = admin_client.post(f"/api/bugs/{bug['id']}/attachments", files=files)
        att_id = r.json()["id"]
        d = admin_client.get(f"/api/bugs/{bug['id']}/attachments/{att_id}/download")
        assert d.status_code == 200
        assert d.content == b"Hello World"

    def test_download_attachment_wrong_bug(self, admin_client):
        """Cross-bug access: downloading bug2's attachment via bug1's URL must 404."""
        p = _create_project(admin_client, name="C6")
        bug1 = _create_bug(admin_client, p["id"])
        bug2 = _create_bug(admin_client, p["id"], title="b2")
        files = {"file": ("x.txt", io.BytesIO(b"secret"), "text/plain")}
        r = admin_client.post(f"/api/bugs/{bug2['id']}/attachments", files=files)
        att_id = r.json()["id"]
        d = admin_client.get(f"/api/bugs/{bug1['id']}/attachments/{att_id}/download")
        assert d.status_code == 404

    def test_attachment_filename_with_special_chars(self, admin_client):
        """Filenames with quotes can break the Content-Disposition header."""
        p = _create_project(admin_client, name="C8")
        bug = _create_bug(admin_client, p["id"])
        files = {"file": ('weird"name.txt', io.BytesIO(b"x"), "text/plain")}
        r = admin_client.post(f"/api/bugs/{bug['id']}/attachments", files=files)
        if r.status_code == 201:
            att_id = r.json()["id"]
            d = admin_client.get(f"/api/bugs/{bug['id']}/attachments/{att_id}/download")
            assert d.status_code != 500, "header injection from filename crashed server"

    def test_uploader_cannot_delete_their_own_attachment_v25(self, admin_client):
        """Attachment deletion is admin-only: even the uploader gets 403."""
        p = _create_project(admin_client, name="C9")
        bug = _create_bug(admin_client, p["id"])
        u = _create_user(admin_client, "Up", "up@x.com", password="TestUserPwd9X")
        admin_client.put(f"/api/bugs/{bug['id']}", json={"assignee_ids": [u["id"]]})
        _logout(admin_client)
        _login_as(admin_client, "up@x.com", "TestUserPwd9X")
        files = {"file": ("u.txt", io.BytesIO(b"u"), "text/plain")}
        r = admin_client.post(f"/api/bugs/{bug['id']}/attachments", files=files)
        att_id = r.json()["id"]
        d = admin_client.delete(f"/api/bugs/{bug['id']}/attachments/{att_id}")
        assert d.status_code == 403
        assert "admin" in d.json()["detail"].lower()
        # Admin can delete.
        _logout(admin_client)
        _login_as(admin_client, "admin@test.local", "Admin1234")
        d = admin_client.delete(f"/api/bugs/{bug['id']}/attachments/{att_id}")
        assert d.status_code == 200

    def test_attachment_count_increments(self, admin_client):
        p = _create_project(admin_client, name="C10")
        bug = _create_bug(admin_client, p["id"])
        files = {"file": ("a.txt", io.BytesIO(b"a"), "text/plain")}
        admin_client.post(f"/api/bugs/{bug['id']}/attachments", files=files)
        admin_client.post(f"/api/bugs/{bug['id']}/attachments", files=files)
        # Re-fetch
        r = admin_client.get(f"/api/bugs/{bug['id']}")
        assert r.json()["attachment_count"] == 2


class TestStatsAudit:
    def test_stats_includes_recent_bugs_in_timeline(self, admin_client):
        p = _create_project(admin_client, name="ST1")
        _create_bug(admin_client, p["id"])
        r = admin_client.get("/api/stats")
        body = r.json()
        assert isinstance(body["timeline"], list)
        assert len(body["timeline"]) == 14
        assert sum(d["count"] for d in body["timeline"]) >= 1

    def test_audit_filter_by_entity_type(self, admin_client):
        p = _create_project(admin_client, name="ST2")
        r = admin_client.get("/api/audit?entity_type=project")
        assert r.status_code == 200
        rows = r.json()
        assert all(row["entity_type"] == "project" for row in rows)
        assert any(row["action"] == "project_created" for row in rows)

    def test_audit_search_by_text(self, admin_client):
        p = _create_project(admin_client, name="UniqueAuditMarker_xyz")
        r = admin_client.get("/api/audit?q=UniqueAuditMarker_xyz")
        rows = r.json()
        assert len(rows) >= 1

    def test_bug_delete_creates_audit_record(self, admin_client):
        p = _create_project(admin_client, name="ST3")
        bug = _create_bug(admin_client, p["id"], title="DELETED_MARKER_777")
        admin_client.delete(f"/api/bugs/{bug['id']}")
        # The audit record uses bug_id=None so it survives the cascade delete.
        r = admin_client.get("/api/audit?q=DELETED_MARKER_777")
        rows = r.json()
        assert any(row["action"] == "bug_deleted" for row in rows), \
            "Bug delete didn't create a non-bug audit record"


class TestEdgeCases:
    def test_title_with_only_whitespace_rejected(self, admin_client):
        """A title that is whitespace-only should be rejected as empty."""
        p = _create_project(admin_client, name="E1")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "         ",
            "priority": "Medium", "environment": "DEV",
        })
        assert r.status_code == 422, f"got {r.status_code}: {r.text}"

    def test_title_with_padded_whitespace_below_min_length(self, admin_client):
        """Length must be re-checked after stripping: '  a  ' trims to 1 char and must 422."""
        p = _create_project(admin_client, name="E2")
        r = admin_client.post("/api/bugs", json={
            "project_id": p["id"], "title": "  a  ",   # raw len=5, trimmed=1
            "priority": "Medium", "environment": "DEV",
        })
        assert r.status_code == 422, \
            f"BUG: trim-after-min_length lets 1-char title through. status={r.status_code}, body={r.text}"

    def test_bug_create_with_empty_assignees(self, admin_client):
        p = _create_project(admin_client, name="E3")
        bug = _create_bug(admin_client, p["id"], assignee_ids=[])
        assert bug["assignees"] == []

    def test_bug_update_with_empty_dict(self, admin_client):
        """No-op update must not error."""
        p = _create_project(admin_client, name="E4")
        bug = _create_bug(admin_client, p["id"])
        r = admin_client.put(f"/api/bugs/{bug['id']}", json={})
        assert r.status_code == 200

    def test_unicode_in_title_and_description(self, admin_client):
        p = _create_project(admin_client, name="E5")
        bug = _create_bug(admin_client, p["id"], title="🐛 émoji bug ✨",
                          description="Naïve café résumé — Ω₃")
        r = admin_client.get(f"/api/bugs/{bug['id']}")
        assert "🐛" in r.json()["title"]

    def test_concurrent_audit_log_rows(self, admin_client):
        """Two consecutive actions in same second should both appear in audit."""
        p = _create_project(admin_client, name="E6")
        _create_bug(admin_client, p["id"], title="x1")
        _create_bug(admin_client, p["id"], title="x2")
        r = admin_client.get("/api/audit?entity_type=bug&limit=10")
        rows = r.json()
        assert sum(1 for r in rows if r["action"] == "bug_created") >= 2

    def test_user_can_be_assignee_of_their_own_reported_bug(self, admin_client):
        p = _create_project(admin_client, name="E7")
        u = _create_user(admin_client, "Self", "self@x.com", password="TestUserPwd9X")
        bug = _create_bug(admin_client, p["id"], reporter_id=u["id"],
                          assignee_ids=[u["id"]])
        assert bug["reporter"]["id"] == u["id"]
        assert any(a["id"] == u["id"] for a in bug["assignees"])

    def test_negative_user_id_in_filters(self, admin_client):
        """Filter with bogus user id should return empty list, not 500."""
        r = admin_client.get("/api/bugs?assignee_id=-1")
        assert r.status_code == 200
        assert r.json()["items"] == []

    def test_large_search_query(self, admin_client):
        """Over-length q (cap 200 chars) must 422 cleanly, not crash."""
        r = admin_client.get("/api/bugs?q=" + "x" * 5000)
        assert r.status_code == 422  # bounded by the q max_length guard
        r = admin_client.get("/api/bugs?q=" + "x" * 50)
        assert r.status_code == 200

    def test_bug_status_case_insensitive_filter(self, admin_client):
        """List status filter must be case-insensitive, matching creation normalization."""
        p = _create_project(admin_client, name="E8")
        _create_bug(admin_client, p["id"], status="New")
        r = admin_client.get("/api/bugs?status=New")
        n_canonical = r.json()["total"]
        assert n_canonical >= 1
        r = admin_client.get("/api/bugs?status=new")
        n_lower = r.json()["total"]
        assert n_lower == n_canonical, \
            f"Filter is case-SENSITIVE but creation is case-INSENSITIVE: " \
            f"created with 'new'/found 'New' but filter ?status=new yields {n_lower} vs {n_canonical}"


class TestSecurity:
    def test_login_does_not_leak_user_existence(self, client):
        """'No such user' and 'wrong password' must produce identical responses."""
        r1 = client.post("/api/auth/login", json={
            "email": "doesnotexist@nowhere.test", "password": "Whatever1",
        })
        r2 = client.post("/api/auth/login", json={
            "email": "admin@test.local", "password": "wrongpass",
        })
        assert r1.status_code == 401 and r2.status_code == 401
        assert r1.json()["detail"] == r2.json()["detail"]

    def test_unauth_xlsx_export_blocked(self, client):
        """Unauthenticated XLSX export must 401."""
        r = client.post("/api/reports/export.xlsx", json={
            "report_key": "item_detail", "filters": {},
        })
        assert r.status_code == 401

    def test_attachment_download_requires_auth(self, admin_client):
        """Attachment download must require an active session."""
        p = _create_project(admin_client, name="S1")
        bug = _create_bug(admin_client, p["id"])
        files = {"file": ("a.txt", io.BytesIO(b"secret"), "text/plain")}
        r = admin_client.post(f"/api/bugs/{bug['id']}/attachments", files=files)
        att_id = r.json()["id"]
        admin_client.post("/api/auth/logout")
        d = admin_client.get(f"/api/bugs/{bug['id']}/attachments/{att_id}/download")
        assert d.status_code == 401

    def test_audit_endpoint_is_hidden_from_regular_users(self, user_client):
        """Audit trail must be 403 for regular users."""
        r = user_client.get("/api/audit")
        assert r.status_code == 403

    def test_audit_endpoint_visible_to_admin(self, admin_client):
        r = admin_client.get("/api/audit")
        assert r.status_code == 200

    def test_xss_in_bug_title_is_stored_as_is(self, admin_client):
        """The server stores raw HTML; escaping is the frontend's job. Verify the API doesn't double-escape."""
        p = _create_project(admin_client, name="S2")
        bug = _create_bug(admin_client, p["id"], title="<script>alert(1)</script>")
        r = admin_client.get(f"/api/bugs/{bug['id']}")
        assert r.json()["title"] == "<script>alert(1)</script>"  # stored verbatim


class TestV321Security:
    """CSRF origin checks, upload rate limit (20/60s/user), hardened response headers."""

    # ---- CSRF Origin check ----
    def test_csrf_blocks_state_change_from_foreign_origin(self, admin_client):
        """A POST with a foreign Origin must 403 before any work is done."""
        # Valid project first so the 403 clearly comes from the CSRF check.
        p = _create_project(admin_client, name="CSRF1")
        baseline_count = admin_client.get("/api/bugs").json()["total"]

        r = admin_client.post(
            "/api/bugs",
            json={"project_id": p["id"], "title": "Phishy bug",
                  "priority": "Medium", "environment": "DEV"},
            headers={"Origin": "https://evil.example.com"},
        )
        assert r.status_code == 403, r.text
        assert "cross-origin" in r.json()["detail"].lower()

        # No bug was created.
        after = admin_client.get("/api/bugs").json()["total"]
        assert after == baseline_count

    def test_csrf_allows_state_change_from_same_origin(self, admin_client):
        """When Origin matches the request Host the request goes through normally."""
        p = _create_project(admin_client, name="CSRF2")
        r = admin_client.post(
            "/api/bugs",
            json={"project_id": p["id"], "title": "Real bug",
                  "priority": "Medium", "environment": "DEV"},
            headers={"Origin": "http://testserver"},  # TestClient default Host
        )
        assert r.status_code == 201, r.text

    def test_csrf_allows_request_with_no_origin_and_no_referer(self, admin_client):
        """Non-browser clients (no Origin, no Referer) are not a CSRF vector
        and must be accepted. The TestClient omits both by default, which the
        rest of the test suite already relies on; this test makes that contract
        explicit."""
        p = _create_project(admin_client, name="CSRF3")
        r = admin_client.post(
            "/api/bugs",
            json={"project_id": p["id"], "title": "Curl-style bug",
                  "priority": "Medium", "environment": "DEV"},
        )
        assert r.status_code == 201, r.text

    def test_csrf_login_blocks_cross_origin(self, client):
        """Login is not CSRF-exempt: a cross-site login POST (session-fixation
        vector) is blocked. Operator scripts still work because they send no
        Origin/Referer and pass the no-fingerprint branch."""
        r = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin1234"},
            headers={"Origin": "https://other-tool.example.com"},
        )
        assert r.status_code == 403, r.text
        # No Origin header (operator script / same-origin) still authenticates.
        r2 = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin1234"},
        )
        assert r2.status_code == 200, r2.text

    def test_csrf_does_not_apply_to_safe_methods(self, admin_client):
        """GET and HEAD are read-only; the Origin check must not apply to them."""
        r = admin_client.get(
            "/api/bugs",
            headers={"Origin": "https://evil.example.com"},
        )
        assert r.status_code == 200

    def test_csrf_blocks_state_change_via_referer_only(self, admin_client):
        """If Origin is absent but Referer is present and foreign, also block."""
        p = _create_project(admin_client, name="CSRF4")
        r = admin_client.post(
            "/api/bugs",
            json={"project_id": p["id"], "title": "Referer bug",
                  "priority": "Medium", "environment": "DEV"},
            headers={"Referer": "https://evil.example.com/page"},
        )
        assert r.status_code == 403, r.text

    # ---- Upload rate limit ----
    def test_attachment_upload_rate_limit_fires(self, admin_client):
        """20 uploads/min/user is the cap; the 21st in the window must get 429."""
        p = _create_project(admin_client, name="RL1")
        bug = _create_bug(admin_client, p["id"])

        statuses = []
        # Burst well above the cap to make the 429 deterministic.
        for i in range(25):
            r = admin_client.post(
                f"/api/bugs/{bug['id']}/attachments",
                files={"file": (f"f{i}.txt", io.BytesIO(b"x" * 64), "text/plain")},
            )
            statuses.append(r.status_code)
            if r.status_code == 429:
                # Retry-After must be a positive int per RFC 7231.
                assert int(r.headers.get("Retry-After", "0")) > 0
                break

        assert 429 in statuses, f"expected 429 in burst, got {statuses}"
        # Some 201s must precede the 429. We don't assert exactly 20 because
        # the limiter prunes entries as they age out.
        assert any(s == 201 for s in statuses), statuses

    # ---- Response headers ----
    def test_server_header_is_stripped(self, admin_client):
        """The Server header must be absent; exposing the stack (uvicorn) is minor info disclosure."""
        r = admin_client.get("/api/health")
        assert "server" not in {k.lower() for k in r.headers.keys()}

    def test_cross_origin_resource_policy_header(self, admin_client):
        """CORP=same-origin prevents cross-origin embedding of API responses
        as subresources (e.g. via <img src=...>)."""
        r = admin_client.get("/api/health")
        assert r.headers.get("Cross-Origin-Resource-Policy") == "same-origin"

    def test_csp_still_set(self, admin_client):
        """The Content-Security-Policy header is still set with the expected directives."""
        r = admin_client.get("/api/health")
        csp = r.headers.get("Content-Security-Policy", "")
        assert "default-src 'self'" in csp
        assert "frame-ancestors 'none'" in csp


# Not wall-clock timing (flaky on shared CI) — locks in query-count collapse
# on the dashboard, deferred BLOB loading on Attachment, and correct
# attachment counts in list responses.
class TestV321Performance:
    def test_stats_endpoint_returns_expected_kpis(self, admin_client):
        """Stats was refactored from 5 individual queries to a single GROUP BY.
        The returned numbers must still match what the dashboard expects."""
        p = _create_project(admin_client, name="PERF1")
        # 3 New, 1 Resolved, 1 Closed, 1 Resolve Later, 1 Not a Bug.
        for _ in range(3):
            _create_bug(admin_client, p["id"])
        for s in ("Resolved", "Closed", "Resolve Later", "Not a Bug"):
            b = _create_bug(admin_client, p["id"])
            admin_client.put(f"/api/bugs/{b['id']}", json={"status": s})
        r = admin_client.get("/api/stats").json()
        # "Not a Bug" doesn't count towards the total.
        assert r["bugs"] == 6
        assert r["open"] == 3                  # the 3 News
        assert r["resolved"] == 1
        assert r["closed"] == 1
        assert r["resolve_later"] == 1
        # by_status must include "Not a Bug" even though it's excluded from bugs.
        assert r["by_status"].get("Not a Bug") == 1

    def test_attachment_data_is_deferred_but_download_still_works(self, admin_client):
        """Attachment.data is deferred so list/detail queries skip the BLOB.
        The download endpoint explicitly reads .data and must still return
        the full bytes."""
        p = _create_project(admin_client, name="DEFER1")
        bug = _create_bug(admin_client, p["id"])
        payload = b"deferred-loading-canary-" + b"x" * 4096
        r = admin_client.post(
            f"/api/bugs/{bug['id']}/attachments",
            files={"file": ("blob.bin", io.BytesIO(payload), "application/octet-stream")},
        )
        assert r.status_code == 201, r.text
        att_id = r.json()["id"]
        # Detail must include attachment metadata without blowing up.
        d = admin_client.get(f"/api/bugs/{bug['id']}").json()
        assert any(a["id"] == att_id for a in d["attachments"])
        dl = admin_client.get(f"/api/bugs/{bug['id']}/attachments/{att_id}/download")
        assert dl.status_code == 200
        assert dl.content == payload

    def test_bug_list_remains_n_plus_one_free(self, admin_client):
        """Attachment counts must be batched in a single aggregate, not one
        query per bug. We verify the count column rather than query internals
        (which would be a brittle SQLAlchemy-level assertion)."""
        p = _create_project(admin_client, name="N1")
        bugs = [_create_bug(admin_client, p["id"]) for _ in range(3)]
        # No file on the first bug, one on the second, two on the third.
        for f_id, count in ((bugs[1]["id"], 1), (bugs[2]["id"], 2)):
            for i in range(count):
                admin_client.post(
                    f"/api/bugs/{f_id}/attachments",
                    files={"file": (f"f{i}.txt", io.BytesIO(b"x"), "text/plain")},
                )
        listed = admin_client.get(
            f"/api/bugs?project_id={p['id']}&page_size=100"
        ).json()["items"]
        by_id = {b["id"]: b for b in listed}
        assert by_id[bugs[0]["id"]]["attachment_count"] == 0
        assert by_id[bugs[1]["id"]]["attachment_count"] == 1
        assert by_id[bugs[2]["id"]]["attachment_count"] == 2


class TestCascades:
    def test_delete_user_who_is_reporter_sets_reporter_null(self, admin_client):
        p = _create_project(admin_client, name="CA1")
        u = _create_user(admin_client, "Will", "will@x.com")
        bug = _create_bug(admin_client, p["id"], reporter_id=u["id"])
        admin_client.delete(f"/api/users/{u['id']}")
        r = admin_client.get(f"/api/bugs/{bug['id']}")
        assert r.json()["reporter"] is None

    def test_delete_user_who_is_assignee_removes_assignment(self, admin_client):
        p = _create_project(admin_client, name="CA2")
        u = _create_user(admin_client, "A", "asg@x.com")
        bug = _create_bug(admin_client, p["id"], assignee_ids=[u["id"]])
        admin_client.delete(f"/api/users/{u['id']}")
        r = admin_client.get(f"/api/bugs/{bug['id']}")
        assert r.json()["assignees"] == []

    def test_delete_bug_cascades_comments_and_attachments(self, admin_client):
        p = _create_project(admin_client, name="CA3")
        bug = _create_bug(admin_client, p["id"])
        admin_client.post(f"/api/bugs/{bug['id']}/comments", json={"body": "hi"})
        files = {"file": ("a.txt", io.BytesIO(b"a"), "text/plain")}
        admin_client.post(f"/api/bugs/{bug['id']}/attachments", files=files)
        admin_client.delete(f"/api/bugs/{bug['id']}")
        # Comments and attachments cascade with the parent bug.
        r = admin_client.get(f"/api/bugs/{bug['id']}/comments")
        assert r.status_code == 404


class TestV321Chatbot:
    """NLU behavior tests:
      - "me" / "mine" / "my bugs" pronoun resolution
      - "unassigned" / "no assignee" filter
      - "oldest" / "stale" sort hint
      - "minor" / "trivial" priority synonyms
      - Context-aware unknown-intent fallback
    """

    def _ask(self, c, message):
        r = c.post("/api/chat/ask", json={"message": message})
        assert r.status_code == 200, r.text
        return r.json()

    def test_my_bugs_resolves_to_actor(self, admin_client):
        """'my bugs' must filter by the logged-in user without them typing
        their own name."""
        p = _create_project(admin_client, name="ME1")
        me = admin_client.get("/api/auth/me").json()
        bug = _create_bug(admin_client, p["id"], title="Mine to fix",
                          assignee_ids=[me["id"]])
        _create_bug(admin_client, p["id"], title="Someone else's")

        r = self._ask(admin_client, "show my bugs")
        # Summary text must name the actor, confirming the pronoun resolved.
        text_blocks = [b["payload"].get("text", "")
                       for b in r["blocks"] if b["kind"] == "text"]
        full = " ".join(text_blocks).lower()
        assert me["name"].lower() in full, full
        table = next((b for b in r["blocks"] if b["kind"] == "table"), None)
        assert table is not None
        flat = " ".join(" ".join(row) for row in table["payload"]["rows"])
        assert "Mine to fix" in flat
        assert "Someone else's" not in flat

    def test_bugs_i_reported_uses_reporter_role(self, admin_client):
        """'bugs I reported' must filter by reporter, not assignee."""
        p = _create_project(admin_client, name="ME2")
        _create_bug(admin_client, p["id"], title="Found by admin")
        r = self._ask(admin_client, "bugs I reported")
        table = next((b for b in r["blocks"] if b["kind"] == "table"), None)
        assert table is not None
        flat = " ".join(" ".join(row) for row in table["payload"]["rows"])
        assert "Found by admin" in flat

    def test_unassigned_filter(self, admin_client):
        """'unassigned bugs' / 'bugs with no assignee' must return only
        bugs where the assignees list is empty."""
        p = _create_project(admin_client, name="UN1")
        u = _create_user(admin_client, "Bob", "bob@x.com")
        _create_bug(admin_client, p["id"], title="Loose ticket")          # unassigned
        _create_bug(admin_client, p["id"], title="Has owner",
                    assignee_ids=[u["id"]])
        r = self._ask(admin_client, "show unassigned bugs")
        table = next((b for b in r["blocks"] if b["kind"] == "table"), None)
        assert table is not None
        flat = " ".join(" ".join(row) for row in table["payload"]["rows"])
        assert "Loose ticket" in flat
        assert "Has owner" not in flat

    def test_oldest_sort_hint(self, admin_client):
        """'oldest bugs' must sort ASC by updated_at."""
        p = _create_project(admin_client, name="OLD1")
        # Timestamps have second resolution, so sleep >1s between writes
        # to make updated_at deltas measurable.
        first = _create_bug(admin_client, p["id"], title="Older one")
        time.sleep(1.1)
        second = _create_bug(admin_client, p["id"], title="Newer one")
        time.sleep(1.1)
        # Touch the first bug so its updated_at moves past the second.
        admin_client.put(f"/api/bugs/{first['id']}", json={"priority": "High"})
        # After the bump: "Newer one" (never touched) has the oldest updated_at
        # and should appear first in ASC order.
        r = self._ask(admin_client, "show oldest bugs")
        table = next((b for b in r["blocks"] if b["kind"] == "table"), None)
        assert table is not None
        rows = table["payload"]["rows"]
        assert "Newer one" in rows[0][1], rows
        # Default (no sort hint) should be newest-updated first.
        r2 = self._ask(admin_client, "show all bugs")
        table2 = next((b for b in r2["blocks"] if b["kind"] == "table"), None)
        assert "Older one" in table2["payload"]["rows"][0][1], table2["payload"]["rows"]

    def test_priority_synonyms_minor_and_blocker(self, admin_client):
        """'minor bugs' maps to Low and 'showstopper bugs' maps to Critical."""
        p = _create_project(admin_client, name="SYN1")
        _create_bug(admin_client, p["id"], title="Tiny issue", priority="Low")
        _create_bug(admin_client, p["id"], title="Crashing!", priority="Critical")
        r = self._ask(admin_client, "list minor bugs")
        table = next((b for b in r["blocks"] if b["kind"] == "table"), None)
        assert table is not None
        flat = " ".join(" ".join(row) for row in table["payload"]["rows"])
        assert "Tiny issue" in flat
        assert "Crashing!" not in flat

    def test_unknown_intent_hint_is_context_aware(self, admin_client):
        """A query mentioning 'user' that doesn't match a known intent
        should produce a users-flavoured hint rather than the generic bugs hint."""
        r = self._ask(admin_client, "tell me about user habits")
        text_blocks = [b["payload"].get("text", "")
                       for b in r["blocks"] if b["kind"] == "text"]
        full = " ".join(text_blocks).lower()
        if r["intent"] == "unknown":
            assert ("users" in full) or ("admins" in full), full

    def test_typo_tolerant_priority(self, admin_client):
        """A common typo in 'critical' should still find Critical bugs via fuzzy match."""
        p = _create_project(admin_client, name="TYPO1")
        _create_bug(admin_client, p["id"], title="The crash",
                    priority="Critical")
        _create_bug(admin_client, p["id"], title="Trivial thing",
                    priority="Low")
        r = self._ask(admin_client, "show ctitical bugs")
        table = next((b for b in r["blocks"] if b["kind"] == "table"), None)
        if table is not None:
            flat = " ".join(" ".join(row) for row in table["payload"]["rows"])
            assert "The crash" in flat
            assert "Trivial thing" not in flat

    def test_typo_tolerant_environment(self, admin_client):
        """A typo of 'production' should still resolve to PROD."""
        p = _create_project(admin_client, name="TYPO2")
        _create_bug(admin_client, p["id"], title="In prod", environment="PROD")
        _create_bug(admin_client, p["id"], title="In dev", environment="DEV")
        r = self._ask(admin_client, "bugs in produciton")
        table = next((b for b in r["blocks"] if b["kind"] == "table"), None)
        if table is not None:
            flat = " ".join(" ".join(row) for row in table["payload"]["rows"])
            assert "In prod" in flat, flat

    def test_time_phrase_this_year(self, admin_client):
        """'this year' must parse as a time window without a 500."""
        p = _create_project(admin_client, name="TIME1")
        _create_bug(admin_client, p["id"], title="Yearly")
        r = self._ask(admin_client, "bugs created this year")
        assert r["intent"] in ("list_bugs", "count_bugs"), r
