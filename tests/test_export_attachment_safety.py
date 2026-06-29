"""Security hardening and performance tests.

Covers:
  1. Chatbot Excel export: formula triggers defanged (same policy as CSV/reports).
  2. Attachment downloads: inline rendering is safelist-based; unknown MIME types
     get Content-Disposition: attachment.
  3. Security headers: X-Permitted-Cross-Domain-Policies present alongside COOP/CORP.
  4. Cached HTML serving still replaces all placeholders.
  5. Events list: batched aggregate counts match what per-event queries produce.
"""
from __future__ import annotations


# ---------------------------------------------------------------------------
# 1. Chatbot Excel formula defang
# ---------------------------------------------------------------------------
class TestChatbotExcelDefang:
    def test_defang_helper(self, client):
        from app.chatbot.excel import _defang_formula_text
        assert _defang_formula_text("=cmd|'/c calc'!A1") == "'=cmd|'/c calc'!A1"
        assert _defang_formula_text("+1") == "'+1"
        assert _defang_formula_text("-1") == "'-1"
        assert _defang_formula_text("@x") == "'@x"
        assert _defang_formula_text("\tx") == "'\tx"
        assert _defang_formula_text("normal title") == "normal title"
        assert _defang_formula_text("") == ""

    def test_workbook_cells_defanged(self, client):
        import io
        from openpyxl import load_workbook
        from app.chatbot.excel import _build_workbook
        rows = [{
            "id": 1, "title": "=HYPERLINK(\"http://evil\")", "project": "P",
            "status": "New", "priority": "High", "environment": "DEV",
            "reporter": "@victim", "assignees": "+someone", "due_date": "",
            "created_at": "2026-01-01", "updated_at": "2026-01-01",
        }]
        data = _build_workbook(rows, "test export")
        ws = load_workbook(io.BytesIO(data)).active
        # Row 3 is the first data row (rows 1-2 are the banner and header).
        assert ws.cell(row=3, column=2).value.startswith("'=")
        assert ws.cell(row=3, column=7).value.startswith("'@")
        assert ws.cell(row=3, column=8).value.startswith("'+")


# ---------------------------------------------------------------------------
# 2. Attachment inline-disposition safelist
# ---------------------------------------------------------------------------
class TestAttachmentDispositionSafelist:
    def _make_bug_with_attachment(self, client, filename, content, ctype):
        r = client.post("/api/projects", json={"name": "AttProj", "color": "#000000"})
        assert r.status_code == 201, r.text
        project = r.json()
        r = client.post("/api/bugs", json={
            "title": "att bug", "project_id": project["id"],
            "priority": "Low", "environment": "DEV",
        })
        assert r.status_code == 201, r.text
        bug = r.json()
        r = client.post(
            f"/api/bugs/{bug['id']}/attachments",
            files={"file": (filename, content, ctype)},
        )
        assert r.status_code == 201, r.text
        return bug["id"], r.json()["id"]

    def test_image_still_inline(self, admin_client):
        png = (b"\x89PNG\r\n\x1a\n" + b"\x00" * 24)
        bug_id, att_id = self._make_bug_with_attachment(
            admin_client, "x.png", png, "image/png")
        r = admin_client.get(f"/api/bugs/{bug_id}/attachments/{att_id}/download")
        assert r.headers["content-disposition"].startswith("inline;")

    def test_text_plain_still_inline(self, admin_client):
        bug_id, att_id = self._make_bug_with_attachment(
            admin_client, "log.txt", b"hello", "text/plain")
        r = admin_client.get(f"/api/bugs/{bug_id}/attachments/{att_id}/download")
        assert r.headers["content-disposition"].startswith("inline;")

    def test_unknown_type_forced_to_attachment(self, admin_client):
        # application/x-anything is not on the safelist, so the browser must
        # download it rather than render it inline.
        bug_id, att_id = self._make_bug_with_attachment(
            admin_client, "blob.bin", b"\x00\x01", "application/x-custom")
        r = admin_client.get(f"/api/bugs/{bug_id}/attachments/{att_id}/download")
        assert r.headers["content-disposition"].startswith("attachment;")

    def test_active_type_still_neutralised(self, admin_client):
        bug_id, att_id = self._make_bug_with_attachment(
            admin_client, "evil.svg", b"<svg/>", "image/svg+xml")
        r = admin_client.get(f"/api/bugs/{bug_id}/attachments/{att_id}/download")
        assert r.headers["content-disposition"].startswith("attachment;")
        assert r.headers["content-type"].startswith("application/octet-stream")


# ---------------------------------------------------------------------------
# 3. Security headers
# ---------------------------------------------------------------------------
class TestSecurityHeaders:
    def test_full_header_set_on_api(self, client):
        r = client.get("/api/health")
        assert r.headers.get("cross-origin-opener-policy") == "same-origin"
        assert r.headers.get("cross-origin-resource-policy") == "same-origin"
        assert r.headers.get("x-permitted-cross-domain-policies") == "none"
        assert r.headers.get("x-content-type-options") == "nosniff"
        assert r.headers.get("x-frame-options") == "DENY"
        assert "content-security-policy" in r.headers


# ---------------------------------------------------------------------------
# 4. Cached HTML serving
# ---------------------------------------------------------------------------
class TestCachedHtml:
    def test_login_page_renders_placeholders_repeatedly(self, client):
        # Second request is served from the render cache; both responses must
        # be identical with no unreplaced placeholders.
        r1 = client.get("/login.html")
        r2 = client.get("/login.html")
        assert r1.status_code == r2.status_code == 200
        assert r1.text == r2.text
        assert "__APP_VERSION__" not in r1.text
        assert "__ASSET_VERSION__" not in r1.text


# ---------------------------------------------------------------------------
# 5. Events list batched aggregates
# ---------------------------------------------------------------------------
class TestEventListAggregates:
    def test_counts_match_after_batching(self, admin_client):
        # One event gets two items, the other gets none. The list endpoint
        # (which uses batched aggregates) must report the same counts.
        r = admin_client.post("/api/projects", json={"name": "EvProj", "color": "#000000"})
        project = r.json()
        r = admin_client.post("/api/events", json={"name": "Ev One"})
        assert r.status_code == 201, r.text
        ev1 = r.json()
        r = admin_client.post("/api/events", json={"name": "Ev Two"})
        ev2 = r.json()
        for title in ("item one", "item two"):
            r = admin_client.post("/api/bugs", json={
                "title": title, "project_id": project["id"],
                "priority": "Low", "environment": "DEV",
                "event_id": ev1["id"],
            })
            assert r.status_code == 201, r.text

        r = admin_client.get("/api/events")
        assert r.status_code == 200
        by_id = {e["id"]: e for e in r.json()}
        assert by_id[ev1["id"]]["item_count"] == 2
        assert by_id[ev2["id"]]["item_count"] == 0
        # creator name must also be resolved via the batched lookup
        assert by_id[ev1["id"]]["created_by_name"] == "Test Admin"
