"""Tests for Bug/Requirement/Task work-item types (one bugs table, item_type column).
Meta exposure, CRUD/filtering, 'Bug' default, stats by_type, idempotent init_db, audit nouns.
"""
from __future__ import annotations


def _make_project(client, name="Eng", color="#c9764f"):
    r = client.post("/api/projects", json={"name": name, "color": color})
    assert r.status_code == 201, r.text
    return r.json()


def _make_item(client, project_id, item_type="Bug", **extra):
    body = {
        "title": f"{item_type} smoke",
        "project_id": project_id,
        "item_type": item_type,
        "priority": "Medium",
        "environment": "DEV",
    }
    body.update(extra)
    r = client.post("/api/bugs", json=body)
    assert r.status_code == 201, r.text
    return r.json()


# Meta
def test_meta_exposes_item_types(client):
    body = client.get("/api/meta").json()
    assert "item_types" in body
    assert set(body["item_types"]) == {"Bug", "Requirement", "Task"}


# Create / update
def test_create_each_item_type(admin_client):
    p = _make_project(admin_client)
    for t in ("Bug", "Requirement", "Task"):
        row = _make_item(admin_client, p["id"], item_type=t)
        assert row["item_type"] == t


def test_create_normalizes_case(admin_client):
    """Case-insensitive match like every other enum."""
    p = _make_project(admin_client)
    r = admin_client.post("/api/bugs", json={
        "title": "case-insensitive type",
        "project_id": p["id"],
        "item_type": "task",
    })
    assert r.status_code == 201, r.text
    assert r.json()["item_type"] == "Task"


def test_create_rejects_unknown_type(admin_client):
    p = _make_project(admin_client)
    r = admin_client.post("/api/bugs", json={
        "title": "bad type",
        "project_id": p["id"],
        "item_type": "Epic",
    })
    assert r.status_code == 422, r.text
    assert "item_type" in r.text


def test_default_item_type_is_bug(admin_client):
    """Omitting item_type defaults to 'Bug' — keeps legacy clients working."""
    p = _make_project(admin_client)
    r = admin_client.post("/api/bugs", json={
        "title": "no type specified",
        "project_id": p["id"],
    })
    assert r.status_code == 201, r.text
    assert r.json()["item_type"] == "Bug"


def test_update_item_type(admin_client):
    """Changing the type updates the field and logs an activity row."""
    p = _make_project(admin_client)
    row = _make_item(admin_client, p["id"], item_type="Bug")
    r = admin_client.put(f"/api/bugs/{row['id']}", json={"item_type": "Requirement"})
    assert r.status_code == 200, r.text
    assert r.json()["item_type"] == "Requirement"
    # Activity log captures the change.
    acts = admin_client.get(f"/api/bugs/{row['id']}/activity").json()
    actions = [a["action"] for a in acts]
    assert "item_type_changed" in actions


# Filter / list
def test_filter_by_item_type(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], item_type="Bug", title="A bug")
    _make_item(admin_client, p["id"], item_type="Requirement", title="A req")
    _make_item(admin_client, p["id"], item_type="Task", title="A task")
    rows = admin_client.get("/api/bugs?item_type=Task").json()
    assert rows["total"] == 1
    assert rows["items"][0]["item_type"] == "Task"
    # Multi-value filter.
    rows = admin_client.get("/api/bugs?item_type=Task&item_type=Requirement").json()
    assert rows["total"] == 2
    assert {r["item_type"] for r in rows["items"]} == {"Task", "Requirement"}


def test_filter_by_due_date(admin_client):
    """Exact-day due_date match, used by reporting and the events detail date filter."""
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], item_type="Task", due_date="2026-05-28")
    _make_item(admin_client, p["id"], item_type="Task", due_date="2026-05-29")
    rows = admin_client.get(
        "/api/bugs?item_type=Task&due_date=2026-05-28"
    ).json()
    assert rows["total"] == 1
    assert rows["items"][0]["due_date"] == "2026-05-28"


# Stats
def test_stats_includes_by_type(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], item_type="Bug")
    _make_item(admin_client, p["id"], item_type="Bug")
    _make_item(admin_client, p["id"], item_type="Requirement")
    _make_item(admin_client, p["id"], item_type="Task")
    # An event makes by_type['Event'] non-zero; the SPA reads it for the Events KPI pill.
    admin_client.post("/api/events", json={"name": "kickoff"})
    s = admin_client.get("/api/stats").json()
    assert "by_type" in s
    assert s["by_type"].get("Bug") == 2
    assert s["by_type"].get("Requirement") == 1
    assert s["by_type"].get("Task") == 1
    assert s["by_type"].get("Event") == 1


# XLSX export carries the type column (verified against the Item Detail report).
def test_xlsx_export_includes_type_column(admin_client):
    import io
    from openpyxl import load_workbook
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], item_type="Task")
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail", "filters": {},
    })
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    # Row 1 is the banner; row 2 is the column-header row.
    header_row = [str(v).lower() if v is not None else "" for v in rows[1]]
    assert "type" in header_row
    # The Task value appears somewhere in the body.
    body_text = " ".join(str(v) for r in rows[2:] for v in r if v is not None)
    assert "Task" in body_text


# Email subject / body uses the work-item type
def test_email_subject_uses_item_type(monkeypatch):
    """Email subjects and bodies should reflect the actual item type, not always 'bug'."""
    from app.email_service import (
        BugSnapshot, UserSnapshot,
        notify_assignment, notify_bug_created, notify_bug_updated,
        notify_comment_added,
    )

    sent: list[tuple[str, list[str], str]] = []

    def fake_deliver(subject, to, body):
        sent.append((subject, list(to), body))

    monkeypatch.setattr("app.email_service.deliver", fake_deliver)

    reporter = UserSnapshot(id=1, name="Alice", email="a@x.test")
    assignee = UserSnapshot(id=2, name="Bob",   email="b@x.test")

    def snap(item_type):
        return BugSnapshot(
            id=42, title="do the thing", project_name="P",
            status="New", priority="Medium", environment="DEV",
            description="", reporter=reporter, assignees=(assignee,),
            item_type=item_type,
        )

    # CREATE
    sent.clear()
    notify_bug_created(snap("Task"), actor_user_id=None)
    assert sent, "expected exactly one email for create"
    subj, _, body = sent[-1]
    assert "task" in subj.lower(), subj
    assert "new task" in subj.lower(), subj
    assert "Type:        Task" in body
    # Must not report it as a bug.
    assert "new bug" not in subj.lower()
    assert "a new bug" not in body.lower()

    sent.clear()
    notify_bug_created(snap("Requirement"), actor_user_id=None)
    subj, _, body = sent[-1]
    assert "requirement" in subj.lower()
    assert "Type:        Requirement" in body
    assert "new bug" not in subj.lower()

    sent.clear()
    notify_bug_created(snap("Bug"), actor_user_id=None)
    subj, _, body = sent[-1]
    assert "new bug" in subj.lower(), subj  # back-compat preserved

    # ASSIGNMENT
    sent.clear()
    notify_assignment(snap("Task"), [assignee], actor_name="Alice")
    subj, _, body = sent[-1]
    assert "task" in subj.lower(), subj
    assert "assigned you to a task" in body.lower()
    assert "assigned you to a bug" not in body.lower()

    # UPDATE
    sent.clear()
    notify_bug_updated(snap("Requirement"), [("status", "New", "In Progress")],
                       actor_name="Alice", actor_user_id=None)
    subj, _, body = sent[-1]
    assert "requirement" in subj.lower()
    assert "updated requirement" in body.lower()
    assert "updated bug" not in body.lower()

    # COMMENT
    sent.clear()
    notify_comment_added(snap("Task"), "Alice", 1, "lgtm")
    subj, _, body = sent[-1]
    assert "task" in subj.lower()
    assert "commented on task" in body.lower()
    assert "commented on bug" not in body.lower()


# Audit-log detail strings carry the right noun
def test_audit_log_detail_uses_item_type(admin_client):
    """The create audit row says e.g. 'Task created with status ...', not 'Bug created ...'."""
    p = _make_project(admin_client)
    row = _make_item(admin_client, p["id"], item_type="Task")
    acts = admin_client.get(f"/api/bugs/{row['id']}/activity").json()
    creates = [a for a in acts if a["action"] == "bug_created"]
    assert creates, acts
    assert "Task" in creates[0]["detail"], creates[0]
    assert "Bug created" not in creates[0]["detail"]


def test_audit_log_delete_uses_item_type(admin_client):
    """Deletion of a Requirement should log 'Deleted requirement #N: …'."""
    p = _make_project(admin_client)
    row = _make_item(admin_client, p["id"], item_type="Requirement")
    rid = row["id"]
    r = admin_client.delete(f"/api/bugs/{rid}")
    assert r.status_code == 200, r.text
    # The detached audit row is in the global audit feed.
    audit = admin_client.get("/api/audit").json()
    matching = [a for a in audit if a["action"] == "bug_deleted" and a.get("entity_id") == rid]
    assert matching, audit
    assert "requirement" in matching[0]["detail"].lower()


# Migration safety
def test_init_db_is_idempotent(admin_client):
    """Re-running init_db() on a populated DB is a no-op; protects prod on every redeploy."""
    from app.database import init_db
    # The first call already ran at app startup; a second and third must be safe.
    init_db()
    init_db()
    # Writes must still work afterward.
    p = _make_project(admin_client, name="Idempotent")
    row = _make_item(admin_client, p["id"], item_type="Requirement")
    assert row["item_type"] == "Requirement"


def test_existing_row_without_type_defaults_to_bug(client, tmp_path, monkeypatch):
    """Legacy row without item_type: init_db adds the column DEFAULT 'Bug' so rows stay bugs."""
    # Build a minimal old SQLite DB without item_type, point the engine at it, run init_db().
    import sqlite3
    db_file = tmp_path / "legacy.db"
    con = sqlite3.connect(db_file)
    con.executescript(
        """
        CREATE TABLE projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            description TEXT NOT NULL DEFAULT '',
            color TEXT NOT NULL DEFAULT '#c9764f',
            created_at DATETIME NOT NULL,
            updated_at DATETIME NOT NULL
        );
        CREATE TABLE bugs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL REFERENCES projects(id),
            reporter_id INTEGER,
            title TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'New',
            priority TEXT NOT NULL DEFAULT 'Medium',
            environment TEXT NOT NULL DEFAULT 'DEV',
            due_date TEXT,
            created_at DATETIME NOT NULL,
            updated_at DATETIME NOT NULL
        );
        INSERT INTO projects (name, created_at, updated_at)
            VALUES ('legacy', '2026-01-01', '2026-01-01');
        INSERT INTO bugs (project_id, title, status, priority, environment,
                          created_at, updated_at)
            VALUES (1, 'pre-migration bug', 'New', 'Medium', 'DEV',
                    '2026-01-01', '2026-01-01');
        """
    )
    con.commit()
    con.close()

    # Boot against the legacy DB; init_db() should add item_type DEFAULT 'Bug'.
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_file}")
    monkeypatch.setenv("API_KEY", "")
    monkeypatch.setenv("EMAIL_BACKEND", "disabled")
    monkeypatch.setenv("SESSION_SECRET", "legacy_secret")
    monkeypatch.setenv("BOOTSTRAP_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("BOOTSTRAP_ADMIN_PASSWORD", "Admin1234")
    monkeypatch.setenv("BOOTSTRAP_ADMIN_NAME", "Admin")

    import sys
    for mod in list(sys.modules):
        if mod == "app" or mod.startswith("app."):
            del sys.modules[mod]
    from app.config import get_settings
    get_settings.cache_clear()  # type: ignore[attr-defined]

    from fastapi.testclient import TestClient
    from app.main import app
    with TestClient(app) as c:
        r = c.post("/api/auth/login", json={
            "email": "admin@test.local", "password": "Admin1234",
        })
        assert r.status_code == 200, r.text
        rows = c.get("/api/bugs").json()
        assert rows["total"] == 1
        assert rows["items"][0]["item_type"] == "Bug"
