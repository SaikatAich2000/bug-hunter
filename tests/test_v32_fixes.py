"""Tests for the v3.2 audit-fix pass (High + Medium + Low + Nit).

One+ test per behaviour change. Mirrors the suite's conventions: app modules are
imported INSIDE test bodies (the `client` fixture purges sys.modules per test),
and DB assertions open their own SessionLocal.
"""
from __future__ import annotations

import pytest

BOOTSTRAP_EMAIL = "admin@test.local"
BOOTSTRAP_PASSWORD = "Admin1234"
_PW = "User12345"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _session():
    from app.database import SessionLocal
    return SessionLocal()


def _login(c, email, pw):
    r = c.post("/api/auth/login", json={"email": email, "password": pw})
    assert r.status_code == 200, r.text


def _mk_user(admin_c, email, role="user"):
    r = admin_c.post("/api/users", json={
        "name": email.split("@")[0], "email": email, "role": role, "password": _PW,
    })
    assert r.status_code == 201, r.text
    return r.json()


def _mk_project(admin_c, name):
    return admin_c.post("/api/projects", json={"name": name}).json()


# ---------------------------------------------------------------------------
# M1 — create_bug item_type role gate
# ---------------------------------------------------------------------------
def test_regular_user_cannot_create_task(client):
    _login(client, BOOTSTRAP_EMAIL, BOOTSTRAP_PASSWORD)
    proj = _mk_project(client, "Proj A")
    _mk_user(client, "u1@test.local")
    client.post("/api/auth/logout")
    _login(client, "u1@test.local", _PW)
    # A Bug is fine for a regular user…
    assert client.post("/api/bugs", json={"project_id": proj["id"], "title": "a bug"}).status_code == 201
    # …a Task / Requirement is admin/manager-only.
    for t in ("Task", "Requirement"):
        r = client.post("/api/bugs", json={"project_id": proj["id"], "title": f"a {t}", "item_type": t})
        assert r.status_code == 403, r.text


# ---------------------------------------------------------------------------
# M4 — inactive users can't be newly assigned / set as reporter
# ---------------------------------------------------------------------------
def test_create_bug_rejects_inactive_assignee(admin_client):
    proj = _mk_project(admin_client, "PInactive")
    bob = _mk_user(admin_client, "bob.inactive@test.local")
    admin_client.put(f"/api/users/{bob['id']}", json={"is_active": False})
    r = admin_client.post("/api/bugs", json={
        "project_id": proj["id"], "title": "to inactive", "assignee_ids": [bob["id"]],
    })
    assert r.status_code == 400, r.text
    assert "deactivated" in r.json()["detail"].lower()


# ---------------------------------------------------------------------------
# M2 — comment/attachment writes honour the per-type edit policy
# ---------------------------------------------------------------------------
def test_user_cannot_comment_or_attach_on_task(client):
    _login(client, BOOTSTRAP_EMAIL, BOOTSTRAP_PASSWORD)
    proj = _mk_project(client, "PTask")
    task = client.post("/api/bugs", json={
        "project_id": proj["id"], "title": "a task", "item_type": "Task",
    }).json()
    bug = client.post("/api/bugs", json={"project_id": proj["id"], "title": "a bug"}).json()
    _mk_user(client, "u2@test.local")
    client.post("/api/auth/logout")
    _login(client, "u2@test.local", _PW)
    # Blocked on the Task…
    assert client.post(f"/api/bugs/{task['id']}/comments", json={"body": "hi"}).status_code == 403
    att = client.post(f"/api/bugs/{task['id']}/attachments",
                      files={"file": ("n.txt", b"x", "text/plain")})
    assert att.status_code == 403
    # …allowed on the collaborative Bug.
    assert client.post(f"/api/bugs/{bug['id']}/comments", json={"body": "hi"}).status_code == 201


# ---------------------------------------------------------------------------
# M3 — audit numeric search over int4 range must not 500
# ---------------------------------------------------------------------------
def test_audit_search_overrange_number_no_500(admin_client):
    r = admin_client.get("/api/audit", params={"q": "9" * 25})
    assert r.status_code == 200, r.text


# ---------------------------------------------------------------------------
# M17 — removing an assignee notifies them
# ---------------------------------------------------------------------------
def test_unassign_notifies_removed_user(admin_client):
    proj = _mk_project(admin_client, "PUnassign")
    bob = _mk_user(admin_client, "bob.unassign@test.local")
    bug = admin_client.post("/api/bugs", json={
        "project_id": proj["id"], "title": "owned", "assignee_ids": [bob["id"]],
    }).json()
    r = admin_client.put(f"/api/bugs/{bug['id']}", json={"assignee_ids": []})
    assert r.status_code == 200, r.text
    from sqlalchemy import select
    from app.models import Notification
    db = _session()
    notes = list(db.scalars(select(Notification).where(Notification.user_id == bob["id"])).all())
    assert any("Unassigned" in n.title for n in notes), [n.title for n in notes]


# ---------------------------------------------------------------------------
# M5 — bulk optimistic concurrency
# ---------------------------------------------------------------------------
def test_bulk_version_conflict_is_skipped(admin_client):
    proj = _mk_project(admin_client, "PBulkVer")
    bug = admin_client.post("/api/bugs", json={"project_id": proj["id"], "title": "ver bug"}).json()
    # Stale version → conflict, no change.
    r = admin_client.post("/api/bugs/bulk", json={
        "action": "set_priority", "ids": [bug["id"]], "value": "High",
        "expected_versions": {str(bug["id"]): bug["version"] + 5},
    })
    body = r.json()
    assert body["conflicts"] == 1 and body["updated"] == 0, body
    assert admin_client.get(f"/api/bugs/{bug['id']}").json()["priority"] != "High"
    # Correct version → applies.
    fresh = admin_client.get(f"/api/bugs/{bug['id']}").json()
    r2 = admin_client.post("/api/bugs/bulk", json={
        "action": "set_priority", "ids": [bug["id"]], "value": "High",
        "expected_versions": {str(bug["id"]): fresh["version"]},
    })
    assert r2.json()["updated"] == 1, r2.json()


# ---------------------------------------------------------------------------
# H1 — XLSX Filters Applied sheet defangs free-text filters
# ---------------------------------------------------------------------------
def test_xlsx_filters_sheet_defangs_formula():
    from app.reports.engine import ReportResult
    from app.reports import xlsx
    result = ReportResult(
        report_key="item_detail", report_label="X", columns=[], rows=[],
        filters={"text_search": "=cmd|'/c calc'!A1", "label": "+evil()"},
    )
    data = xlsx.build_workbook_bytes(result)
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Filters Applied"]
    cells = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any(v.startswith("'=cmd") for v in cells), cells
    assert any(v.startswith("'+evil") for v in cells), cells


# ---------------------------------------------------------------------------
# M12 — aging report with a non-open status filter returns empty
# ---------------------------------------------------------------------------
def test_aging_empty_status_intersection(admin_client):
    from app.reports.engine import run_report, Filters
    proj = _mk_project(admin_client, "PAge")
    admin_client.post("/api/bugs", json={"project_id": proj["id"], "title": "open one"})
    db = _session()
    res = run_report("aging", Filters(statuses=["Closed"]), db)
    assert res.rows == []


# ---------------------------------------------------------------------------
# M13 — resolution parse takes the RIGHTMOST status clause
# ---------------------------------------------------------------------------
def test_resolution_regex_last_match():
    from app.reports.engine import _parse_resolution_status
    detail = "#5 'weird status: 'x' → 'y'' — status: 'New' → 'Resolved'"
    assert _parse_resolution_status(detail) == "Resolved"


# ---------------------------------------------------------------------------
# M15 — a reopened (now-open) bug shows no stale resolved info
# ---------------------------------------------------------------------------
def test_reopened_bug_has_no_resolved_info(admin_client):
    from app.reports.engine import _fetch_resolution_info
    proj = _mk_project(admin_client, "PReopen")
    bug = admin_client.post("/api/bugs", json={"project_id": proj["id"], "title": "reopen bug"}).json()
    admin_client.put(f"/api/bugs/{bug['id']}", json={"status": "Resolved"})
    admin_client.put(f"/api/bugs/{bug['id']}", json={"status": "Reopened"})
    db = _session()
    info = _fetch_resolution_info(db, [bug["id"]])
    assert bug["id"] not in info, info


# ---------------------------------------------------------------------------
# M11 — throughput marks itself truncated when the scan hits the cap
# ---------------------------------------------------------------------------
def test_throughput_truncated_flag(admin_client, monkeypatch):
    import app.reports.engine as engine
    from app.reports.engine import run_report, Filters
    monkeypatch.setattr(engine, "_detail_cap", lambda: 1)
    proj = _mk_project(admin_client, "PThru")
    for i in range(2):
        b = admin_client.post("/api/bugs", json={"project_id": proj["id"], "title": f"thru bug {i}"}).json()
        admin_client.put(f"/api/bugs/{b['id']}", json={"status": "Resolved"})
    db = _session()
    res = run_report("throughput", Filters(), db)
    assert res.truncated is True


# ---------------------------------------------------------------------------
# config — _env_int / _env_float are crash-safe and clamped
# ---------------------------------------------------------------------------
def test_env_int_garbage_and_clamp(monkeypatch):
    from app import config
    monkeypatch.setenv("BH_TEST_INT", "not-a-number")
    assert config._env_int("BH_TEST_INT", 7) == 7
    monkeypatch.setenv("BH_TEST_INT", "0")
    assert config._env_int("BH_TEST_INT", 7, minimum=3) == 3
    monkeypatch.setenv("BH_TEST_FLOAT", "xx")
    assert config._env_float("BH_TEST_FLOAT", 1.5) == pytest.approx(1.5)
    monkeypatch.setenv("BH_TEST_FLOAT", "-2")
    assert config._env_float("BH_TEST_FLOAT", 0.5, minimum=0.0) == pytest.approx(0.0)


# ---------------------------------------------------------------------------
# scheduler — "N/step" expands to a series, not a single value
# ---------------------------------------------------------------------------
def test_cron_value_step_expands_series():
    from app.scheduler import CronSchedule
    sch = CronSchedule("5/15 * * * *")
    assert sch.minute == {5, 20, 35, 50}


# ---------------------------------------------------------------------------
# cloud_llm — _extract_json scans past a non-dict brace
# ---------------------------------------------------------------------------
def test_extract_json_skips_non_dict():
    from app.chatbot.cloud_llm import _extract_json
    assert _extract_json('[]  {"mode": "answer"}') == {"mode": "answer"}


# ---------------------------------------------------------------------------
# memory — staged ingest expires after the confirm window
# ---------------------------------------------------------------------------
def test_ingest_confirm_ttl(monkeypatch):
    import types
    from app.chatbot import memory
    store = memory._Store()
    store.stage_ingest(1, {"specs": [{"title": "x"}]})
    assert store.peek_ingest(1) is not None
    real_time = memory.time.time
    future = types.SimpleNamespace(
        time=lambda: real_time() + memory._CONFIRM_TTL_SECONDS + 5
    )
    monkeypatch.setattr(memory, "time", future)
    assert store.take_ingest(1) is None


# ---------------------------------------------------------------------------
# push_service — register survives a unique-token insert race
# ---------------------------------------------------------------------------
def test_push_register_insert_race(admin_client, monkeypatch):
    from sqlalchemy.exc import IntegrityError
    from sqlalchemy import select
    from app import push_service
    from app.models import PushSubscription, User
    db0 = _session()
    admin_id = db0.scalar(select(User.id).where(User.email == BOOTSTRAP_EMAIL))
    # The row a concurrent request committed while we were mid-insert.
    db0.add(PushSubscription(user_id=admin_id, token="race-tok", platform="web"))
    db0.commit()

    db = _session()
    real_scalar, real_flush = db.scalar, db.flush
    n = {"scalar": 0, "flush": 0}

    def fake_scalar(*a, **k):
        n["scalar"] += 1
        # Initial existence check MISSES (the race window); recovery FINDS it.
        return None if n["scalar"] == 1 else real_scalar(*a, **k)

    def fake_flush(*a, **k):
        n["flush"] += 1
        if n["flush"] == 1:
            raise IntegrityError("dup", {}, ValueError("dup token"))
        return real_flush(*a, **k)

    monkeypatch.setattr(db, "scalar", fake_scalar)
    monkeypatch.setattr(db, "flush", fake_flush)
    sub = push_service.register(db, user_id=admin_id, token="race-tok", platform="web")
    assert sub.token == "race-tok" and sub.user_id == admin_id


def test_push_register_race_reraises_when_row_gone(monkeypatch):
    # IntegrityError but the recovery lookup also finds nothing → re-raise.
    from sqlalchemy.exc import IntegrityError
    from app import push_service
    db = _session()

    def boom_flush(*a, **k):
        raise IntegrityError("dup", {}, ValueError("x"))

    monkeypatch.setattr(db, "scalar", lambda *a, **k: None)
    monkeypatch.setattr(db, "flush", boom_flush)
    with pytest.raises(IntegrityError):
        push_service.register(db, user_id=1, token="gone-tok", platform="web")


# ---------------------------------------------------------------------------
# download — a Range request returns the exact sliced bytes
# ---------------------------------------------------------------------------
def test_download_range_slice(admin_client):
    proj = _mk_project(admin_client, "PDl")
    bug = admin_client.post("/api/bugs", json={"project_id": proj["id"], "title": "download bug"}).json()
    body = b"0123456789abcdef"
    att = admin_client.post(f"/api/bugs/{bug['id']}/attachments",
                            files={"file": ("f.bin", body, "application/octet-stream")}).json()
    r = admin_client.get(f"/api/bugs/{bug['id']}/attachments/{att['id']}/download",
                         headers={"Range": "bytes=3-7"})
    assert r.status_code == 206
    assert r.content == body[3:8]
    full = admin_client.get(f"/api/bugs/{bug['id']}/attachments/{att['id']}/download")
    assert full.content == body


# ---------------------------------------------------------------------------
# main — login is no longer CSRF-exempt; error responses carry no-store
# ---------------------------------------------------------------------------
def test_login_not_csrf_exempt(client):
    r = client.post("/api/auth/login",
                    headers={"Origin": "https://evil.example"},
                    json={"email": BOOTSTRAP_EMAIL, "password": BOOTSTRAP_PASSWORD})
    assert r.status_code == 403
    assert r.headers.get("Cache-Control") == "no-store"


def test_sessions_list_sweeps_without_error(admin_client):
    # Exercises the set-based expired-sweep DELETE path.
    assert admin_client.get("/api/sessions").status_code == 200


# ---------------------------------------------------------------------------
# M4 — setting an inactive user as reporter is rejected
# ---------------------------------------------------------------------------
def test_update_reporter_to_inactive_rejected(admin_client):
    proj = _mk_project(admin_client, "PRep")
    carol = _mk_user(admin_client, "carol.rep@test.local")
    admin_client.put(f"/api/users/{carol['id']}", json={"is_active": False})
    bug = admin_client.post("/api/bugs", json={"project_id": proj["id"], "title": "rep bug"}).json()
    r = admin_client.put(f"/api/bugs/{bug['id']}", json={"reporter_id": carol["id"]})
    assert r.status_code == 400, r.text
    assert "deactivated" in r.json()["detail"].lower()


# ---------------------------------------------------------------------------
# database — _add_column_safely logs-and-skips a failing ALTER (savepoint)
# ---------------------------------------------------------------------------
def test_add_column_safely_skips_on_error(tmp_path):
    from sqlalchemy import create_engine
    import app.database as database
    eng = create_engine(f"sqlite:///{tmp_path / 'leg.db'}")
    with eng.begin() as conn:
        # ALTER on a missing table raises inside the SAVEPOINT; it must be
        # swallowed (logged) so one bad column can't abort the whole pass.
        database._add_column_safely(conn, "ALTER TABLE nope ADD COLUMN x INTEGER")
    eng.dispose()  # no exception propagated == pass
