"""Reports feature: auth gates, /types catalog, filters, throughput, XLSX export, Sleuth report intent, removed legacy CSV."""
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone

import pytest

from openpyxl import load_workbook

from tests.conftest import BOOTSTRAP_EMAIL, BOOTSTRAP_PASSWORD


# Test-only credentials; hermetic to the temp SQLite DB the fixtures spin up.
_MANAGER_LOGIN = ("mona@test.local", "Mana123456")
_DEFAULT_USER_SECRET = ("UserPass99",)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------
@pytest.fixture()
def manager_client(client):
    """Return a TestClient logged in as a newly-created manager."""
    email, secret = _MANAGER_LOGIN
    res = client.post("/api/auth/login", json={
        "email": BOOTSTRAP_EMAIL, "password": BOOTSTRAP_PASSWORD,
    })
    assert res.status_code == 200
    res = client.post("/api/users", json={
        "name": "Mona Manager",
        "email": email,
        "role": "manager",
        "password": secret,
    })
    assert res.status_code == 201, res.text
    client.post("/api/auth/logout")
    res = client.post("/api/auth/login", json={
        "email": email, "password": secret,
    })
    assert res.status_code == 200
    return client


def _make_user(client, name, email, role="user", password=_DEFAULT_USER_SECRET[0]):
    r = client.post("/api/users", json={
        "name": name, "email": email, "role": role, "password": password,
    })
    assert r.status_code == 201, r.text
    return r.json()


def _make_project(client, name="ReportProj"):
    r = client.post("/api/projects", json={"name": name, "color": "#000000"})
    assert r.status_code == 201, r.text
    return r.json()


def _make_item(client, project_id, **extra):
    body = {
        "title": "reports smoke item",
        "project_id": project_id,
        "priority": "High",
        "environment": "DEV",
        "item_type": "Bug",
    }
    body.update(extra)
    r = client.post("/api/bugs", json=body)
    assert r.status_code == 201, r.text
    return r.json()


def _change_status(client, bug_id, new_status):
    """PUT a status change; the activity log records the actor and timestamp."""
    r = client.put(f"/api/bugs/{bug_id}", json={"status": new_status})
    assert r.status_code == 200, r.text
    return r.json()


# ---------------------------------------------------------------------------
# Auth gate
# ---------------------------------------------------------------------------
def test_reports_unauthenticated_401(client):
    for path in ("/api/reports/types",):
        r = client.get(path)
        assert r.status_code == 401, f"{path}: expected 401, got {r.status_code}"
    r = client.post("/api/reports/run", json={"report_key": "item_detail"})
    assert r.status_code == 401


def test_reports_regular_user_403(user_client):
    r = user_client.get("/api/reports/types")
    assert r.status_code == 403
    r = user_client.post("/api/reports/run", json={"report_key": "item_detail"})
    assert r.status_code == 403
    r = user_client.post("/api/reports/export.xlsx", json={"report_key": "item_detail"})
    assert r.status_code == 403


def test_reports_admin_can_run(admin_client):
    r = admin_client.get("/api/reports/types")
    assert r.status_code == 200
    catalog = r.json()
    keys = {t["key"] for t in catalog["types"]}
    expected = {
        "item_detail", "throughput", "pending_snapshot",
        "status_distribution", "priority_distribution",
        "project_breakdown", "aging", "timeline", "time_to_resolution",
    }
    assert expected.issubset(keys), f"missing: {expected - keys}"
    # Vocab is exposed so the frontend can populate filter pickers.
    v = catalog["vocab"]
    assert "Bug" in v["item_types"]
    assert "Critical" in v["priorities"]
    assert v["environments"] == ["DEV", "UAT", "PROD"]


def test_reports_manager_can_run(manager_client):
    r = manager_client.get("/api/reports/types")
    assert r.status_code == 200


# ---------------------------------------------------------------------------
# Item Detail Export — universal SQL-query view
# ---------------------------------------------------------------------------
def test_item_detail_returns_every_matching_item(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="alpha")
    _make_item(admin_client, p["id"], title="beta", item_type="Bug")
    _make_item(admin_client, p["id"], title="gamma", item_type="Task")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",
        "filters": {},
    })
    assert r.status_code == 200
    body = r.json()
    assert body["report_key"] == "item_detail"
    assert body["total"] == 3
    titles = {row["title"] for row in body["rows"]}
    assert {"alpha", "beta", "gamma"} == titles
    # Check that the standard columns, including computed resolved/days_open, are present.
    keys = set(body["rows"][0].keys())
    for must in ("id", "item_type", "title", "project", "status",
                 "priority", "environment", "reporter_name",
                 "assignees", "created_at", "updated_at",
                 "resolved_at", "resolved_by", "days_open"):
        assert must in keys, f"missing column {must}"


def test_item_detail_filter_by_item_type(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="bug item one", item_type="Bug")
    _make_item(admin_client, p["id"], title="requirement item one", item_type="Requirement")
    _make_item(admin_client, p["id"], title="task item one", item_type="Task")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",
        "filters": {"item_types": ["Bug"]},
    })
    body = r.json()
    assert body["total"] == 1
    assert body["rows"][0]["item_type"] == "Bug"


def test_item_detail_filter_by_project(admin_client):
    p1 = _make_project(admin_client, name="P-One")
    p2 = _make_project(admin_client, name="P-Two")
    _make_item(admin_client, p1["id"], title="only-in-p1")
    _make_item(admin_client, p2["id"], title="only-in-p2")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",
        "filters": {"project_ids": [p1["id"]]},
    })
    body = r.json()
    assert body["total"] == 1
    assert body["rows"][0]["title"] == "only-in-p1"


def test_item_detail_filter_excludes_not_a_bug_by_default(admin_client):
    p = _make_project(admin_client)
    b = _make_item(admin_client, p["id"], title="nab-target")
    _change_status(admin_client, b["id"], "Not a Bug")
    _make_item(admin_client, p["id"], title="regular")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",
        "filters": {},
    })
    titles = {row["title"] for row in r.json()["rows"]}
    assert titles == {"regular"}
    # Including Not a Bug brings it back.
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",
        "filters": {"include_not_a_bug": True},
    })
    titles = {row["title"] for row in r.json()["rows"]}
    assert titles == {"nab-target", "regular"}


def test_item_detail_text_search(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="payments crash on cart")
    _make_item(admin_client, p["id"], title="navigation glitch")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",
        "filters": {"text_search": "payments"},
    })
    body = r.json()
    assert body["total"] == 1
    assert "payments" in body["rows"][0]["title"]


def test_item_detail_attribute_and_entity_filters(admin_client):
    """Priority, environment, assignee, reporter, event, and date filters each narrow the result set."""
    me = admin_client.get("/api/auth/me").json()
    p = _make_project(admin_client, name="FilterProj")
    ev = admin_client.post("/api/events", json={"name": "FilterEvent"}).json()
    # The target matches every filter; the decoy matches none.
    target = _make_item(
        admin_client, p["id"], title="filter-target",
        priority="Critical", environment="PROD",
        assignee_ids=[me["id"]], event_id=ev["id"],
    )
    _make_item(admin_client, p["id"], title="filter-decoy",
               priority="Low", environment="DEV")
    today = datetime.now(timezone.utc).date().isoformat()
    r = admin_client.post("/api/reports/run", json={
        "report_key": "item_detail",
        "filters": {
            "priorities": ["Critical"],
            "environments": ["PROD"],
            "assignee_ids": [me["id"]],
            "reporter_ids": [me["id"]],
            "event_id": ev["id"],
            "date_from": today,
            "date_to": today,
        },
    })
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total"] == 1
    assert body["rows"][0]["title"] == "filter-target"
    assert body["rows"][0]["id"] == target["id"]


# ---------------------------------------------------------------------------
# Throughput — derived from activity_log
# ---------------------------------------------------------------------------
def test_throughput_counts_resolution_events_per_user(admin_client):
    """Basic throughput check: admin resolves two items, both appear in the summary."""
    p = _make_project(admin_client)
    # Both bugs are reported and resolved by admin (the only user in scope).
    b1 = _make_item(admin_client, p["id"], title="resolvable-1")
    b2 = _make_item(admin_client, p["id"], title="resolvable-2")
    _change_status(admin_client, b1["id"], "Resolved")
    _change_status(admin_client, b2["id"], "Closed")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "throughput",
        "filters": {},
    })
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["total_resolved"] == 2
    admin_row = next(row for row in body["rows"] if row["user_name"] == "Test Admin")
    assert admin_row["resolved"] == 2
    assert admin_row["bugs"] == 2
    assert admin_row["requirements"] == 0


def test_throughput_respects_per_type_solved_map(admin_client):
    """Bug→Resolved, Requirement→Implemented, and Task→Done all count;
    other status transitions (e.g. "In Progress") do not."""
    p = _make_project(admin_client)
    bug = _make_item(admin_client, p["id"], title="bug-x", item_type="Bug")
    req = _make_item(admin_client, p["id"], title="req-x", item_type="Requirement")
    task = _make_item(admin_client, p["id"], title="task-x", item_type="Task")
    other = _make_item(admin_client, p["id"], title="other", item_type="Bug")
    _change_status(admin_client, bug["id"], "Resolved")
    _change_status(admin_client, req["id"], "Implemented")
    _change_status(admin_client, task["id"], "Done")
    _change_status(admin_client, other["id"], "In Progress")  # must not count
    r = admin_client.post("/api/reports/run", json={
        "report_key": "throughput",
        "filters": {},
    })
    body = r.json()
    assert body["summary"]["total_resolved"] == 3
    admin_row = next(row for row in body["rows"] if row["user_name"] == "Test Admin")
    assert admin_row["bugs"] == 1
    assert admin_row["requirements"] == 1
    assert admin_row["tasks"] == 1
    # Per-item drill-down has one entry per resolution event.
    assert body["has_detail"] is True
    assert body["detail_total"] == 3


def test_throughput_filters_date_window(admin_client):
    """date_from/date_to filter on the activity_log timestamp, not created_at."""
    p = _make_project(admin_client)
    b = _make_item(admin_client, p["id"], title="will-be-resolved")
    _change_status(admin_client, b["id"], "Resolved")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "throughput",
        "filters": {"date_from": "2000-01-01", "date_to": "2000-12-31"},
    })
    assert r.json()["summary"]["total_resolved"] == 0
    # Open-ended from today → the just-made resolution is included.
    today = datetime.now(timezone.utc).date().isoformat()
    r = admin_client.post("/api/reports/run", json={
        "report_key": "throughput",
        "filters": {"date_from": today},
    })
    assert r.json()["summary"]["total_resolved"] == 1


# ---------------------------------------------------------------------------
# Pending snapshot — currently-open items
# ---------------------------------------------------------------------------
def test_pending_snapshot_includes_only_open_items(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="still-open-1")
    _make_item(admin_client, p["id"], title="still-open-2")
    closed = _make_item(admin_client, p["id"], title="will-close")
    _change_status(admin_client, closed["id"], "Closed")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "pending_snapshot",
        "filters": {},
    })
    body = r.json()
    titles = {row["title"] for row in body["rows"]}
    assert {"still-open-1", "still-open-2"} == titles


# ---------------------------------------------------------------------------
# Distributions
# ---------------------------------------------------------------------------
def test_status_distribution_groups_by_status(admin_client):
    p = _make_project(admin_client)
    b1 = _make_item(admin_client, p["id"], title="status-1")
    b2 = _make_item(admin_client, p["id"], title="status-2")
    _make_item(admin_client, p["id"], title="status-3")
    _change_status(admin_client, b1["id"], "Resolved")
    _change_status(admin_client, b2["id"], "Resolved")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "status_distribution", "filters": {},
    })
    body = r.json()
    by_status = {row["status"]: row["count"] for row in body["rows"]}
    assert by_status.get("Resolved") == 2
    assert by_status.get("New") == 1


def test_priority_distribution_groups_by_priority(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="hi-1", priority="High")
    _make_item(admin_client, p["id"], title="hi-2", priority="High")
    _make_item(admin_client, p["id"], title="lo-1", priority="Low")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "priority_distribution", "filters": {},
    })
    body = r.json()
    by_priority = {row["priority"]: row["count"] for row in body["rows"]}
    assert by_priority.get("High") == 2
    assert by_priority.get("Low") == 1


# ---------------------------------------------------------------------------
# Aging
# ---------------------------------------------------------------------------
def test_aging_orders_oldest_first(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="open-1")
    _make_item(admin_client, p["id"], title="open-2")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "aging", "filters": {},
    })
    body = r.json()
    assert body["total"] == 2
    for row in body["rows"]:
        assert "age_bucket" in row


# ---------------------------------------------------------------------------
# Timeline — created vs resolved per day
# ---------------------------------------------------------------------------
def test_timeline_counts_created_and_resolved_per_day(admin_client):
    p = _make_project(admin_client)
    a = _make_item(admin_client, p["id"], title="tl-1")
    _make_item(admin_client, p["id"], title="tl-2")
    _change_status(admin_client, a["id"], "Resolved")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "timeline", "filters": {},
    })
    assert r.status_code == 200
    body = r.json()
    assert body["report_key"] == "timeline"
    # Default window is 30 days, so there are 30 daily rows.
    assert body["summary"]["window_days"] == len(body["rows"])
    assert body["summary"]["total_created"] == 2
    assert body["summary"]["total_resolved"] == 1
    assert body["summary"]["net"] == 1  # created - resolved
    for row in body["rows"]:
        assert set(row) >= {"date", "created", "resolved", "delta"}
    today = datetime.now(timezone.utc).date().isoformat()
    today_row = next(row for row in body["rows"] if row["date"] == today)
    assert today_row["created"] == 2
    assert today_row["resolved"] == 1


def test_timeline_respects_explicit_window(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="tl-window")
    # A 3-day window entirely in the past → no items, but still 3 rows.
    r = admin_client.post("/api/reports/run", json={
        "report_key": "timeline",
        "filters": {"date_from": "2000-01-01", "date_to": "2000-01-03"},
    })
    body = r.json()
    assert body["summary"]["window_days"] == 3
    assert len(body["rows"]) == 3
    assert body["summary"]["total_created"] == 0


# ---------------------------------------------------------------------------
# Time to Resolution — hours from creation to resolution + stats
# ---------------------------------------------------------------------------
def test_time_to_resolution_reports_resolved_items(admin_client):
    p = _make_project(admin_client)
    b1 = _make_item(admin_client, p["id"], title="ttr-1")
    b2 = _make_item(admin_client, p["id"], title="ttr-2")
    # The open item must not appear in the result.
    _make_item(admin_client, p["id"], title="ttr-open")
    _change_status(admin_client, b1["id"], "Resolved")
    _change_status(admin_client, b2["id"], "Closed")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "time_to_resolution", "filters": {},
    })
    assert r.status_code == 200
    body = r.json()
    assert body["report_key"] == "time_to_resolution"
    assert body["summary"]["count"] == 2
    assert body["total"] == 2
    titles = {row["title"] for row in body["rows"]}
    assert titles == {"ttr-1", "ttr-2"}
    for k in ("average_hours", "median_hours", "p95_hours",
              "fastest_hours", "slowest_hours"):
        assert k in body["summary"]
    # Resolved just now, so hours should be ~0 but never negative.
    for row in body["rows"]:
        assert row["hours_to_resolve"] >= 0
        assert "days_to_resolve" in row


def test_time_to_resolution_empty_when_nothing_resolved(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="never-resolved")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "time_to_resolution", "filters": {},
    })
    body = r.json()
    assert body["total"] == 0
    assert body["summary"]["count"] == 0
    assert body["summary"]["average_hours"] == 0


# ---------------------------------------------------------------------------
# Project breakdown
# ---------------------------------------------------------------------------
def test_project_breakdown_counts_per_project(admin_client):
    p1 = _make_project(admin_client, name="ProjA")
    p2 = _make_project(admin_client, name="ProjB")
    _make_item(admin_client, p1["id"], title="proj-A-one")
    _make_item(admin_client, p1["id"], title="proj-A-two")
    _make_item(admin_client, p2["id"], title="proj-B-one")
    r = admin_client.post("/api/reports/run", json={
        "report_key": "project_breakdown", "filters": {},
    })
    body = r.json()
    by_proj = {row["project"]: row for row in body["rows"]}
    assert by_proj["ProjA"]["created"] == 2
    assert by_proj["ProjB"]["created"] == 1


# ---------------------------------------------------------------------------
# Unknown report key
# ---------------------------------------------------------------------------
def test_unknown_report_key_returns_400(admin_client):
    r = admin_client.post("/api/reports/run", json={
        "report_key": "not_a_real_report", "filters": {},
    })
    assert r.status_code == 400


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------
def test_xlsx_export_returns_valid_workbook(admin_client):
    p = _make_project(admin_client)
    _make_item(admin_client, p["id"], title="export-this")
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail",
        "filters": {"label": "Sanity Pull"},
    })
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    cd = r.headers.get("content-disposition", "")
    assert "report-item_detail" in cd
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    sheet_names = set(wb.sheetnames)
    assert "Filters Applied" in sheet_names
    # Main sheet: banner + header + at least one data row.
    # Title is the report's label, truncated to Excel's 31-char sheet name limit.
    main_ws = wb[wb.sheetnames[0]]
    rows = list(main_ws.iter_rows(values_only=True))
    assert len(rows) >= 3, "expected banner + header + data row"
    filters_ws = wb["Filters Applied"]
    text_dump = " ".join(
        str(v) for r in filters_ws.iter_rows(values_only=True) for v in r if v is not None
    )
    assert "Sanity Pull" in text_dump


def test_xlsx_export_for_throughput_has_drilldown_sheet(admin_client):
    p = _make_project(admin_client)
    b = _make_item(admin_client, p["id"], title="resolve-me")
    _change_status(admin_client, b["id"], "Resolved")
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "throughput",
        "filters": {},
    })
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    # "Items" drill-down sheet is only written when detail_rows is non-empty.
    assert "Items" in wb.sheetnames


# ---------------------------------------------------------------------------
# Sleuth chatbot integration
# ---------------------------------------------------------------------------
def test_sleuth_report_intent_returns_report_with_file(admin_client):
    p = _make_project(admin_client)
    b = _make_item(admin_client, p["id"], title="for-sleuth")
    _change_status(admin_client, b["id"], "Resolved")
    r = admin_client.post("/api/chat/ask", json={
        "message": "report of who resolved how many bugs",
    })
    assert r.status_code == 200
    body = r.json()
    assert body["intent"] == "report", body
    kinds = [b["kind"] for b in body["blocks"]]
    assert "text" in kinds
    assert "table" in kinds
    # One file block carries the XLSX download token.
    file_blocks = [b for b in body["blocks"] if b["kind"] == "file"]
    assert len(file_blocks) == 1, body
    token = file_blocks[0]["payload"]["download_token"]
    assert token and isinstance(token, str)
    r2 = admin_client.get(f"/api/chat/download/{token}")
    assert r2.status_code == 200
    assert r2.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_sleuth_report_intent_forbidden_for_regular_user(user_client):
    r = user_client.post("/api/chat/ask", json={
        "message": "report of who resolved how many bugs last week",
    })
    # The chat endpoint always returns 200; the refusal is in the body.
    assert r.status_code == 200
    body = r.json()
    assert body["intent"] == "report_forbidden"


# ---------------------------------------------------------------------------
# Legacy CSV endpoint must be gone
# ---------------------------------------------------------------------------
def test_legacy_csv_endpoint_removed(admin_client):
    """Removed /api/bugs/export.csv now 404s (no handler) or 422s (catchall can't parse "export.csv" as int)."""
    r = admin_client.get("/api/bugs/export.csv")
    assert r.status_code in (404, 422), r.text
    r = admin_client.post("/api/reports/export.xlsx", json={
        "report_key": "item_detail", "filters": {},
    })
    assert r.status_code == 200
